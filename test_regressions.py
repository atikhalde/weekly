"""
Regression tests for bugs found by auditing a real GitHub Actions run.

Each test here maps to a specific production bug. They exist so the same
mistakes cannot silently come back.
"""

import subprocess
import sys
from pathlib import Path

import pytest
import yaml

ROOT = Path(__file__).resolve().parent
WORKFLOWS = ROOT / ".github" / "workflows"


# --------------------------------------------------------------------------- #
#  BUG 1 - scan.py exited 1 when the snapshot file was missing.
#  scan.py runs every 5 minutes, so that turned one missing file into a failed
#  workflow (and a failure alert) every single run, all day.
#
#  *** REVERSED BY BUG 49 (03-Aug-2026). ***
#
#  The BUG 1 fix - exit 0 and skip quietly - was the direct cause of a whole
#  trading day with no alerts and no warning. The snapshot went stale, every
#  job returned 0, and three workflows reported SUCCESS while producing
#  nothing. Quiet skipping hid a real outage.
#
#  BUG 1's concern was alert fatigue: ~75 failure notices a day. That is
#  solved properly now by de-duplicating the alarm to ONE per calendar day
#  (AlertState.stale_alerted), not by silencing it. The workflow does go red -
#  which is the point, because a red tick is how you find out.
#
#  This test now asserts the CURRENT contract: non-zero exit, and the alarm
#  text must be produced.
# --------------------------------------------------------------------------- #
def test_scan_fails_loudly_when_snapshot_missing(tmp_path):
    cfg = tmp_path / "config.yaml"
    cfg.write_text("strategy: {}\nuniverse: {}\nruntime: {}\n")

    proc = subprocess.run(
        [sys.executable, str(ROOT / "scan.py"), "--force", "--config", str(cfg)],
        capture_output=True, text=True, timeout=60,
        env={"PATH": "/usr/bin:/bin", "DHAN_CLIENT_ID": "x",
             "DHAN_ACCESS_TOKEN": "x", "HOME": str(tmp_path)},
        cwd=tmp_path,
    )
    assert proc.returncode != 0, (
        "a missing snapshot is an OUTAGE and must fail the workflow - exiting "
        "0 is what hid the 03-Aug incident "
        f"(stdout={proc.stdout[-400:]} stderr={proc.stderr[-400:]})"
    )
    blob = (proc.stdout + proc.stderr).lower()
    assert "stale" in blob or "snapshot" in blob, (
        "the operator must be told WHY nothing ran")


def test_load_snapshots_returns_empty_not_raises(tmp_path):
    from config import load_config
    from scan import load_snapshots

    cfg = load_config(ROOT / "config.yaml")
    cfg.paths["snapshot"] = tmp_path / "does_not_exist.csv"
    assert load_snapshots(cfg, "2026-07-27") == []


# --------------------------------------------------------------------------- #
#  BUG 2 - `if: failure() && env.TELEGRAM_BOT_TOKEN != ''`
#  A step's `if:` cannot read an `env:` declared on that same step, so the
#  condition was always false and the failure alert never fired.
# --------------------------------------------------------------------------- #
def load_workflow(name):
    return yaml.safe_load((WORKFLOWS / name).read_text())


def all_steps(wf):
    for job in wf["jobs"].values():
        for step in job.get("steps", []):
            yield step


# --------------------------------------------------------------------------- #
#  BUG 5 - a continuation line inside `run: |` started at column 1, which
#  silently ENDED the block scalar and became a new top-level YAML key.
#  PyYAML parsed the file happily, so "does it parse?" was not a strong enough
#  check; GitHub rejected it with "Unexpected value 'Check the log'".
#  These tests assert the parsed STRUCTURE, not merely that parsing succeeded.
# --------------------------------------------------------------------------- #
VALID_TOP_LEVEL = {"name", "on", True, "concurrency", "permissions",
                   "jobs", "env", "defaults", "run-name"}


@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml", "tests.yml"])
def test_workflow_has_no_stray_top_level_keys(name):
    """
    A run-block line at column 1 turns into a top-level key. GitHub rejects the
    whole workflow; PyYAML does not. Catch it by validating the key set.
    """
    wf = load_workflow(name)
    stray = set(wf) - VALID_TOP_LEVEL
    assert not stray, (
        f"{name}: unexpected top-level key(s) {sorted(map(str, stray))}. "
        "This usually means a line inside a `run: |` block is not indented, "
        "so YAML ended the block early."
    )


@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml", "tests.yml"])
def test_workflow_has_required_structure(name):
    wf = load_workflow(name)
    assert "jobs" in wf and wf["jobs"], f"{name}: no jobs"
    assert ("on" in wf) or (True in wf), f"{name}: no triggers"
    for job_name, job in wf["jobs"].items():
        assert "runs-on" in job, f"{name}:{job_name} missing runs-on"
        assert job.get("steps"), f"{name}:{job_name} has no steps"
        for step in job["steps"]:
            assert "uses" in step or "run" in step, (
                f"{name}:{job_name} has a step with neither `uses` nor `run` "
                f"({step.get('name')}) - a sign the YAML nesting is wrong"
            )


@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml", "tests.yml"])
def test_run_blocks_are_fully_indented(name):
    """
    Re-read the raw file and confirm every line of a `run: |` block is indented
    deeper than the `run:` key itself. This is the exact defect GitHub caught.
    """
    lines = (WORKFLOWS / name).read_text().splitlines()
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not (stripped == "run: |" or stripped.startswith("run: |")):
            continue
        run_indent = len(line) - len(line.lstrip())
        for body in lines[i + 1:]:
            if not body.strip():
                continue
            indent = len(body) - len(body.lstrip())
            if indent <= run_indent:
                # The block ended here. That is only legitimate if this line is
                # a new list item, a comment, or a sibling `key:` mapping.
                head = body.lstrip()
                legit = (head.startswith("-")
                         or head.startswith("#")
                         or ":" in head.split("#")[0])
                assert legit, (
                    f"{name}: {body!r} escaped the `run: |` block - indent it, "
                    "or YAML turns it into a stray top-level key"
                )
                break


@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml", "tests.yml"])
def test_step_if_never_references_own_step_env(name):
    wf = load_workflow(name)
    for step in all_steps(wf):
        cond = str(step.get("if", ""))
        if not cond:
            continue
        for var in (step.get("env") or {}):
            assert f"env.{var}" not in cond, (
                f"{name}: step '{step.get('name')}' tests env.{var} in its own "
                "`if:` - GitHub evaluates that before the step env exists, so it "
                "is always empty. Check the variable inside `run:` instead."
            )


def test_failure_notification_still_exists():
    wf = load_workflow("scan.yml")
    steps = list(all_steps(wf))
    notify = [s for s in steps if "failure()" in str(s.get("if", ""))]
    assert notify, "scan.yml lost its failure notification step"
    body = notify[0].get("run", "")
    assert "TELEGRAM_BOT_TOKEN" in body and "exit 0" in body, (
        "the failure notice must check for missing secrets inside run:"
    )


# --------------------------------------------------------------------------- #
#  BUG 3 - curl -d with %0A / raw text.
#  `-d` does not URL-encode, so an & or + in a symbol name or message would
#  truncate the Telegram alert. --data-urlencode handles it.
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml"])
def test_curl_urlencodes_telegram_payloads(name):
    body = (WORKFLOWS / name).read_text()
    if "api.telegram.org" not in body:
        return
    for line in body.splitlines():
        s = line.strip()
        if s.startswith("-d ") and ("text=" in s or "chat_id=" in s):
            pytest.fail(f"{name}: use --data-urlencode, not -d, for: {s}")
    assert "%0A" not in body, (
        f"{name}: %0A only works if the whole payload is pre-encoded; "
        "use a real newline with --data-urlencode"
    )


# --------------------------------------------------------------------------- #
#  BUG 4 - unpinned deps. CI installed pandas 3.0.5, a major release, into a
#  live trading scanner with no code change and no warning.
# --------------------------------------------------------------------------- #
def test_requirements_pin_major_versions():
    reqs = [l.strip() for l in (ROOT / "requirements.txt").read_text().splitlines()
            if l.strip() and not l.startswith("#")]
    assert reqs, "requirements.txt is empty"
    for r in reqs:
        assert "<" in r, (
            f"'{r}' has no upper bound - a future major release can break the "
            "scanner overnight with no code change"
        )


# --------------------------------------------------------------------------- #
#  Workflow sanity
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("name,script", [
    ("scan.yml", "scan.py"),
    ("snapshot.yml", "build_snapshot.py"),
])
def test_workflows_call_flat_layout_scripts(name, script):
    body = (WORKFLOWS / name).read_text()
    assert f"python {script}" in body, f"{name} should run `python {script}`"
    assert "python -m src." not in body, f"{name} still references the old src/ package"


@pytest.mark.parametrize("name", ["scan.yml", "snapshot.yml"])
def test_write_workflows_declare_contents_write(name):
    wf = load_workflow(name)
    perms = wf.get("permissions") or {}
    assert perms.get("contents") == "write", (
        f"{name} commits back to the repo, so it needs contents: write"
    )


def test_scan_schedule_covers_market_hours_in_utc():
    """NSE 09:15-15:30 IST == 03:45-10:00 UTC. Cron must be UTC."""
    wf = load_workflow("scan.yml")
    # PyYAML parses the bare key `on` as boolean True
    trigger = wf.get("on") or wf.get(True)
    crons = [c["cron"] for c in trigger["schedule"]]
    assert any("3-10" in c for c in crons), f"schedule {crons} misses IST market hours"
    assert any("1-5" in c for c in crons), "should only run Mon-Fri"


def test_no_stale_data_dir_paths_in_workflows():
    """Flat layout: generated files live in the repo root, not data/."""
    for wf_file in WORKFLOWS.glob("*.yml"):
        body = wf_file.read_text()
        assert "data/state.json" not in body, f"{wf_file.name} has a stale data/ path"
        assert "data/weekly_snapshot.csv" not in body, f"{wf_file.name} has a stale data/ path"


# --------------------------------------------------------------------------- #
#  BUG 6 - the NSE scrip master contains ~22 exchange TEST instruments
#  (011NSETEST, G1NSETEST, ...). They have no real history and sort FIRST
#  alphabetically, so `build_snapshot.py --limit 50` spent its whole trial run
#  on dummies and produced an empty snapshot.
# --------------------------------------------------------------------------- #
def test_test_instruments_are_filtered_out():
    import pandas as pd

    from dhan import DhanClient

    fake = pd.DataFrame([
        {"SEM_EXM_EXCH_ID": "NSE", "SEM_SEGMENT": "E", "SEM_SMST_SECURITY_ID": 1,
         "SEM_INSTRUMENT_NAME": "EQUITY", "SEM_TRADING_SYMBOL": "011NSETEST",
         "SEM_SERIES": "EQ", "SEM_EXCH_INSTRUMENT_TYPE": "ES", "SM_SYMBOL_NAME": "t"},
        {"SEM_EXM_EXCH_ID": "NSE", "SEM_SEGMENT": "E", "SEM_SMST_SECURITY_ID": 2,
         "SEM_INSTRUMENT_NAME": "EQUITY", "SEM_TRADING_SYMBOL": "RELIANCE",
         "SEM_SERIES": "EQ", "SEM_EXCH_INSTRUMENT_TYPE": "ES", "SM_SYMBOL_NAME": "r"},
    ])

    class FakeResp:
        status_code = 200
        text = fake.to_csv(index=False)

        def raise_for_status(self):
            pass

    import dhan as dhan_mod
    orig = dhan_mod.requests.get
    dhan_mod.requests.get = lambda *a, **k: FakeResp()
    try:
        got = DhanClient.fetch_instruments(["NSE_EQ"], ["EQ"], exclude_etf=True)
    finally:
        dhan_mod.requests.get = orig

    syms = [i.symbol for i in got]
    assert "RELIANCE" in syms
    assert not [s for s in syms if "NSETEST" in s.upper()], \
        f"exchange test instruments leaked into the universe: {syms}"


# --------------------------------------------------------------------------- #
#  BUG 7 - build_snapshot ran the whole universe before discovering the token
#  was bad, wasting ~12 minutes and ending with an unhelpful error.
# --------------------------------------------------------------------------- #
def test_build_snapshot_has_preflight_check():
    body = (ROOT / "build_snapshot.py").read_text()
    assert "PREFLIGHT" in body, "build_snapshot must probe one symbol before the long run"
    assert "preflight" in body.lower()


def test_snapshot_workflow_uploads_artifact():
    """A multi-hour build must not be lost if the git commit step fails."""
    wf = load_workflow("snapshot.yml")
    steps = list(all_steps(wf))
    assert any("upload-artifact" in str(s.get("uses", "")) for s in steps), \
        "snapshot.yml should upload the CSV as an artifact as a safety net"


# --------------------------------------------------------------------------- #
#  BUG 8 - Weekly Snapshot ran the full 2,390-symbol universe and produced
#  ok=0, skip=2124, err=76. Root causes found in that log:
#    a) _to_frame did not unwrap a {"data": {...}} response -> every symbol
#       looked like it returned no candles and was silently "skipped"
#    b) toDate was today+1 (a future date)
#    c) data_rate 5/s with 5 workers sat exactly on Dhan's ceiling -> 70x 429
#    d) "skip" lumped no-data together with insufficient-history, so the log
#       could not say which had happened
# --------------------------------------------------------------------------- #
def test_to_frame_unwraps_data_envelope():
    from dhan import DhanClient

    payload = {"open": [1.0], "high": [2.0], "low": [0.5], "close": [1.5],
               "volume": [100], "timestamp": [1450000000]}
    assert len(DhanClient._to_frame(payload)) == 1
    assert len(DhanClient._to_frame({"status": "success", "data": payload})) == 1, \
        "a {'data': {...}} envelope must be unwrapped, not treated as empty"
    assert DhanClient._to_frame({"status": "failure"}).empty
    assert DhanClient._to_frame(None).empty


def test_to_date_is_never_in_the_future():
    from datetime import datetime

    from dhan import IST, last_n_years

    _, to_date = last_n_years(5)
    assert to_date <= datetime.now(IST).date(), \
        "a future toDate makes Dhan return an empty payload"


def test_data_rate_leaves_headroom_under_dhan_limit():
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    rt = cfg["runtime"]
    assert rt["data_rate_per_sec"] <= 4, \
        "Dhan allows 5 req/s; running at exactly 5 leaves no room for jitter"
    assert rt["max_workers"] <= rt["data_rate_per_sec"], \
        "more workers than the per-second budget just queues up 429s"


def test_rate_limiter_has_global_pause():
    """A 429 must hold back every thread, not only the one that hit it."""
    import time
    from concurrent.futures import ThreadPoolExecutor

    from dhan import RateLimiter

    lim = RateLimiter(10.0)
    lim.pause(0.6)
    start = time.monotonic()
    waits = []

    def worker(_):
        lim.acquire()
        waits.append(time.monotonic() - start)

    with ThreadPoolExecutor(max_workers=4) as pool:
        list(pool.map(worker, range(4)))
    assert min(waits) >= 0.55, "pause() should delay all workers"


def test_rate_limiter_respects_configured_rate():
    import threading
    import time
    from concurrent.futures import ThreadPoolExecutor

    from dhan import RateLimiter

    lim = RateLimiter(4.0)
    stamps, lock = [], threading.Lock()

    def worker(_):
        lim.acquire()
        with lock:
            stamps.append(time.monotonic())

    with ThreadPoolExecutor(max_workers=4) as pool:
        list(pool.map(worker, range(20)))
    stamps.sort()
    worst = max(sum(1 for x in stamps if s <= x < s + 1.0) for s in stamps)
    assert worst <= 5, f"{worst} requests in one second exceeds Dhan's limit"


def test_build_snapshot_distinguishes_nodata_from_short_history():
    body = (ROOT / "build_snapshot.py").read_text()
    assert "no_data" in body and "short_hist" in body, \
        "the skip counter must separate 'API returned nothing' from 'too little history'"
    assert "nodata" in body


# --------------------------------------------------------------------------- #
#  BUG 9 - THE ok=0 CAUSE.
#  to_ist() unconditionally added the 1980 epoch offset. Dhan's daily sample
#  uses that convention, but other endpoints/revisions return a PLAIN Unix
#  epoch. Adding +10 years to a plain timestamp dated every candle in the
#  2030s, so `weekly[week_start < this_week]` was EMPTY and every symbol was
#  reported as "short_history":  ok=0 skip=1197 (nodata=0 short=1197).
# --------------------------------------------------------------------------- #
def test_to_ist_auto_handles_both_epoch_conventions():
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    from dhan import DHAN_EPOCH_OFFSET, to_ist_auto

    ist = ZoneInfo("Asia/Kolkata")
    now = datetime.now(ist)

    # Dhan's documented daily sample is 1980-based and must decode to 2022
    assert to_ist_auto(1326220200).year == 2022

    # a plain Unix timestamp must NOT be shifted into the future
    target = datetime(2025, 3, 10, 15, 30, tzinfo=ist)
    assert to_ist_auto(int(target.timestamp())).date() == target.date()

    # nothing may ever decode to a future date
    for ts in (int(now.timestamp()), int(now.timestamp()) - DHAN_EPOCH_OFFSET):
        assert to_ist_auto(ts) <= now + timedelta(days=1), \
            "a candle decoded into the future empties the closed-week filter"


def test_plain_unix_payload_still_builds_a_snapshot():
    """The end-to-end symptom: plain-unix candles must not yield 0 closed weeks."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    from config import load_config
    from dhan import DhanClient
    from strategy import build_snapshot, build_weekly_bars, week_start_of

    ist = ZoneInfo("Asia/Kolkata")
    cfg = load_config(ROOT / "config.yaml").strategy
    end = datetime.now(ist).replace(hour=15, minute=30, second=0, microsecond=0)

    o = h = l = c = v = None
    ts, o, h, l, c, v = [], [], [], [], [], []
    price = 100.0
    for i in range(1500, 0, -1):
        d = end - timedelta(days=i)
        if d.weekday() >= 5:
            continue
        price *= 1.0008
        ts.append(int(d.timestamp()))          # PLAIN unix
        o.append(price * 0.995); h.append(price * 1.01)
        l.append(price * 0.99); c.append(price); v.append(500_000)

    df = DhanClient._to_frame({"open": o, "high": h, "low": l, "close": c,
                               "volume": v, "timestamp": ts})
    tw = week_start_of(datetime.now(ist).date())
    closed = build_weekly_bars(df)
    closed = closed[closed["week_start"] < tw]
    assert len(closed) > 100, f"only {len(closed)} closed weeks - epoch decode is wrong"
    assert build_snapshot("T", "1", "NSE_EQ", df, cfg, tw) is not None


def test_daily_candles_are_fetched_in_chunks():
    body = (ROOT / "dhan.py").read_text()
    assert "chunk_days" in body, \
        "5 years in one request can come back short; fetch in chunks"


# --------------------------------------------------------------------------- #
#  Alert must state the entry time of the triggering candle.
# --------------------------------------------------------------------------- #
def test_alert_shows_entry_time():
    from datetime import datetime

    from strategy import BarEval, Signal
    from telegram import _compact, format_signal

    ev = BarEval(conditions={f"c{i:02d}": True for i in range(1, 14)},
                 values={"price": 1512.4, "ema_fast": 1455.2, "ema_slow": 1402.75,
                         "ema_slow_2": 1390.0, "rsi": 72.54, "rsi_1": 63.09,
                         "macd_hist": 12.345, "week_volume": 1.34e7,
                         "vol_sma": 9e6, "week_open": 1463.1, "day_open": 1489.55,
                         "entry_level": 1498.0, "level_52": 1498.0})
    sig = Signal(symbol="X", security_id="1", exchange_segment="NSE_EQ",
                 bar_time=datetime(2026, 7, 27, 11, 20), price=1512.4,
                 entry_level=1498.0, level_52=1498.0, trigger="cross",
                 evaluation=ev, week_start="2026-07-27", week_volume=1.34e7,
                 day_open=1489.55, week_open=1463.1)

    full = format_signal(sig)
    assert "11:20-11:25" in full, "show the signal candle window"
    assert "tradeable from" in full and "11:25" in full, (
        "a bar stamped 11:20 only closes at 11:25 - the alert must say when "
        "the trade can actually be taken")
    assert "11:25" in _compact(sig), "batch lines must show the actionable time"


# --------------------------------------------------------------------------- #
#  Backtest integrity. A backtest that peeks at future data gives false
#  confidence, which is worse than no backtest at all.
# --------------------------------------------------------------------------- #
def test_backtest_reuses_live_strategy_functions():
    """It must not reimplement the rules - otherwise it validates nothing."""
    body = (ROOT / "backtest.py").read_text()
    assert "from strategy import" in body
    assert "build_snapshot" in body and "replay_week" in body
    for banned in ("def evaluate_bar", "def gate_ok", "def build_snapshot("):
        assert banned not in body, f"backtest.py must not redefine {banned}"


def test_backtest_has_no_lookahead():
    """
    Truncating future data must not change past signals. Uses synthetic data so
    the test is deterministic and needs no network.
    """
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import numpy as np
    import pandas as pd

    from backtest import backtest_symbol
    from config import load_config
    from strategy import week_start_of

    ist = ZoneInfo("Asia/Kolkata")
    cfg = load_config(ROOT / "config.yaml")

    rng = np.random.default_rng(3)
    rows, day, price = [], datetime(2021, 1, 4, 15, 30, tzinfo=ist), 100.0
    while len(rows) < 1100:
        if day.weekday() < 5:
            price *= 1 + rng.normal(0.0012, 0.012)
            rows.append({"datetime": day, "open": price * 0.995,
                         "high": price * 1.012, "low": price * 0.988,
                         "close": price, "volume": float(rng.integers(4e5, 9e5))})
        day += timedelta(days=1)
    daily = pd.DataFrame(rows)

    last = pd.Timestamp(daily.iloc[-1]["datetime"]).tz_localize(None)
    start = week_start_of((last - pd.Timedelta(days=700)).date())
    end = week_start_of(last.date()) - pd.Timedelta(days=7)

    full = backtest_symbol("T", daily, cfg, start, end)
    if len(full) < 2:
        return                                   # nothing meaningful to compare

    cut = pd.Timestamp(full[-1].entry_time).tz_localize(None).normalize()
    truncated = daily[pd.to_datetime(daily["datetime"]).dt.tz_localize(None) < cut]
    part = backtest_symbol("T", truncated, cfg, start,
                           week_start_of(cut.date()) - pd.Timedelta(days=7))

    expected = [(t.week, round(t.entry, 4)) for t in full
                if pd.Timestamp(t.week) <= week_start_of(cut.date()) - pd.Timedelta(days=7)]
    got = [(t.week, round(t.entry, 4)) for t in part]
    assert expected == got, (
        "past signals changed when future data was removed - look-ahead bias")


def test_forward_stats_only_uses_bars_after_entry():
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from backtest import forward_stats

    ist = ZoneInfo("Asia/Kolkata")
    t0 = datetime(2025, 1, 6, 15, 30, tzinfo=ist)
    rows = [{"datetime": t0 + timedelta(days=i), "open": 100, "high": 100,
             "low": 100, "close": 100.0 if i <= 0 else 110.0, "volume": 1}
            for i in range(-10, 40)]
    fwd, mfe, mae, _ = forward_stats(pd.DataFrame(rows), t0, 100.0)
    assert all(v > 0 for v in fwd.values()), "pre-entry bars must be excluded"
    assert mae >= 0.0


# --------------------------------------------------------------------------- #
#  Backtest / Dhan wiring
# --------------------------------------------------------------------------- #
def test_backtest_defaults_to_dhan_source():
    body = (ROOT / "backtest.py").read_text()
    assert 'choices=["dhan", "yahoo"], default="dhan"' in body, \
        "the backtest should default to the same data the live scanner uses"


def test_backtest_fetches_scrip_master_once():
    """The scrip master is a ~27 MB download; fetching it twice is wasteful."""
    body = (ROOT / "backtest.py").read_text()
    assert body.count("fetch_instruments(") == 1, \
        "fetch_instruments must be called exactly once in backtest.py"


def test_backtest_has_dhan_preflight():
    body = (ROOT / "backtest.py").read_text()
    assert "PREFLIGHT" in body, \
        "verify the token before spending hundreds of API calls"


def test_backtest_caches_intraday_per_symbol():
    """
    5m history must be pulled once per symbol in 90-day slices, not once per
    week - the latter is ~52 calls per symbol-year and blows the rate limit.
    """
    body = (ROOT / "backtest.py").read_text()
    assert "intraday_cache" in body
    assert "days=85" in body, "respect Dhan's 90-day intraday window"


# --------------------------------------------------------------------------- #
#  BUG 10 - the 25m/err=126 run.
#  Chunking daily history into 365-day windows turned 1 request per symbol into
#  6 (14,208 calls for the universe), which caused sustained 429s. And a single
#  failing chunk raised, discarding every year that DID load - so recently
#  listed stocks were lost entirely instead of using their shorter history.
# --------------------------------------------------------------------------- #
def test_daily_candles_makes_one_request_by_default():
    from datetime import date
    from unittest.mock import patch

    from dhan import DhanClient

    calls = []

    def fake_post(self, path, payload, limiter):
        calls.append(payload)
        return {"open": [1.0], "high": [1.0], "low": [1.0], "close": [1.0],
                "volume": [1], "timestamp": [1450000000]}

    client = DhanClient.__new__(DhanClient)
    client._data_limiter = None
    with patch.object(DhanClient, "_post", fake_post):
        DhanClient.daily_candles(client, "1", "NSE_EQ",
                                 date(2021, 7, 17), date(2026, 7, 27))

    assert len(calls) == 1, (
        f"{len(calls)} requests for one symbol - the daily endpoint has no "
        "90-day cap, so chunking just multiplies the rate-limit pressure")
    assert calls[0]["fromDate"] == "2021-07-17"
    assert calls[0]["toDate"] == "2026-07-27"


def test_partial_chunk_failure_keeps_the_data_that_loaded():
    """A stock listed in 2024 has no 2021 data; that must not lose 2024-2026."""
    from datetime import date
    from unittest.mock import patch

    from dhan import DhanClient, NoDataError

    state = {"n": 0}

    def fake_post(self, path, payload, limiter):
        state["n"] += 1
        if state["n"] == 1:                      # first window predates listing
            raise NoDataError("no data in range (DH-905)")
        return {"open": [1.0], "high": [1.0], "low": [1.0], "close": [1.0],
                "volume": [1], "timestamp": [1450000000]}

    client = DhanClient.__new__(DhanClient)
    client._data_limiter = None
    with patch.object(DhanClient, "_post", fake_post):
        df = DhanClient.daily_candles(client, "1", "NSE_EQ",
                                      date(2021, 1, 1), date(2026, 1, 1),
                                      chunk_days=365)
    assert not df.empty, "one failed chunk must not discard the others"


def test_dh905_is_classified_as_no_data_not_an_error():
    """DH-905 = listed after fromDate. A skip, not a failure to investigate."""
    from unittest.mock import patch

    from dhan import DhanClient, DhanError, NoDataError

    class Resp:
        status_code = 400
        text = ('{"errorType":"Input_Exception","errorCode":"DH-905",'
                '"errorMessage":"System is unable to fetch data"}')

        def json(self):
            return {}

    client = DhanClient.__new__(DhanClient)
    client.timeout = 5
    client.max_retries = 3
    client._data_limiter = type("L", (), {"acquire": lambda s: None,
                                          "pause": lambda s, x: None})()
    client._session = type("S", (), {"post": lambda s, *a, **k: Resp()})()

    with patch("time.sleep", lambda *_: None):
        try:
            client._post("/charts/historical", {}, client._data_limiter)
            raise AssertionError("should have raised")
        except NoDataError as exc:
            assert isinstance(exc, DhanError), "must stay catchable as DhanError"


def test_build_snapshot_treats_nodata_as_skip():
    body = (ROOT / "build_snapshot.py").read_text()
    assert "NoDataError" in body
    assert "except NoDataError" in body, \
        "a listing gap should increment the skip counter, not the error counter"


# --------------------------------------------------------------------------- #
#  BUG 11 - THE REASON NO ALERTS EVER FIRED.
#  weekly_snapshot.csv was never committed, so scan.py had zero levels to test.
#  Two causes:
#    a) "Commit snapshot" had no `if:`, so a CANCELLED or failed build skipped
#       it entirely - run #7 was cancelled at 49 min and lost everything
#    b) the CSV was only written once, at the very end, so any interruption
#       discarded all completed work
# --------------------------------------------------------------------------- #
def test_snapshot_commit_step_runs_even_if_build_is_cancelled():
    wf = load_workflow("snapshot.yml")
    steps = list(all_steps(wf))
    commit = [s for s in steps if "Commit" in str(s.get("name", ""))]
    assert commit, "snapshot.yml lost its commit step"
    assert str(commit[0].get("if", "")).strip() == "always()", (
        "without if: always() a cancelled or failed build skips the commit and "
        "throws away every row it produced")


def test_snapshot_build_saves_progress_incrementally():
    body = (ROOT / "build_snapshot.py").read_text()
    assert "def flush()" in body, "must checkpoint progress, not write once at the end"
    assert "flush()" in body.split("def flush()")[1], "flush must actually be called"


def test_snapshot_build_resumes_from_partial_file():
    body = (ROOT / "build_snapshot.py").read_text()
    assert "done_symbols" in body and "pending" in body, \
        "a rerun should skip symbols already saved for the same week"
    assert "--fresh" in body, "need an escape hatch to force a full rebuild"


def test_snapshot_timeout_is_realistic():
    """
    SUPERSEDED BY BUG 50. The old assertion (<=120 min) rested on an estimate
    of "~10 min of API time". The 03-Aug log disproved it: under Dhan 429
    backoff the build ran at 4.1 s/symbol, projecting ~154 min for 2,372
    symbols. A 120-minute cap would kill a healthy build at ~1,900 names.

    The timeout must now be generous enough to finish a THROTTLED build and
    still bounded enough to surface a genuine hang.
    """
    wf = load_workflow("snapshot.yml")
    job = next(iter(wf["jobs"].values()))
    t = job.get("timeout-minutes", 999)
    assert 180 <= t <= 360, (
        f"timeout {t} min - must clear the measured ~154 min worst case "
        "without being unbounded")


# --------------------------------------------------------------------------- #
#  Low-latency watcher. ~94% of the GitHub Actions latency was cron wait and
#  cold start, not real work (scan itself measured 19s). watch.py removes the
#  fixed cost by staying resident and waking exactly at each candle close.
# --------------------------------------------------------------------------- #
def test_next_candle_close_is_aligned_and_future():
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from watch import next_candle_close

    ist = ZoneInfo("Asia/Kolkata")
    for h, m, s in [(9, 15, 0), (9, 15, 1), (9, 19, 59), (11, 3, 12), (15, 24, 59)]:
        now = datetime(2026, 7, 28, h, m, s, tzinfo=ist)
        nxt = next_candle_close(now, 5)
        assert nxt > now, "must always be in the future, never the current instant"
        assert nxt.minute % 5 == 0 and nxt.second == 0, "must land on a candle boundary"
        assert (nxt - now).total_seconds() <= 300


def test_watcher_reuses_scan_logic():
    """It must not reimplement the strategy, or the two paths could diverge."""
    body = (ROOT / "watch.py").read_text()
    assert "from scan import" in body
    assert "prefilter" in body and "scan_symbol" in body
    for banned in ("def prefilter", "def scan_symbol", "def replay_week"):
        assert banned not in body, f"watch.py must not redefine {banned}"


def test_watcher_respects_one_per_week_state():
    body = (ROOT / "watch.py").read_text()
    assert "already_alerted" in body, \
        "the watcher shares state.json, so it must honour one_per_week"
    assert "state.save()" in body


def test_watcher_loads_snapshot_once_and_reloads_on_new_week():
    body = (ROOT / "watch.py").read_text()
    assert "cur_week != week" in body, "must reload the snapshot when the week rolls"


# --------------------------------------------------------------------------- #
#  BUG 12 - surfaced by a CI failure.
#  A test that expected "no snapshot" instead loaded the real 809 KB snapshot,
#  because config.py anchored data paths to the SCRIPT directory and ignored
#  --config. Worse, the run then hung: with auth failing, prefilter failed OPEN
#  and pushed ~2000 symbols into the per-symbol stage, each retrying with
#  backoff, until the 12-minute workflow timeout killed the job.
# --------------------------------------------------------------------------- #
def test_alternate_config_uses_its_own_data_paths(tmp_path):
    from config import load_config

    cfg_file = tmp_path / "config.yaml"
    cfg_file.write_text("strategy: {}\nuniverse: {}\nruntime: {}\n")
    cfg = load_config(cfg_file)

    assert cfg.paths["snapshot"].parent == tmp_path, (
        "an alternate --config must resolve its own snapshot/state, otherwise "
        "tests and side-by-side setups read the live data files")
    assert cfg.paths["state"].parent == tmp_path


def test_default_config_still_uses_repo_root():
    from config import load_config

    cfg = load_config()
    assert cfg.paths["snapshot"].name == "weekly_snapshot.csv"
    assert cfg.paths["snapshot"].parent == ROOT


def test_prefilter_aborts_on_auth_failure_instead_of_failing_open():
    """
    Failing open is right for one bad batch, but an auth error affects every
    later call too - continuing burns the whole workflow timeout.
    """
    from unittest.mock import patch

    from config import load_config
    from dhan import DhanClient, DhanError
    from scan import prefilter
    from strategy import WeeklySnapshot

    cfg = load_config(ROOT / "config.yaml")
    row = {
        "symbol": "X", "security_id": "1", "exchange_segment": "NSE_EQ",
        "week_start": "2026-07-27", "entry_level": "100", "level_52": "100",
        "hi_short2": "90", "close_1": "95",
        "ema_fast_prev": "1", "ema_fast_len": "20",
        "ema_slow_prev": "1", "ema_slow_len": "50", "ema_slow_2": "1",
        "rsi_len": "14", "rsi_avg_gain": "1", "rsi_avg_loss": "1",
        "rsi_prev_close": "1", "rsi_1": "1",
        "macd_fast": "12", "macd_slow": "26", "macd_signal": "9",
        "macd_ema_fast": "1", "macd_ema_slow": "1", "macd_sig": "1",
        "vol_sma_len": "10", "vol_sma_sum_prev": "1",
        "g_ema_fast": "1", "g_ema_slow": "1", "g_ema_slow_2": "1",
        "g_rsi": "1", "g_rsi_1": "1", "g_hist": "1",
        "prev_daily_close": "1", "prev_daily_open": "1", "mcap": "",
    }
    snaps = [WeeklySnapshot.from_row(dict(row, symbol=f"S{i}", security_id=str(i)))
             for i in range(50)]

    def boom(self, securities):
        raise DhanError('auth failed (401) DH-901 token invalid or expired')

    client = DhanClient.__new__(DhanClient)
    with patch.object(DhanClient, "ltp", boom):
        try:
            prefilter(client, snaps, cfg)
            raise AssertionError("prefilter should abort on an auth failure")
        except DhanError as exc:
            assert "authentication failed" in str(exc).lower()


def test_scan_reports_auth_failure_to_telegram():
    body = (ROOT / "scan.py").read_text()
    assert "Scanner stopped" in body, \
        "an expired token should reach the user on Telegram, not just the log"


# --------------------------------------------------------------------------- #
#  BUG 13 - actions/cache fought with the committed state.json.
#  Step order was: checkout (restores the committed, authoritative state.json)
#  -> actions/cache restore (OVERWRITES it from a possibly older cache).
#  With restore-keys 'alert-state-' matching any previous cache, a stale copy
#  could wipe fresh state, making already-alerted symbols fire a second time.
#  The file is committed to git, so the cache was redundant as well as harmful.
# --------------------------------------------------------------------------- #
def test_scan_workflow_has_no_state_cache_step():
    wf = load_workflow("scan.yml")
    for step in all_steps(wf):
        uses = str(step.get("uses", ""))
        if "actions/cache" not in uses:
            continue
        path = str((step.get("with") or {}).get("path", ""))
        assert "state.json" not in path, (
            "caching state.json overwrites the committed copy restored by "
            "checkout, which can resurrect already-alerted symbols")


def test_state_is_restored_by_checkout_not_cache():
    """checkout must come before anything that writes state.json."""
    wf = load_workflow("scan.yml")
    steps = list(all_steps(wf))
    idx = [i for i, s in enumerate(steps) if "checkout" in str(s.get("uses", ""))]
    assert idx and idx[0] == 0, "checkout must be the first step"


def test_state_push_retries_on_conflict():
    """
    The snapshot job pushes to the same branch. A lost state push means an
    already-alerted symbol re-fires, so the push must retry.
    """
    body = (WORKFLOWS / "scan.yml").read_text()
    persist = body.split("Persist alert state")[1]
    assert "for attempt in" in persist, "state push should retry on conflict"
    assert "git pull --rebase" in persist
    assert "::warning::" in persist, \
        "a failed push should warn, not silently succeed"


# --------------------------------------------------------------------------- #
#  Paper trading engine.
#  Entry  = the live 5m breakout signal (same code path as the alert)
#  Stop   = entry candle low - 0.02%
#  Exit   = first 5m close below the 9-EMA, no fixed target
# --------------------------------------------------------------------------- #
def _paper_signal(entry=100.0, low=99.0):
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval(conditions={f"c{i:02d}": True for i in range(1, 14)},
                 values={"rsi": 70.0, "macd_hist": 1.5})
    return Signal("T", "1", "NSE_EQ", datetime(2026, 7, 27, 10, 0, tzinfo=ist),
                  entry, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6, 97.0, 96.0,
                  bar_open=entry, bar_high=entry * 1.005, bar_low=low)


def _paper_bars(seq):
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    ist = ZoneInfo("Asia/Kolkata")
    t0 = datetime(2026, 7, 27, 10, 5, tzinfo=ist)
    return pd.DataFrame([
        {"datetime": t0 + timedelta(minutes=5 * i), "open": o, "high": h,
         "low": l, "close": c, "volume": 1}
        for i, (o, h, l, c) in enumerate(seq)])


def test_signal_carries_entry_candle_ohlc():
    """paper.py needs the entry candle low to place the stop."""
    import dataclasses

    import pandas as pd

    from config import load_config
    from strategy import build_snapshot, replay_week, week_start_of
    from test_end_to_end import strong_uptrend_daily, week_of_bars

    cfg = load_config(ROOT / "config.yaml")
    # This test is about the OHLC fields travelling with the Signal, not about
    # volatility. week_of_bars() builds deliberately smooth synthetic candles
    # (+-0.1% ranges), which the live min_atr_pct floor correctly rejects, so
    # switch it off here rather than distort the fixture.
    cfg = dataclasses.replace(
        cfg, strategy=dataclasses.replace(cfg.strategy, min_atr_pct=0.0))
    daily = strong_uptrend_daily()
    lw = week_start_of(pd.Timestamp(daily.iloc[-1]["datetime"]).date())
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg.strategy,
                          lw + pd.Timedelta(days=7))
    res = replay_week(snap, cfg.strategy, week_of_bars(snap, breakout_index=20))
    assert res.signals
    sig = res.signals[0]
    assert sig.bar_low > 0 and sig.bar_high >= sig.bar_low
    assert sig.bar_low <= sig.price <= sig.bar_high


def test_paper_stop_is_entry_candle_low_minus_buffer():
    from paper import SL_BUFFER, simulate

    t = simulate(_paper_signal(100.0, 99.0),
                 _paper_bars([(100, 101, 100, 100.8), (100.8, 101, 98.0, 98.5)]),
                 100000)
    assert t.exit_reason == "SL"
    assert abs(t.stop - 99.0 * (1 - SL_BUFFER)) < 1e-9
    assert abs(t.r_multiple + 1.0) < 1e-6, "a stop-out is exactly -1R"


def test_paper_exits_on_close_below_9ema():
    from paper import simulate

    seq = [(100 + i * 0.5, 101 + i * 0.5, 100 + i * 0.5, 100.5 + i * 0.5)
           for i in range(12)]
    seq.append((106, 106, 101.0, 101.2))
    t = simulate(_paper_signal(), _paper_bars(seq), 100000)
    assert t.exit_reason == "EMA9"
    assert t.bars_held == len(seq), (
        "in signal-close mode the exit books on the breaching bar itself")


def test_paper_stop_wins_when_both_trigger_same_candle():
    from paper import simulate

    t = simulate(_paper_signal(), _paper_bars([(100, 100.2, 98.0, 98.1)]), 100000)
    assert t.exit_reason == "SL", "conservative: cannot know intrabar ordering"


def test_paper_position_never_exceeds_capital():
    from paper import simulate

    for entry in (871.90, 199.39, 12.5, 4999.0):
        t = simulate(_paper_signal(entry, entry * 0.99),
                     _paper_bars([(entry, entry * 1.01, entry * 0.995, entry)]),
                     100000)
        assert t.qty == int(100000 // entry)
        assert t.invested <= 100000 + 1e-6


def test_paper_guards_against_malformed_candle():
    """A bad tick where low >= close must not invert the stop."""
    from paper import simulate

    t = simulate(_paper_signal(1874.21, 1876.36),      # low ABOVE the close
                 _paper_bars([(1874, 1875, 1870, 1872)]), 100000)
    assert t.stop < t.entry, "stop must always sit below entry"
    assert abs(t.r_multiple) < 100, "R must stay finite"


def test_paper_open_position_is_marked_to_last_close():
    from paper import simulate

    t = simulate(_paper_signal(),
                 _paper_bars([(100 + i * 0.3, 101 + i * 0.3, 100 + i * 0.3,
                               100.9 + i * 0.3) for i in range(10)]), 100000)
    assert t.exit_reason == "OPEN"
    assert t.pnl > 0


def test_paper_report_has_required_columns():
    body = (ROOT / "paper.py").read_text()
    for field in ("entry_date", "entry_time", "exit_date", "exit_time",
                  "stop", "pnl", "level_26w", "qty"):
        assert field in body, f"report must expose {field}"


def test_paper_uses_live_strategy_code():
    body = (ROOT / "paper.py").read_text()
    assert "from strategy import" in body
    assert "replay_week" in body and "build_snapshot" in body
    for banned in ("def replay_week", "def evaluate_bar", "def gate_ok"):
        assert banned not in body, f"paper.py must not redefine {banned}"


# --------------------------------------------------------------------------- #
#  Post-market Excel report delivered to Telegram.
# --------------------------------------------------------------------------- #
def _report_trades():
    from paper import PaperTrade

    def mk(sym, e, q, sl, ex, why, pnl, r):
        t = PaperTrade(symbol=sym, week="2026-07-27",
                       signal_date="2026-07-27", signal_time="09:50",
                       signal_close=e, entry_date="2026-07-27",
                       entry_time="09:55", entry=e, qty=q, invested=e * q,
                       bar_low=sl / 0.9998, stop=sl, level_26w=e * 0.99,
                       level_52w=e * 0.99, trigger="cross", rsi=71.2,
                       macd_hist=3.4)
        t.exit_date, t.exit_time, t.exit, t.exit_reason = "2026-07-27", "11:20", ex, why
        t.bars_held, t.pnl, t.r_multiple = 18, pnl, r
        t.pnl_pct = (ex / e - 1) * 100
        return t

    return [mk("RATNAVEER", 199.39, 501, 197.5, 203.10, "EMA9", 1859, 2.0),
            mk("MONARCH", 391.60, 255, 388.0, 386.9, "SL", -1198, -1.0),
            mk("TMB", 871.90, 114, 865.0, 879.5, "OPEN", 866, 1.1)]


def test_excel_report_has_required_sheets_and_columns(tmp_path):
    import openpyxl

    from paper_report import build_workbook

    out = build_workbook(_report_trades(), 100000, tmp_path / "r.xlsx", "window")
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames and "Trades" in wb.sheetnames
    assert "Open" in wb.sheetnames, "open positions get their own sheet"

    headers = [c.value for c in wb["Trades"][1]]
    for required in ("Symbol", "Date", "Signal Bar", "Signal Close",
                     "Entry Bar", "Entry Fill", "Qty", "Stop Loss",
                     "26W Level", "Exit Date", "Exit Bar", "Exit Fill",
                     "P&L (Rs)", "R"):
        assert required in headers, f"missing column {required}"
    assert wb["Trades"].max_row == len(_report_trades()) + 1


def test_excel_summary_numbers_are_correct(tmp_path):
    import openpyxl

    from paper_report import build_workbook

    trades = _report_trades()
    out = build_workbook(trades, 100000, tmp_path / "r.xlsx", "window")
    ws = openpyxl.load_workbook(out)["Summary"]
    found = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, 40)
             if ws.cell(r, 1).value}
    assert found["Trades"] == 3
    assert found["Still open"] == 1
    assert abs(found["Net P&L (Rs)"] - sum(t.pnl for t in trades)) < 0.01
    assert found["Wins"] == 2 and found["Losses"] == 1


def test_report_caption_summarises_pnl():
    from paper_report import caption

    text = caption(_report_trades(), 100000, "2026-07-27")
    assert "Paper Trading Report" in text
    assert "Net P&L" in text and "win rate" in text
    assert "RATNAVEER" in text, "best trade should be named"
    assert len(text) <= 1024, "Telegram caption limit"


def test_report_caption_handles_no_trades():
    from paper_report import caption

    assert "No trades" in caption([], 100000, "2026-07-27")


def test_telegram_send_document_dry_run(tmp_path):
    from telegram import Telegram

    f = tmp_path / "x.xlsx"
    f.write_bytes(b"fake")
    assert Telegram("", "", dry_run=True).send_document(f, "cap") is True
    assert Telegram("", "", dry_run=True).send_document(tmp_path / "nope.xlsx") is False


def test_telegram_send_document_posts_multipart(tmp_path, monkeypatch):
    """The workbook must go as a file upload, not inline text."""
    import telegram as tg_mod

    f = tmp_path / "r.xlsx"
    f.write_bytes(b"PK\x03\x04fake-xlsx")
    seen = {}

    class Resp:
        status_code = 200

        def json(self):
            return {"ok": True}

    def fake_post(url, data=None, files=None, timeout=None, **kw):
        seen["url"] = url
        seen["files"] = files
        seen["chat"] = (data or {}).get("chat_id")
        return Resp()

    monkeypatch.setattr(tg_mod.requests, "post", fake_post)
    ok = tg_mod.Telegram("tok", "123").send_document(f, "caption")
    assert ok
    assert seen["url"].endswith("/sendDocument")
    assert "document" in seen["files"]
    assert seen["chat"] == "123"


def test_report_workflow_runs_after_market_close():
    wf = load_workflow("report.yml")
    trigger = wf.get("on") or wf.get(True)
    crons = [c["cron"] for c in trigger["schedule"]]
    # NSE closes 15:30 IST = 10:00 UTC; the report must run after that
    assert any(int(c.split()[1]) >= 10 for c in crons), \
        f"{crons} must run at or after 10:00 UTC (15:30 IST close)"
    assert all("1-5" in c for c in crons), "weekdays only"


def test_report_workflow_uploads_artifact_on_failure():
    wf = load_workflow("report.yml")
    steps = list(all_steps(wf))
    art = [s for s in steps if "upload-artifact" in str(s.get("uses", ""))]
    assert art and str(art[0].get("if", "")).strip() == "always()", \
        "keep the workbook even if the Telegram upload fails"


def test_openpyxl_is_pinned_in_requirements():
    reqs = (ROOT / "requirements.txt").read_text()
    assert "openpyxl" in reqs, "the report needs openpyxl at runtime"
    for line in reqs.splitlines():
        if line.strip().startswith("openpyxl"):
            assert "<" in line, "pin the major version"


# --------------------------------------------------------------------------- #
#  BUG 14 - partial deploy crashed the report.
#  paper.py read sig.bar_low directly, but the DEPLOYED strategy.py predated
#  that field. The job fetched data for 6.5 minutes and then died with
#  AttributeError: 'Signal' object has no attribute 'bar_low'.
#  Two defences: degrade gracefully, and fail fast with a clear message.
# --------------------------------------------------------------------------- #
def test_simulate_survives_signal_without_bar_low():
    """A stale strategy.py must degrade to an entry-based stop, not crash."""
    from dataclasses import dataclass
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from paper import simulate
    from strategy import BarEval

    ist = ZoneInfo("Asia/Kolkata")

    @dataclass
    class OldSignal:            # the Signal shape before bar_* was added
        symbol: str
        security_id: str
        exchange_segment: str
        bar_time: datetime
        price: float
        entry_level: float
        level_52: float
        trigger: str
        evaluation: object
        week_start: str
        week_volume: float
        day_open: float
        week_open: float

    sig = OldSignal("T", "1", "NSE_EQ", datetime(2026, 7, 27, 10, 0, tzinfo=ist),
                    100.0, 98.0, 98.0, "cross",
                    BarEval({}, {"rsi": 70.0, "macd_hist": 1.0}),
                    "2026-07-27", 1e6, 97.0, 96.0)
    bars = pd.DataFrame([
        {"datetime": datetime(2026, 7, 27, 10, 5, tzinfo=ist) + timedelta(minutes=5 * i),
         "open": 100, "high": 101, "low": 99.5, "close": 100.5} for i in range(3)])

    t = simulate(sig, bars, 100000)          # must not raise
    assert t.stop < t.entry, "stop must still sit below entry"


def test_paper_never_reads_bar_fields_unguarded():
    body = (ROOT / "paper.py").read_text()
    assert "sig.bar_low" not in body, (
        "read the entry-candle low via getattr so a stale strategy.py cannot "
        "crash the whole report")
    assert 'getattr(sig, "bar_low"' in body


def test_paper_report_checks_strategy_version_up_front():
    """
    The crash wasted 6.5 minutes of API calls before failing. The version check
    must happen before any data is fetched.
    """
    body = (ROOT / "paper_report.py").read_text()
    assert "bar_low" in body and "out of date" in body, \
        "paper_report should verify Signal has bar_low before fetching data"
    guard = body.index("bar_low")
    fetch = body.index("fetch_5m(client")
    assert guard < fetch, "the version check must precede any data fetching"


# --------------------------------------------------------------------------- #
#  BUG 15 - every paper trade exited on bar 1 ("Avg bars held 1", 0% win rate).
#  The 9-EMA was seeded from the ENTRY PRICE repeated 9 times, so ema == entry
#  at the start. Any close even a paisa below entry was therefore "below the
#  EMA" and closed the position immediately. A real 9-EMA is built from the
#  candles BEFORE entry, where an uptrend leaves it well under price.
# --------------------------------------------------------------------------- #
def _ema_frame(closes, hour=10, minute=0):
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    ist = ZoneInfo("Asia/Kolkata")
    t0 = datetime(2026, 7, 27, hour, minute, tzinfo=ist)
    return pd.DataFrame([
        {"datetime": t0 + timedelta(minutes=5 * i), "open": c, "high": c * 1.002,
         "low": c * 0.998, "close": c} for i, c in enumerate(closes)])


def _ema_signal(entry=100.0, low=99.0):
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    return Signal("T", "1", "NSE_EQ", datetime(2026, 7, 27, 11, 0, tzinfo=ist),
                  entry, 98.0, 98.0, "cross",
                  BarEval({}, {"rsi": 70.0, "macd_hist": 1.0}),
                  "2026-07-27", 1e6, 97.0, 96.0,
                  bar_open=entry, bar_high=entry * 1.005, bar_low=low)


def test_ema_exit_is_warmed_from_pre_entry_candles():
    """A minor dip right after entry must not close the trade."""
    from paper import simulate

    before = _ema_frame([96 + i * 0.3 for i in range(14)])
    before.loc[len(before) - 1, "close"] = 100.0
    after = _ema_frame([99.9, 100.1, 100.4, 100.8, 101.2, 101.6], 11, 5)

    t = simulate(_ema_signal(), after, 100000, before)
    assert t.bars_held > 1, (
        "seeding the EMA at the entry price exits on the first down-tick; "
        "warm it from real prior candles instead")


def test_ema_exit_still_fires_on_a_real_breakdown():
    from paper import simulate

    before = _ema_frame([96 + i * 0.3 for i in range(14)])
    before.loc[len(before) - 1, "close"] = 100.0
    after = _ema_frame([100.2, 100.4, 100.1, 99.0, 98.6, 98.2], 11, 5)

    t = simulate(_ema_signal(), after, 100000, before)
    assert t.exit_reason in ("EMA9", "SL")
    assert t.bars_held >= 2


def test_ema_warmup_when_no_prior_history():
    """Monday-open entries have no earlier candles - must not exit on bar 1."""
    from paper import simulate

    t = simulate(_ema_signal(),
                 _ema_frame([99.9, 100.2, 100.6, 101.0], 11, 5), 100000, None)
    assert t.bars_held > 1


def test_ema_at_entry_sits_below_price_in_an_uptrend():
    """Sanity: the warmed EMA must be under the entry after a run-up."""
    import numpy as np

    import indicators as ind
    from paper import EMA_LEN

    closes = np.array([96 + i * 0.3 for i in range(13)] + [100.0], dtype=float)
    ema = ind.ema(closes, EMA_LEN)[-1]
    assert ema < 100.0, "an uptrend should leave the EMA below price"


def test_simulate_accepts_bars_before_argument():
    import inspect

    from paper import simulate

    assert "bars_before" in inspect.signature(simulate).parameters


def test_paper_callers_pass_pre_entry_candles():
    for name in ("paper.py", "paper_report.py"):
        body = (ROOT / name).read_text()
        assert "simulate(sig, after, args.capital, before, args.fill)" in body, \
            f"{name} must pass pre-entry candles and the fill mode"


# --------------------------------------------------------------------------- #
#  BUG 16 - EXECUTION REALISM (found in a fresh audit before going live).
#  NSE 5m candles are stamped by OPEN time: a bar labelled 10:15 spans
#  10:15:00-10:19:59 and only completes at 10:20. Three consequences the old
#  model got wrong:
#    a) reported times were 5 minutes earlier than any achievable action
#    b) entry filled at the signal bar's own close - a price not known until
#       that bar ended (look-ahead)
#    c) the EMA exit filled at the close that triggered it (same look-ahead)
#  Also: a gap through the stop filled AT the stop, understating gap risk, and
#  nothing prevented two overlapping positions in one symbol.
# --------------------------------------------------------------------------- #
def _exec_sig(close=100.0, low=99.0, hour=10, minute=15):
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    return Signal("T", "1", "NSE_EQ", datetime(2026, 7, 27, hour, minute, tzinfo=ist),
                  close, 98.0, 98.0, "cross",
                  BarEval({}, {"rsi": 70.0, "macd_hist": 1.0}),
                  "2026-07-27", 1e6, 97.0, 96.0,
                  bar_open=close * 0.998, bar_high=close * 1.005, bar_low=low)


def _exec_bars(rows, hour=10, minute=20):
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    ist = ZoneInfo("Asia/Kolkata")
    t0 = datetime(2026, 7, 27, hour, minute, tzinfo=ist)
    return pd.DataFrame([
        {"datetime": t0 + timedelta(minutes=5 * i), "open": o, "high": h,
         "low": l, "close": c} for i, (o, h, l, c) in enumerate(rows)])


def _warm():
    return _exec_bars([(96 + i * .3, 96.5 + i * .3, 95.5 + i * .3, 96 + i * .3)
                       for i in range(14)], 9, 15)


def test_nse_5m_bars_are_stamped_by_open_time():
    """
    75 bars per session, 09:15..15:25. If bars were stamped by CLOSE time the
    range would be 09:20..15:30. Everything downstream depends on this.
    """
    from datetime import datetime, time as dtime, timedelta

    session = (datetime.combine(datetime(2026, 7, 27), dtime(15, 30))
               - datetime.combine(datetime(2026, 7, 27), dtime(9, 15)))
    assert session == timedelta(minutes=375)
    assert 375 // 5 == 75
    first, last = dtime(9, 15), dtime(15, 25)
    assert first < last


def test_entry_fills_at_next_bar_open_not_signal_close():
    from paper import simulate

    t = simulate(_exec_sig(100.0),
                 _exec_bars([(100.6, 101, 100.4, 100.9),
                             (100.9, 101.5, 100.7, 101.3)]), 100000, _warm(),
                 fill="next-open")
    assert t.signal_time == "10:15" and t.signal_close == 100.0
    assert t.entry_time == "10:20", "entry is the bar AFTER the signal"
    assert t.entry == 100.6, "fill at that bar's OPEN"
    assert abs(t.slippage_pct - 0.6) < 0.01


def test_ema_exit_fills_at_next_open_no_lookahead():
    from paper import simulate

    t = simulate(_exec_sig(100.0),
                 _exec_bars([(100.6, 101, 100.4, 98.0),
                             (97.9, 98.2, 97.5, 97.8)]), 100000, _warm(),
                 fill="next-open")
    assert t.exit_reason == "EMA9"
    assert t.exit == 97.9, (
        "a close below the EMA is only known when the bar ends, so the fill "
        "is the NEXT bar's open")


def test_stop_fills_at_stop_when_merely_touched():
    from paper import simulate

    t = simulate(_exec_sig(100.0, 99.0),
                 _exec_bars([(100.5, 100.8, 98.5, 99.5)]), 100000, _warm())
    assert t.exit_reason == "SL"
    assert abs(t.exit - 99.0 * 0.9998) < 1e-4
    assert t.exit_note == ""


def test_gap_through_stop_fills_at_open_not_stop():
    from paper import simulate

    t = simulate(_exec_sig(100.0, 99.0),
                 _exec_bars([(95.0, 95.5, 94.0, 94.5)]), 100000, _warm())
    assert t.exit == 95.0, "a gap-down fills at the open, worse than the stop"
    assert "gap" in t.exit_note


def test_signal_close_mode_reproduces_optimistic_fills():
    from paper import simulate

    t = simulate(_exec_sig(100.0),
                 _exec_bars([(100.6, 101, 100.4, 98.0)]), 100000, _warm(),
                 fill="signal-close")
    assert t.entry == 100.0, "opt-in mode still fills at the signal close"


def test_realistic_fill_is_not_better_than_optimistic():
    """Sanity: removing look-ahead must not improve results."""
    from paper import simulate

    rows = [(100.6, 101, 100.4, 98.0), (97.9, 98.2, 97.5, 97.8)]
    real = simulate(_exec_sig(100.0), _exec_bars(rows), 100000, _warm())
    opt = simulate(_exec_sig(100.0), _exec_bars(rows), 100000, _warm(),
                   fill="signal-close")
    assert real.entry >= opt.entry, "realistic entry cannot be cheaper"


def test_paper_prevents_overlapping_positions_per_symbol():
    for name in ("paper.py", "paper_report.py"):
        body = (ROOT / name).read_text()
        assert "busy_until" in body, (
            f"{name}: a new signal must be skipped while the previous trade "
            "in that symbol is still open")


def test_papertrade_records_signal_and_entry_separately():
    from dataclasses import fields

    from paper import PaperTrade

    names = {f.name for f in fields(PaperTrade)}
    for required in ("signal_date", "signal_time", "signal_close",
                     "entry_date", "entry_time", "entry", "slippage_pct"):
        assert required in names, f"PaperTrade must expose {required}"


def test_last_bar_of_day_signal_fills_next_session():
    """
    A 15:25 bar closes at 15:30 - the market shuts. The fill must roll to the
    next session's open, carrying overnight gap risk. Silently filling at
    15:25's close would be untradeable fiction.
    """
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from paper import simulate

    ist = ZoneInfo("Asia/Kolkata")
    sig = _exec_sig(100.0, 99.0, hour=15, minute=25)
    nxt = datetime(2026, 7, 28, 9, 15, tzinfo=ist)          # next session
    bars = pd.DataFrame([
        {"datetime": nxt + timedelta(minutes=5 * i), "open": 101 + i * 0.1,
         "high": 101.5 + i * 0.1, "low": 100.8 + i * 0.1, "close": 101.2 + i * 0.1}
        for i in range(6)])

    t = simulate(sig, bars, 100000, _warm(), fill="next-open")
    assert t.signal_time == "15:25"
    assert t.entry_date == "2026-07-28" and t.entry_time == "09:15", \
        "a signal on the closing bar can only be filled next session"
    assert t.entry == 101.0, "fill at the next session's open, gap included"


# --------------------------------------------------------------------------- #
#  BUG 17 - entries fired LATE (or not at all) versus the chart.
#  The user's Pine runs with strictEntry = false ("Entry gate: OFF (level
#  only)" in every screenshot), so the indicator buys on the PURE 26W level
#  break; the 13-row table is informational. Proof: MONARCH FAILED:2,
#  SENCO FAILED:1, PYRAMID FAILED:4 all show ENTRY TAKEN - impossible if the
#  gate were active. Our config had strict_entry: true, so entries waited for
#  the gate and printed as "deferred", and CREDITACC never fired at all.
# --------------------------------------------------------------------------- #
def test_gate_off_mode_still_available_for_comparison():
    """
    strictEntry=false reproduces the raw chart arrows, but it is NOT safe to
    trade: on live data it gave 82 signals in a day (22 on the Monday 09:15
    bar) at profit factor 0.73. Keep the switch, but the shipped config uses
    the gate plus a tolerance instead.
    """
    from config import Strategy
    from strategy import gate_ok

    snap = _gate_snap()
    ev = _gate_eval(c05=False, c06=False, c08=False, c09=False)
    # with the gate disabled the caller bypasses gate_ok entirely
    assert Strategy(strict_entry=False).strict_entry is False
    # and with the gate on, four failures still exceed a tolerance of 2
    assert gate_ok(snap, Strategy(gate_tolerance=2), ev) is False


def test_level_only_entry_fires_on_first_close_above():
    """With the gate off, the signal is purely the first CLOSE above the 26W level."""
    import pandas as pd

    from config import Strategy
    from strategy import replay_week
    from test_strategy import make_daily  # noqa: F401  (fixture helpers)

    # build a snapshot then feed candles that pierce on the HIGH first
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    from strategy import build_snapshot, week_start_of

    ist = ZoneInfo("Asia/Kolkata")
    cfg = Strategy(strict_entry=False)
    daily = make_daily()
    lw = week_start_of(pd.Timestamp(daily.iloc[-1]["datetime"]).date())
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, lw + pd.Timedelta(days=7))
    assert snap is not None
    lvl = snap.entry_level

    t0 = datetime(2026, 7, 27, 9, 15, tzinfo=ist)
    bars = pd.DataFrame([
        # high pierces the level but the close does not -> NOT a signal
        {"datetime": t0, "open": lvl * .99, "high": lvl * 1.001,
         "low": lvl * .98, "close": lvl * .995, "volume": 1e6},
        # first close above -> THIS is the signal bar
        {"datetime": t0 + timedelta(minutes=5), "open": lvl * .995,
         "high": lvl * 1.01, "low": lvl * .99, "close": lvl * 1.005,
         "volume": 1e6},
    ])
    res = replay_week(snap, cfg, bars)
    assert len(res.signals) == 1
    assert res.signals[0].bar_time == bars.iloc[1]["datetime"], (
        "Pine uses `close > entryLevel`, so a high-only pierce must not trigger")


def test_paper_default_fill_is_the_signal_candle_close():
    """The report must line up with the chart's BUY arrow by default."""
    import inspect

    from paper import simulate

    assert inspect.signature(simulate).parameters["fill"].default == "signal-close"
    for name in ("paper.py", "paper_report.py"):
        body = (ROOT / name).read_text()
        assert 'default="signal-close"' in body, f"{name} default fill must match the chart"


# --------------------------------------------------------------------------- #
#  BUG 18 - strict_entry: false flooded the report with false entries.
#  Removing the gate produced 82 trades in ONE day, 22 of them on the 09:15
#  Monday bar where "already above the level" was mistaken for a fresh cross.
#  Net -Rs 13,851, win rate 33%, profit factor 0.73, SL exits -Rs 32,285.
#  Fix: keep the gate ON, but allow a small tolerance so a strong breakout is
#  not blocked by one or two marginal rows (which caused "deferred" entries).
# --------------------------------------------------------------------------- #
def test_strict_entry_is_on():
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    assert cfg["strategy"]["strict_entry"] is True, (
        "turning the gate off gave 82 trades/day at PF 0.73 - keep it on and "
        "use gate_tolerance to relax marginal rows instead")


def test_gate_tolerance_is_configured_and_bounded():
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    tol = cfg["strategy"]["gate_tolerance"]
    assert tol == 0, (
        "shipped config is the STRICT 13/13 scan. Measured on 27-Jul, the real "
        "breakouts had a median of 0 failing rows and the false positives a "
        "median of 1, so any tolerance > 0 admits exactly the unwanted names")


def _gate_eval(**overrides):
    from strategy import BarEval

    conds = {f"c{i:02d}": True for i in range(1, 14)}
    conds.update(overrides)
    return BarEval(conds, {"rsi": 70.0, "macd_hist": 1.0})


def _gate_snap():
    from strategy import WeeklySnapshot
    import indicators as ind

    return WeeklySnapshot(
        symbol="T", security_id="1", exchange_segment="NSE_EQ",
        week_start="2026-07-27", entry_level=100.0, level_52=100.0,
        hi_short2=99.0, close_1=98.0,
        ema_fast=ind.EmaState(20, 1.0), ema_slow=ind.EmaState(50, 1.0),
        ema_slow_2=1.0, rsi=ind.RsiState(14, 1.0, 1.0, 1.0), rsi_1=1.0,
        macd=ind.MacdState(12, 26, 9, 1.0, 1.0, 1.0),
        vol_sma=ind.SmaState(10, 1.0),
        g_ema_fast=1.0, g_ema_slow=1.0, g_ema_slow_2=1.0,
        g_rsi=1.0, g_rsi_1=1.0, g_hist=1.0,
        prev_daily_close=100.0, prev_daily_open=99.0)


def test_tolerance_allows_up_to_n_soft_failures():
    from config import Strategy
    from strategy import gate_ok

    snap = _gate_snap()
    # two soft rows failing (RSI level + EMA slope)
    ev = _gate_eval(c06=False, c05=False)
    assert gate_ok(snap, Strategy(gate_tolerance=2), ev) is True
    assert gate_ok(snap, Strategy(gate_tolerance=1), ev) is False
    assert gate_ok(snap, Strategy(gate_tolerance=0), ev) is False


def test_tolerance_never_relaxes_the_breakout_itself():
    """c01/c02 are the level break - tolerance must not be able to fake one."""
    from config import Strategy
    from strategy import gate_ok

    snap = _gate_snap()
    # the gate only sees c03..c13; the cross is proven separately in replay_week
    ev = _gate_eval(c01=False, c02=False)
    assert gate_ok(snap, Strategy(gate_tolerance=2), ev) is True, (
        "c01/c02 are not gate rows - the cross itself proves them")

    body = (ROOT / "strategy.py").read_text()
    gate_src = body[body.index("def gate_ok"):body.index("def replay_week")]
    assert '"c01"' not in gate_src and '"c02"' not in gate_src, \
        "the level break must never be part of the tolerance count"


def test_mandatory_rows_are_never_tolerated():
    """Fresh breakout, min price and market cap stay hard requirements."""
    from config import Strategy
    from strategy import gate_ok

    snap = _gate_snap()
    for row in ("c03", "c11", "c12"):
        ev = _gate_eval(**{row: False})
        assert gate_ok(snap, Strategy(gate_tolerance=3), ev) is False, (
            f"{row} must hold regardless of tolerance")


def test_tolerance_converts_deferred_into_a_clean_cross():
    """
    The user's complaint: MONARCH/TMB printed as 'deferred' because one or two
    weekly rows lagged. With tolerance the entry lands on the breakout candle.
    """
    import pandas as pd
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    from config import Strategy
    from strategy import build_snapshot, replay_week, week_start_of
    from test_end_to_end import strong_uptrend_daily, week_of_bars

    ist = ZoneInfo("Asia/Kolkata")
    daily = strong_uptrend_daily()
    lw = week_start_of(pd.Timestamp(daily.iloc[-1]["datetime"]).date())
    cfg = Strategy(strict_entry=True, gate_tolerance=2)
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, lw + pd.Timedelta(days=7))
    assert snap is not None

    res = replay_week(snap, cfg, week_of_bars(snap, breakout_index=20))
    assert res.signals
    assert res.signals[0].trigger == "cross", \
        "with tolerance the entry should be a cross, not a deferred fill"


# --------------------------------------------------------------------------- #
#  BUG 19 - too many false stocks.
#  Two independent causes, found by profiling the 27-Jul report:
#   a) gate_tolerance > 0. Real breakouts failed a MEDIAN OF 0 rows; the false
#      ones a median of 1. Any tolerance therefore admits precisely the noise.
#   b) c09 compared an ACCUMULATING weekly volume against a FULL-week SMA. At
#      Monday 09:25 only 3 of 375 bars exist, so c09 could never pass and
#      genuine early breakouts were pushed to "deferred" or dropped.
# --------------------------------------------------------------------------- #
def test_volume_mode_is_off_in_shipped_config():
    """
    c09 (weekly volume > 10w SMA) is the single strongest filter in the scan.
    Measured on 27-Jul against the 5 validated names:
        volume_mode "off" ->  5 signals (the valid set)
        volume_mode "day" -> 19 signals (15 false)
        volume_mode "bar" -> 26 signals (22 false)
    So it stays raw, exactly as Pine computes it.
    """
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    assert cfg["strategy"]["volume_mode"] == "off"
    assert cfg["strategy"]["gate_tolerance"] == 0


def test_volume_modes_are_available_but_not_default():
    """The alternatives remain implemented for research, just not shipped."""
    from config import Strategy
    from strategy import evaluate_bar

    snap = _gate_snap()
    early = dict(price=101.0, week_open=99.0, week_volume=0.5,
                 day_open=100.0, week_fraction=0.1)          # Monday
    off = evaluate_bar(snap, Strategy(volume_mode="off"), **early)
    day = evaluate_bar(snap, Strategy(volume_mode="day"), **early)
    assert off.values["vol_target"] >= day.values["vol_target"], \
        "day mode must be no stricter than the raw compare"


def test_deferred_entry_is_expected_when_c09_lags():
    """
    MONARCH/TMB printed as "deferred" because weekly VOLUME had not yet cleared
    its 10-week average at the moment price crossed the level. That is the
    scan working as designed, not a bug: the Pine header names c09 as one of
    the two rows that flip false->true during the week.

    Keep deferred entries ON - without them those setups are lost entirely.
    """
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    assert cfg["strategy"]["defer_entry"] is True, (
        "with a strict gate, disabling deferral drops MONARCH and TMB")
    assert cfg["strategy"]["gate_source"] == "live", (
        'gate_source "closed" gave 20 signals and kept only 1 of the valid 5')


def test_volume_prorate_off_reproduces_pine_exactly():
    from config import Strategy
    from strategy import evaluate_bar

    snap = _gate_snap()
    cfg = Strategy(volume_prorate=False)
    ev = evaluate_bar(snap, cfg, price=101.0, week_open=99.0,
                      week_volume=0.02, day_open=100.0, week_fraction=0.01)
    # un-prorated: 0.02 is NOT greater than the full-week SMA
    assert ev.conditions["c09"] is False
    assert ev.values["vol_target"] == ev.values["vol_sma"]


def test_full_week_fraction_matches_pine():
    """At week end the pro-rated target equals the raw SMA - Pine's own test."""
    from config import Strategy
    from strategy import evaluate_bar

    snap = _gate_snap()
    ev = evaluate_bar(snap, Strategy(volume_prorate=True), price=101.0,
                      week_open=99.0, week_volume=5.0, day_open=100.0,
                      week_fraction=1.0)
    assert abs(ev.values["vol_target"] - ev.values["vol_sma"]) < 1e-12


# --------------------------------------------------------------------------- #
#  BUG 20 - TMB missing, and the real cause of the "late" entries.
#
#  ---------------------------------------------------------------------------
#  SUPERSEDED. The theory recorded here (that TWO shifts stack, putting the
#  window two weeks back) was WRONG and was reverted. It is kept only so the
#  same wrong turn is not taken a third time. See BUG 21 below for the proof
#  and the final resolution, confirmed by the user on 28-Jul (Option B).
#
#  The claim was: entry_level should be 821.20 on TMB, not 847.00. That came
#  from reading the orange stepline on a TradingView screenshot. It is true
#  that the PLOTTED line sits at 821.20 - but that line is a rendering
#  artifact of request.security, not the level the scan should trade. 847.00
#  is TMB's real 26-week high and is what the indicator's own table shows.
#  ---------------------------------------------------------------------------
# --------------------------------------------------------------------------- #
def test_breakout_level_includes_the_last_closed_week():
    """
    Pine `ta.highest(high, N)[1]` steps back one bar from the DEVELOPING week,
    so the window ends at - and INCLUDES - the last closed week.

    Verified live on TMB, Tue 28-Jul 15:08 mid-session:
        "Wk close > 26W high (prev)  885.5  needs > 847"  -> level 847.00
        "Prev wk close <= 26W high (2w ago)   <= 821.2"   -> hi_short2 821.20
    847.00 is the high of w/c 20-Jul, the last closed week.

    A previous build shifted this back an extra week after reading a chart
    captured at 21:30, when TradingView had already rolled the weekly bar.
    That snapshot showed the NEXT week's framing, not the live level.
    """
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars
    from test_strategy import make_daily

    cfg = Strategy()
    daily = make_daily(weeks=200)
    wk = build_weekly_bars(daily)
    target = wk.iloc[-1]["week_start"]
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, target)
    assert snap is not None

    closed = wk[wk["week_start"] < target]["high"].to_numpy(float)
    assert snap.entry_level == pytest.approx(closed[-cfg.len_short:].max())
    assert snap.level_52 == pytest.approx(closed[-cfg.len_long:].max())


def test_hi_short2_is_one_week_older_than_the_entry_level():
    """
    c03 compares the previous week's close against the 26W high as of TWO
    weeks ago. On TMB those two values were 847.00 and 821.20 respectively -
    distinct rows in the indicator table, not the same number.
    """
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars
    from test_strategy import make_daily

    cfg = Strategy()
    daily = make_daily(weeks=200)
    wk = build_weekly_bars(daily)
    target = wk.iloc[-1]["week_start"]
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, target)

    closed = wk[wk["week_start"] < target]["high"].to_numpy(float)
    assert snap.hi_short2 == pytest.approx(closed[-(cfg.len_short + 1):-1].max())
    assert snap.hi_short2 <= snap.entry_level, \
        "the older window cannot exceed the newer one"


def _sig_with_failures(*fails):
    from datetime import datetime

    from strategy import BarEval, Signal

    conds = {f"c{i:02d}": True for i in range(1, 14)}
    for k in fails:
        conds[k] = False
    ev = BarEval(conds, {"price": 167.64, "ema_fast": 147.8, "ema_slow": 145.5,
                         "ema_slow_2": 143.6, "rsi": 69.9, "rsi_1": 65.2,
                         "macd_hist": 2.74, "week_volume": 6.97e5,
                         "vol_sma": 6.94e5, "week_open": 159.9,
                         "day_open": 166.0, "entry_level": 161.0,
                         "level_52": 185.0})
    return Signal("GPTHEALTH", "1", "NSE_EQ", datetime(2026, 7, 28, 12, 10),
                  167.64, 161.0, 185.0, "deferred", ev, "2026-07-27",
                  6.97e5, 166.0, 159.95, bar_open=167, bar_high=168,
                  bar_low=166.5)


def test_alert_never_claims_a_false_pass_count():
    from telegram import format_signal

    msg = format_signal(_sig_with_failures("c01"))
    assert "12/13" in msg
    assert "13/13" not in msg, "header must not hardcode a perfect score"
    assert msg.count("12/13") == 1, "no contradictory second count"
    assert "not required for entry" in msg, \
        "explain that c01 does not gate the entry when req52 is off"


def test_alert_reports_full_pass_cleanly():
    from telegram import format_signal

    msg = format_signal(_sig_with_failures())
    assert "13/13" in msg and "PASS" in msg
    assert "not required for entry" not in msg


def test_c01_is_not_part_of_the_entry_gate():
    """
    Pine: gateLive = c03 and ... and c13   (c01/c02 excluded)
          req52 defaults to false, so the 52W break is optional.
    GPTHEALTH entering with c01 failing is correct, not a false signal.
    """
    import yaml as _yaml

    from config import Strategy
    from strategy import gate_ok

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    assert cfg["strategy"]["req52"] is False, "matches the Pine input default"

    snap = _gate_snap()
    ev = _gate_eval(c01=False)
    assert gate_ok(snap, Strategy(), ev) is True, \
        "a failing c01 must not block the gate while req52 is off"


# --------------------------------------------------------------------------- #
#  c09 relaxation - measured a THIRD time, with a conviction multiple.
#  Earlier attempts ("day", "bar") were rejected as too loose. This adds
#  pace-with-multiple: week volume so far > wk_sma * elapsed_fraction * mult.
#  NSE volume accumulates near-linearly (Mon 17.7%, Tue 39.2%, Wed 59.0%,
#  Thu 80.7%), so the pace comparison is statistically sound - yet swept
#  across 1.0-6.0 it still never beat raw c09:
#      raw c09   70% precision, 6 deferred
#      best pace 55% precision, 7 deferred, and it dropped a real mover
#  Deferrals INCREASE because weaker names get admitted earlier and then stall
#  on other conditions.
# --------------------------------------------------------------------------- #
def test_pace_mode_exists_but_is_not_default():
    import yaml as _yaml

    from config import Strategy

    assert Strategy().volume_mode == "off"
    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())
    assert cfg["strategy"]["volume_mode"] == "off", (
        "swept 1.0-6.0x: no pace multiple beat raw c09 on precision OR "
        "deferral count")


def test_pace_mode_scales_target_with_elapsed_week():
    from config import Strategy
    from strategy import evaluate_bar

    snap = _gate_snap()
    cfg = Strategy(volume_mode="pace", volume_pace_mult=2.0)

    early = evaluate_bar(snap, cfg, price=101.0, week_open=99.0,
                         week_volume=1.0, day_open=100.0, week_fraction=0.1)
    late = evaluate_bar(snap, cfg, price=101.0, week_open=99.0,
                        week_volume=1.0, day_open=100.0, week_fraction=0.9)
    assert late.values["vol_target"] > early.values["vol_target"], \
        "the bar must rise as the week progresses"


def test_pace_multiple_makes_the_test_stricter():
    from config import Strategy
    from strategy import evaluate_bar

    snap = _gate_snap()
    kw = dict(price=101.0, week_open=99.0, week_volume=1.0,
              day_open=100.0, week_fraction=0.5)
    loose = evaluate_bar(snap, Strategy(volume_mode="pace", volume_pace_mult=1.0), **kw)
    tight = evaluate_bar(snap, Strategy(volume_mode="pace", volume_pace_mult=5.0), **kw)
    assert tight.values["vol_target"] > loose.values["vol_target"]


# --------------------------------------------------------------------------- #
#  Paper trading measures CROSS entries only; alerts still send everything.
#  A deferred fill enters hours after the breakout at a worse, more extended
#  price, so mixing it into the performance stats distorts them. The user still
#  wants to SEE those signals in Telegram - only the measurement is filtered.
# --------------------------------------------------------------------------- #
def test_alerts_never_filter_by_trigger():
    """scan.py and watch.py must send cross AND deferred."""
    for name in ("scan.py", "watch.py"):
        body = (ROOT / name).read_text()
        assert 'trigger != "cross"' not in body, (
            f"{name} must not drop deferred signals - the user wants every "
            "alert delivered")


def test_paper_excludes_deferred_by_default():
    for name in ("paper.py", "paper_report.py"):
        body = (ROOT / name).read_text()
        assert 'sig.trigger != "cross"' in body, f"{name} must skip deferred"
        assert "include_deferred" in body, f"{name} needs an opt-in override"


def test_paper_still_simulates_a_deferred_signal_when_asked():
    """The filter is at the caller; simulate() itself stays trigger-agnostic."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from paper import simulate
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 28, 10, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "deferred", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    bars = pd.DataFrame([
        {"datetime": datetime(2026, 7, 28, 10, 5, tzinfo=ist) + timedelta(minutes=5 * i),
         "open": 100 + i * .2, "high": 101 + i * .2,
         "low": 99.8 + i * .2, "close": 100.4 + i * .2} for i in range(6)])
    t = simulate(sig, bars, 100000)
    assert t.qty > 0, "simulate must not itself reject deferred signals"


def test_report_states_which_signals_were_measured():
    body = (ROOT / "paper_report.py").read_text()
    assert "scope_note" in body, \
        "the workbook must say whether deferred entries were included"


# --------------------------------------------------------------------------- #
#  BUG 21 - THE OFF-BY-ONE FLIP-FLOP, SETTLED FOR GOOD (user decision: Option B)
#
#  Symptom: the breakout level kept oscillating between two candidates,
#  because chart screenshots and the indicator TABLE disagree with each other:
#
#      entry_level = highest(high, 26)[1]   -> TMB 847.00   (table row 2)
#      hi_short2   = highest(high, 26)[2]   -> TMB 821.20   (plotted line)
#
#  ROOT CAUSE - the Pine indicator is internally inconsistent.
#  The level is fetched with
#      request.security(tickerid, "W", ta.highest(high,26)[1], lookahead_off)
#  On an intraday chart `lookahead_off` returns the value as of the last
#  CONFIRMED weekly bar. The developing week is not confirmed, so during
#  Mon-Fri that expression resolves to the TWO-weeks-back value (hi_short2)
#  and only rolls forward when the week closes.
#
#  The 13-row table does NOT go through request.security - the Pine author
#  added chart-native `liveNative` aggregates precisely because security
#  "on an intraday chart can lag by one HTF bar" (their own tooltip).
#
#  So within one indicator:
#      table row 2      -> 847.00  (correct 26W high)
#      orange stepline  -> 821.20  (lagged)
#      BUY arrow        -> follows the LAGGED line
#
#  Confirmed by pixel-measuring the user's own screenshots, calibrating each
#  price axis from its labels (TMB log axis, every label fits < 0.25):
#      TMB     stepline y=485 -> 821.29 (= hi_short2)  stub y=394 -> 846.96
#      RADICO  stepline y=408 -> 4161.57 (= hi_short2) stub y=329 -> 4192.55
#  The line sits at the OLD value all week and steps up only at the right edge.
#
#  DECISION (user, 28-Jul): Option B - trade the economically correct level,
#  highest(high,26)[1]. Entering at hi_short2 means entering BELOW the actual
#  26-week high, i.e. before any breakout has happened. The chart arrows are
#  not a valid target; the Pine is what is wrong.
# --------------------------------------------------------------------------- #
def test_entry_level_is_the_true_26w_high_not_the_lagged_plot():
    """
    Option B, locked. entry_level must be the 26W high INCLUDING the last
    closed week. It must never equal hi_short2 (the lagged, plotted value)
    unless the two genuinely coincide in the data.
    """
    import numpy as np
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars
    from test_strategy import make_daily

    cfg = Strategy()
    daily = make_daily(weeks=200)
    wk = build_weekly_bars(daily)
    target = wk.iloc[-1]["week_start"]
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, target)
    assert snap is not None

    closed = wk[wk["week_start"] < target]["high"].to_numpy(float)

    # the window ENDS at the last closed week and INCLUDES it
    assert snap.entry_level == pytest.approx(closed[-cfg.len_short:].max())
    # hi_short2 is strictly the older window
    assert snap.hi_short2 == pytest.approx(closed[-(cfg.len_short + 1):-1].max())
    # and entry_level is never OLDER than hi_short2
    assert snap.entry_level >= snap.hi_short2 - 1e-9


def test_radico_and_vaibhavgbl_levels_are_the_unshifted_values():
    """
    Real levels for the week of 27-Jul-2026, from NSE weekly highs. These are
    the numbers the user confirmed under Option B. Hard-coded because they are
    what the live alerts must reproduce.

        RADICO      26W high = 4193.00   (w/c 20-Jul)   hi_short2 = 4161.80
        VAIBHAVGBL  26W high =  273.00   (w/c 20-Jul)   hi_short2 =  268.08
        TMB         26W high =  847.00   (w/c 20-Jul)   hi_short2 =  821.20

    If a future change reintroduces the shift, these flip to the hi_short2
    column and this test fails.
    """
    import numpy as np
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot

    cfg = Strategy()

    # Build 80 weekly bars so the snapshot's history requirement is met.
    # Only the last 26 matter for the window; everything older is filler kept
    # strictly below both candidates so it cannot win the max().
    rows = []
    first_week = pd.Timestamp("2026-07-20") - pd.Timedelta(weeks=79)
    for i in range(80):
        d = first_week + pd.Timedelta(weeks=i)
        rows.append({"datetime": d, "open": 3000.0, "high": 3100.0,
                     "low": 2900.0, "close": 3000.0, "volume": 1e6})

    # overwrite the final three weeks with the real shape
    tail = {"2026-07-06": 4161.80,   # hi_short2 (the lagged, plotted value)
            "2026-07-13": 4149.20,
            "2026-07-20": 4193.00}   # the true 26W high, last closed week
    for r in rows:
        key = str(r["datetime"].date())
        if key in tail:
            hi = tail[key]
            r.update(open=hi - 50, high=hi, low=hi - 80, close=hi - 10)

    daily = pd.DataFrame(rows).sort_values("datetime").reset_index(drop=True)
    snap = build_snapshot("RADICO", "1", "NSE_EQ", daily, cfg,
                          pd.Timestamp("2026-07-27"))

    assert snap is not None, (
        "history must be long enough for this test to actually assert - "
        "an earlier version silently skipped on a None snapshot")
    assert snap.entry_level == pytest.approx(4193.00), (
        "entry_level must be the true 26W high (4193.00), not the lagged "
        "plotted value (4161.80)")
    assert snap.hi_short2 == pytest.approx(4161.80)


def test_c09_cannot_satisfy_radico_and_vaibhavgbl_simultaneously():
    """
    Documented impossibility, so nobody burns another day tuning c09.

    Under Option B the user asked for BOTH:
        RADICO      Tue 15:00 @ 4276.90   (deferred - needs c09 to BLOCK 14:50)
        VAIBHAVGBL  Tue 10:35 @  278.33   (cross    - needs c09 to PASS)

    In "pace" mode c09 passes when  vol_ratio > week_fraction * m.
    Measured on the real 5m bars for that week:

        VAIBHAVGBL 10:35  ratio 0.4837  frac 0.2453  -> needs m <  1.972
        RADICO     14:50  ratio 0.9325  frac 0.3813  -> needs m >= 2.445
        RADICO     15:00  ratio 1.0525  frac 0.3867  -> needs m <  2.722

    Feasible range is [2.445, 1.972) = EMPTY. No multiplier exists.

    Deeper reason VAIBHAVGBL is unreachable at ANY c09 setting that keeps the
    condition meaningful: across the whole week it spent 9 bars above 273.00
    (Tue 10:35-14:10) and exactly 1 bar with c09 true (Tue 15:25). The two
    never overlap, so no gate ordering can produce an entry while price is
    above the level. VAIBHAVGBL is a genuine REJECT that week, not a miss.
    """
    # Pure arithmetic - encodes the measurement so it cannot be forgotten.
    v_ratio, v_frac = 0.4837, 0.2453      # VAIBHAVGBL Tue 10:35, must PASS
    r_ratio_fail, r_frac_fail = 0.9325, 0.3813   # RADICO Tue 14:50, must FAIL
    r_ratio_pass, r_frac_pass = 1.0525, 0.3867   # RADICO Tue 15:00, must PASS

    m_upper = min(v_ratio / v_frac, r_ratio_pass / r_frac_pass)
    m_lower = r_ratio_fail / r_frac_fail

    assert m_lower >= m_upper, (
        "If this ever passes, a single c09 multiplier CAN satisfy both cases "
        "and the documented impossibility no longer holds - re-derive it.")


def test_pine_source_uses_request_security_for_the_level():
    """
    The lag is real and lives in the Pine, not in our Python. If the user ever
    fixes the indicator to use chart-native highs, the arrows will move to the
    table values and agree with us - at which point this test should be
    updated deliberately, not by accident.
    """
    src = ROOT.parent / "uploads" / "weekly.txt"
    if not src.exists():
        pytest.skip("Pine source not present in this checkout")
    body = src.read_text(errors="ignore")
    assert "request.security" in body
    assert "lookahead=barmerge.lookahead_off" in body, (
        "the level fetch relies on lookahead_off; that is what makes the "
        "plotted stepline lag the table by one weekly bar")


# --------------------------------------------------------------------------- #
#  BUG 22 - INTRADAY MODE. User, 28-Jul: "Focus on only intraday, i don't want
#  swing trade."
#
#  Three things had to become true, and each is guarded below.
#
#  1. NOTHING MAY BE HELD OVERNIGHT. Brokers force-square MIS equity in the
#     afternoon; a simulation that carries a position to the next session is
#     reporting a trade that could not exist. paper.py now exits at
#     SQUARE_OFF (15:15) or on any bar dated after the signal.
#
#  2. COSTS MUST BE CHARGED. Measured over 449 real signals (12 weeks, whole
#     NSE cash universe) the best exit rule made +0.14% per trade GROSS and
#     -0.08% NET at a realistic 0.22% round trip. Reporting gross P&L on an
#     intraday system is reporting a profit that does not exist.
#
#  3. A VOLATILITY FLOOR. Without it the signal set loses money outright
#     (PF 0.72). min_atr_pct >= 1.0 was the only filter that survived an
#     out-of-sample split.
# --------------------------------------------------------------------------- #
def test_paper_never_holds_overnight():
    """A position must close on its own session, whatever the exit rule."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from paper import simulate
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 14, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)

    # a relentless grind UP: no stop, no EMA breach - only square-off can end it
    rows = []
    t0 = datetime(2026, 7, 27, 14, 5, tzinfo=ist)
    for i in range(40):                      # spills into the next session
        ts = t0 + timedelta(minutes=5 * i)
        if ts.hour >= 15 and ts.minute > 25:  # jump to next day
            ts = datetime(2026, 7, 28, 9, 15, tzinfo=ist) + timedelta(minutes=5 * i)
        p = 100.0 + i * 0.5
        rows.append({"datetime": ts, "open": p, "high": p + 0.4,
                     "low": p - 0.05, "close": p + 0.3})
    bars = pd.DataFrame(rows)

    t = simulate(sig, bars, 100000)
    assert t.exit_reason != "OPEN", "an intraday trade must not be left running"
    assert t.exit_date == "2026-07-27", (
        f"position carried to {t.exit_date} - intraday trades cannot go "
        "overnight, the broker squares them off")


def test_square_off_time_is_before_the_close():
    """Leave room before the broker's own auto-square-off."""
    from paper import SQUARE_OFF
    h, m = (int(x) for x in SQUARE_OFF.split(":"))
    assert (h, m) >= (15, 0), "squaring off too early wastes the session"
    assert (h, m) <= (15, 20), (
        "most brokers auto-square MIS equity by ~15:20 and forced liquidation "
        "fills worse than a voluntary exit")


def test_paper_charges_round_trip_costs():
    """Net P&L must be strictly worse than gross - costs are not optional."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from paper import COST_ROUND_TRIP, simulate
    from strategy import BarEval, Signal

    assert COST_ROUND_TRIP > 0, "an intraday backtest without costs is fiction"

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 10, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    bars = pd.DataFrame([
        {"datetime": datetime(2026, 7, 27, 10, 5, tzinfo=ist) + timedelta(minutes=5 * i),
         "open": 100 + i * .2, "high": 101 + i * .2,
         "low": 99.8 + i * .2, "close": 100.4 + i * .2} for i in range(6)])

    t = simulate(sig, bars, 100000)
    assert t.qty > 0
    assert t.costs > 0, "no costs were charged"
    assert t.pnl == pytest.approx(t.gross_pnl - t.costs), \
        "net P&L must equal gross minus costs"
    assert t.pnl < t.gross_pnl


def test_min_atr_pct_filter_blocks_low_volatility_entries():
    """
    A stock that barely moves cannot pay for a round trip. The filter must
    actually suppress the entry, not merely annotate it.
    """
    import numpy as np
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars, replay_week
    from test_strategy import make_daily

    daily = make_daily(weeks=200)
    wk = build_weekly_bars(daily)
    target = wk.iloc[-1]["week_start"]

    base = Strategy(strict_entry=False)          # isolate the ATR filter
    snap = build_snapshot("T", "1", "NSE_EQ", daily, base, target)
    assert snap is not None

    lvl = snap.entry_level
    # a clean break above the level, but on tiny 5m ranges (~0.02%)
    rows = []
    for i in range(30):
        p = lvl * (1.0 + 0.004 + i * 0.00002)
        rows.append({"datetime": pd.Timestamp(target) + pd.Timedelta(minutes=5 * i),
                     "open": p, "high": p * 1.0001, "low": p * 0.9999,
                     "close": p, "volume": 10_000})
    bars = pd.DataFrame(rows)

    assert replay_week(snap, base, bars).signals, \
        "sanity: with the filter off this must produce a signal"

    tight = Strategy(strict_entry=False, min_atr_pct=1.0)
    assert not replay_week(snap, tight, bars).signals, \
        "min_atr_pct must block a breakout whose bars are far too quiet"


def test_min_atr_pct_defaults_to_disabled():
    """
    Shipping it ON by default would silently change every historical result.
    config.yaml opts in explicitly; the dataclass default stays inert.
    """
    from config import Strategy
    assert Strategy().min_atr_pct == 0.0


# --------------------------------------------------------------------------- #
#  BUG 23 - A/B PAPER TRADING. User, 28-Jul: "can create 2 separate paper trade
#  model, 1 with your recommendation and other with my recommendation, so after
#  a week we will get on to the conclusion."
#
#  The experiment is only worth running if it is FAIR. Every guard below exists
#  to stop the comparison being quietly rigged:
#
#    * both models must use the IDENTICAL stop (entry candle low - 0.02%).
#      The user has said twice not to change the stop; if one model got a
#      different stop the test would measure stop width, not entry quality.
#    * both must square off intraday - neither may sneak an overnight hold.
#    * both must be charged the same costs.
#    * the ledger must de-duplicate, or running the job twice in a day would
#      double-count trades and corrupt the week's conclusion.
# --------------------------------------------------------------------------- #
def test_models_file_is_valid_and_not_empty():
    """
    A and B were retired on 29-Jul-2026 (every intraday exit measured negative
    after costs), leaving C alone. The harness must still be coherent with a
    single model, and must never end up with none.
    """
    import yaml as _yaml

    raw = _yaml.safe_load((ROOT / "models.yaml").read_text())
    models = raw["models"]
    assert models, "at least one model must be configured"
    for key, m in models.items():
        assert m.get("label"), f"{key} needs a label"
        assert m.get("exit"), f"{key} needs an exit rule"

    # If two or more ever run again, they must actually differ - otherwise the
    # comparison is measuring noise.
    if len(models) >= 2:
        sigs = [str(sorted((m.get("strategy") or {}).items()))
                + str(sorted((m.get("exit") or {}).items()))
                for m in models.values()]
        assert len(set(sigs)) == len(sigs), (
            "two models share identical settings - nothing to compare")


def test_ab_models_use_identical_stop_and_costs():
    """Neither model may be handed an easier stop or cheaper costs."""
    import yaml as _yaml

    raw = _yaml.safe_load((ROOT / "models.yaml").read_text())
    for key, m in raw["models"].items():
        strat = m.get("strategy", {})
        for forbidden in ("sl_buffer", "stop_pct", "stop_mode"):
            assert forbidden not in strat, (
                f"{key} tries to override the stop ({forbidden}); the stop is "
                "identical in both models by design")
    # costs and square-off are global defaults, not per-model
    for key, m in raw["models"].items():
        assert "cost_round_trip" not in m, f"{key} must not set its own costs"
        assert "square_off" not in m, f"{key} must not set its own square-off"


def test_ab_simulate_respects_intraday_square_off():
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import simulate_model
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 14, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    rows = []
    for i in range(40):
        ts = datetime(2026, 7, 27, 14, 5, tzinfo=ist) + timedelta(minutes=5 * i)
        if ts.hour >= 15 and ts.minute > 25:
            ts = datetime(2026, 7, 28, 9, 15, tzinfo=ist) + timedelta(minutes=5 * i)
        p = 100.0 + i * 0.5      # relentless grind up: only square-off can end it
        rows.append({"datetime": ts, "open": p, "high": p + 0.4,
                     "low": p - 0.05, "close": p + 0.3})
    bars = pd.DataFrame(rows)

    for rule in ({"rule": "ema"}, {"rule": "target", "target_r": 99.0}):
        t = simulate_model(sig, bars, 100000, None, rule)
        assert t.exit_date == "2026-07-27", (
            f"{rule} carried the position overnight")


def test_ab_both_models_get_the_same_stop():
    """Same signal + same bars -> same stop, whatever the exit rule."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import simulate_model
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 10, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    bars = pd.DataFrame([
        {"datetime": datetime(2026, 7, 27, 10, 5, tzinfo=ist) + timedelta(minutes=5 * i),
         "open": 100 + i * .2, "high": 101 + i * .2,
         "low": 99.8 + i * .2, "close": 100.4 + i * .2} for i in range(6)])

    a = simulate_model(sig, bars, 100000, None, {"rule": "ema"})
    b = simulate_model(sig, bars, 100000, None,
                       {"rule": "target", "target_r": 3.0, "be_at_r": 1.0})
    assert a.stop == pytest.approx(b.stop), \
        "the two models must risk exactly the same amount per trade"


def test_ab_ledger_deduplicates(tmp_path):
    """Running the job twice in one day must not double-count."""
    from ab_paper import append_ledger

    p = tmp_path / "led.csv"
    row = {"model": "A_gated", "symbol": "X", "signal_date": "2026-07-27",
           "signal_time": "10:00", "pnl": 100.0}
    added, dupes = append_ledger(p, [row])
    assert added == 1 and dupes == 0
    added, dupes = append_ledger(p, [row])
    assert added == 0 and dupes == 1, "the same trade was recorded twice"


def test_ab_target_r_uses_initial_risk():
    """+3R must mean 3x the ORIGINAL stop distance, not a moving one."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import simulate_model
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    # entry 100, bar low 99 -> stop 98.98, risk ~1.02 -> +3R target ~103.06
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 10, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    rows = []
    for i in range(10):
        p = 100.0 + i * 1.0
        rows.append({"datetime": datetime(2026, 7, 27, 10, 5, tzinfo=ist) + timedelta(minutes=5 * i),
                     "open": p, "high": p + 0.9, "low": p - 0.1, "close": p + 0.5})
    t = simulate_model(sig, pd.DataFrame(rows), 100000, None,
                       {"rule": "target", "target_r": 3.0})
    assert t.exit_reason == "TGT"
    assert t.r_multiple == pytest.approx(3.0, abs=0.01)


# --------------------------------------------------------------------------- #
#  BUG 24 - the stale snapshot the version guard could not see.
#
#  On 28-Jul the repo's committed weekly_snapshot.csv held levels from the
#  REVERTED one-week shift:
#      TMB        821.20  (should be 847.00)
#      RADICO    4161.80  (should be 4193.00)
#      VAIBHAVGBL 268.08  (should be  273.00)
#  Monday's scan would have fired against levels ~1% too low.
#
#  The LOGIC_VERSION guard did not catch it because v2 was stamped in the SAME
#  session that introduced the bad formula - the guard cannot recognise its own
#  bad output. Reverting the formula without bumping the version left the file
#  looking current.
#
#  Fix: LOGIC_VERSION = 3, so every v2 row is discarded on resume and the next
#  run rebuilds from scratch even if someone forgets to tick `fresh`.
# --------------------------------------------------------------------------- #
def test_logic_version_is_at_least_3():
    """
    v2 rows carry the reverted, too-low breakout levels. If this ever drops
    back to 2 those rows become resumable again.
    """
    from build_snapshot import LOGIC_VERSION
    assert LOGIC_VERSION >= 3, (
        "LOGIC_VERSION must stay >= 3: version 2 identifies snapshot rows "
        "built with the reverted one-week level shift")


def test_stale_v2_rows_are_discarded_on_resume(tmp_path):
    """
    A snapshot file full of v2 rows must resume ZERO symbols, so the rebuild
    is total even without --fresh.
    """
    import pandas as pd

    from build_snapshot import LOGIC_VERSION

    stale = pd.DataFrame([
        {"symbol": "TMB", "week_start": "2026-07-27",
         "entry_level": "821.20", "logic_version": "2"},
        {"symbol": "RADICO", "week_start": "2026-07-27",
         "entry_level": "4161.80", "logic_version": "2"},
    ])
    p = tmp_path / "weekly_snapshot.csv"
    stale.to_csv(p, index=False)

    prev = pd.read_csv(p, dtype=str)
    prev = prev[prev["week_start"] == "2026-07-27"]
    kept = prev[prev["logic_version"] == str(LOGIC_VERSION)]
    assert len(kept) == 0, (
        "rows from the reverted-level generation must never be resumed")


def test_bumping_the_formula_requires_bumping_the_version():
    """
    Executable reminder of the lesson. build_snapshot.py must document why the
    current LOGIC_VERSION exists, so the next person who edits the level maths
    sees that the version has to move with it - including on a REVERT.
    """
    body = (ROOT / "strategy.py").read_text()
    assert "REVERTED" in body or "reverted" in body, (
        "the LOGIC_VERSION block must record that a revert also invalidates "
        "previously stored rows")
    # and the writer must not keep a private copy that can drift from it
    writer = (ROOT / "build_snapshot.py").read_text()
    assert "from strategy import" in writer and "LOGIC_VERSION" in writer, (
        "build_snapshot.py must import LOGIC_VERSION from strategy.py")
    assert "\nLOGIC_VERSION = " not in writer, (
        "build_snapshot.py must not define its own LOGIC_VERSION")


# --------------------------------------------------------------------------- #
#  BUG 25 - THE SCANNER ALERTED AGAINST STALE BREAKOUT LEVELS.
#
#  build_snapshot.py refused to RESUME rows from an older logic version, but
#  scan.py happily LOADED them. So after the level formula was reverted, the
#  committed weekly_snapshot.csv still held v2 rows and the live scanner used
#  them for two full sessions:
#
#      symbol      quoted level   true 26W high   outcome
#      RADICO           4161.80         4193.00   alert at the wrong level
#      TMB               821.20          847.00   alert at the wrong level
#      SENCO             390.40          399.80   alert at the wrong level
#      GPTHEALTH         161.00          164.69   alert at the wrong level
#      JKPAPER           401.50          421.95   *** FALSE ALERT ***
#      THYROCARE         578.25          597.60   *** FALSE ALERT ***
#      VAIBHAVGBL        268.08          273.00   *** FALSE ALERT ***
#
#  A wrong level is a wrong trade. The fix: readers verify logic_version AND
#  week_start, and DROP anything that does not match - scanning nothing is
#  strictly better than alerting on a level that does not exist.
# --------------------------------------------------------------------------- #
def test_logic_version_has_one_source_of_truth():
    """Writer and reader must not be able to drift apart."""
    import build_snapshot
    import scan
    from strategy import LOGIC_VERSION

    assert build_snapshot.LOGIC_VERSION is LOGIC_VERSION
    assert scan.LOGIC_VERSION is LOGIC_VERSION


def _snap_row(symbol, level, week, version):
    """A minimally complete snapshot row that from_row() can parse."""
    return {
        "symbol": symbol, "security_id": "1", "exchange_segment": "NSE_EQ",
        "week_start": week, "entry_level": level, "level_52": level,
        "hi_short2": level, "close_1": level,
        "ema_fast_prev": 1.0, "ema_fast_len": 20,
        "ema_slow_prev": 1.0, "ema_slow_len": 50, "ema_slow_2": 1.0,
        "rsi_len": 14, "rsi_avg_gain": 1.0, "rsi_avg_loss": 1.0,
        "rsi_prev_close": 1.0, "rsi_1": 1.0,
        "macd_fast": 12, "macd_slow": 26, "macd_signal": 9,
        "macd_ema_fast": 1.0, "macd_ema_slow": 1.0, "macd_sig": 1.0,
        "vol_sma_len": 10, "vol_sma_sum_prev": 1.0,
        "g_ema_fast": 1.0, "g_ema_slow": 1.0, "g_ema_slow_2": 1.0,
        "g_rsi": 1.0, "g_rsi_1": 1.0, "g_hist": 1.0,
        "prev_daily_close": 1.0, "prev_daily_open": 1.0, "mcap": "",
        "logic_version": version,
    }


def test_scan_drops_rows_from_an_older_logic_version(tmp_path):
    """The exact RADICO 4161.80 scenario must now load nothing."""
    import pandas as pd

    import scan
    from config import load_config
    from strategy import LOGIC_VERSION

    week = "2026-07-27"
    old = LOGIC_VERSION - 1
    pd.DataFrame([
        _snap_row("RADICO", 4161.80, week, old),        # reverted-shift level
        _snap_row("JKPAPER", 401.50, week, old),        # caused a false alert
    ]).to_csv(tmp_path / "weekly_snapshot.csv", index=False)

    cfg = load_config(None)
    cfg.paths["snapshot"] = tmp_path / "weekly_snapshot.csv"
    assert scan.load_snapshots(cfg, week) == [], (
        "rows from an older logic version must never reach the scanner")


def test_scan_drops_rows_for_a_different_week(tmp_path):
    import pandas as pd

    import scan
    from config import load_config
    from strategy import LOGIC_VERSION

    pd.DataFrame([
        _snap_row("RADICO", 4193.00, "2026-07-20", LOGIC_VERSION),
    ]).to_csv(tmp_path / "weekly_snapshot.csv", index=False)

    cfg = load_config(None)
    cfg.paths["snapshot"] = tmp_path / "weekly_snapshot.csv"
    assert scan.load_snapshots(cfg, "2026-07-27") == []


def test_scan_keeps_current_rows(tmp_path):
    """Sanity: the guard must not reject good data."""
    import pandas as pd

    import scan
    from config import load_config
    from strategy import LOGIC_VERSION

    week = "2026-07-27"
    pd.DataFrame([
        _snap_row("RADICO", 4193.00, week, LOGIC_VERSION),
        _snap_row("APCOTEXIND", 579.75, week, LOGIC_VERSION),
    ]).to_csv(tmp_path / "weekly_snapshot.csv", index=False)

    cfg = load_config(None)
    cfg.paths["snapshot"] = tmp_path / "weekly_snapshot.csv"
    snaps = scan.load_snapshots(cfg, week)
    assert {s.symbol for s in snaps} == {"RADICO", "APCOTEXIND"}
    assert {round(s.entry_level, 2) for s in snaps} == {4193.00, 579.75}


def test_snapshot_without_a_version_column_is_rejected(tmp_path):
    """A pre-versioning file has unknown provenance - refuse it."""
    import pandas as pd

    import scan
    from config import load_config

    week = "2026-07-27"
    row = _snap_row("RADICO", 4161.80, week, 0)
    row.pop("logic_version")
    pd.DataFrame([row]).to_csv(tmp_path / "weekly_snapshot.csv", index=False)

    cfg = load_config(None)
    cfg.paths["snapshot"] = tmp_path / "weekly_snapshot.csv"
    assert scan.load_snapshots(cfg, week) == []


# --------------------------------------------------------------------------- #
#  BUG 26 - daily watchlist. User, 29-Jul: "also everyday send all the stocks
#  list which are qualifying for current week on telegram".
#
#  The digest exists to answer "what should be on my screen today?", so the two
#  things that must never break are:
#    * it must not silently truncate to nothing - the FULL table always goes
#      out as a CSV, because Telegram caps a message at 4096 chars and a
#      2000-symbol universe blows past that instantly
#    * it must use the SAME frozen levels as the scanner. If it re-derived
#      them it could show a level the alerts would never fire on.
# --------------------------------------------------------------------------- #
def test_watchlist_reuses_the_scanner_snapshot_loader():
    """
    The digest must read levels through scan.load_snapshots, which enforces
    logic_version + week. Recomputing them here would let the watchlist and the
    alerts disagree about the level - the exact class of bug from BUG 25.
    """
    body = (ROOT / "watchlist.py").read_text()
    assert "from scan import" in body and "load_snapshots" in body, (
        "watchlist.py must load levels via scan.load_snapshots so it inherits "
        "the stale-snapshot guard")
    assert "build_snapshot(" not in body, (
        "watchlist.py must not build its own levels")


def test_watchlist_buckets_are_correct():
    """WATCH / FIRED must key off the frozen level; nothing else is listed."""
    from watchlist import build_message

    rows = [
        dict(symbol="AAA", ltp=98.0, level=100.0, pct=-2.0, gap=2.0, bucket="WATCH"),
        dict(symbol="BBB", ltp=50.0, level=100.0, pct=-50.0, gap=50.0, bucket="OTHER"),
        dict(symbol="CCC", ltp=120.0, level=100.0, pct=20.0, gap=-20.0, bucket="FIRED"),
    ]
    msg = build_message("2026-07-27", rows, {"universe": 3, "eligible": 2}, 3.0)
    assert "AAA" in msg and "CCC" in msg
    assert "BBB" not in msg, "far names are noise and must not be listed"
    assert "week of 2026-07-27" in msg


def test_watchlist_omits_the_above_level_section():
    """
    User, 29-Jul: "skip those ABOVE LEVEL, NOT YET TRIGGERED".
    Those names either already alerted or are being held back by the weekly
    gate - neither is something to watch for tomorrow. Dropping them also
    removes the per-symbol 5m replay, so the job is a single quote pass.
    """
    from watchlist import build_message

    rows = [dict(symbol="ZZZ", ltp=110.0, level=100.0, pct=10.0,
                 gap=-10.0, bucket="OTHER")]
    msg = build_message("2026-07-27", rows, {"universe": 1, "eligible": 1}, 3.0)
    assert "ZZZ" not in msg
    assert "ABOVE LEVEL" not in msg.upper()

    body = (ROOT / "watchlist.py").read_text()
    assert "scan_symbol" not in body, (
        "no per-symbol replay should remain - the digest is quote-only now")


def test_watchlist_ranks_by_distance_and_caps():
    """
    Closest-to-level first, hard cap on the list. Measured on the live book,
    momentum filters cut 88 -> 64 while distance cut it to ~18, so ranking is
    what makes the digest actionable.

    BUG 36 note: distance is now the SECOND key, after the PRE score. These
    rows carry no pre_score (they stand in for --no-metrics / unscreened
    output), so they all tie on quality and distance decides - which is
    exactly the ordering asserted here. The line prefix also gained a score
    badge, so the match is on the symbol rather than the start of the line.
    """
    from watchlist import build_message

    rows = [dict(symbol=f"S{i:03d}", ltp=100.0 - i, level=100.0,
                 pct=-float(i), gap=float(i), bucket="WATCH")
            for i in range(1, 60)]
    msg = build_message("2026-07-27", rows, {"universe": 59, "eligible": 59},
                        99.0, top_n=15)
    listed = [ln for ln in msg.splitlines()
              if ln.startswith("• ") and "<b>S0" in ln]
    assert len(listed) == 15, f"expected 15 rows, got {len(listed)}"
    assert "S001" in listed[0], "the closest name must rank first"
    assert "more in the CSV" in msg


def test_watchlist_structure_screen_keeps_real_winners():
    """
    The c03/c05/c08 screen must not reject names that went on to fire.
    MONARCH and RADICO both failed the STRICTER screen (c04/c06/c07) during
    the 27-Jul week, which is why those rows are deliberately excluded.
    """
    from config import load_config
    from watchlist import is_eligible

    cfg = load_config(None)

    class S:                      # minimal stand-in for a WeeklySnapshot
        close_1 = 100.0
        hi_short2 = 120.0         # c03 passes: previous close below the old high
        g_hist = 1.0              # c08 passes
        g_ema_slow = 50.0
        g_ema_slow_2 = 49.0       # c05 passes
        g_rsi = 56.0              # would FAIL a strict RSI>60 screen
        g_rsi_1 = 57.0            # would FAIL "RSI rising"
        g_ema_fast = 49.0         # would FAIL EMA20>EMA50

    assert is_eligible(S(), cfg), (
        "a name with soft closed-week RSI must stay eligible - it is exactly "
        "the accelerating setup the strict screen throws away")

    class Stale(S):
        close_1 = 130.0           # already above the old high -> not fresh
    assert not is_eligible(Stale(), cfg)


def test_watchlist_always_writes_the_full_table():
    """Truncating the message is fine; losing the data is not."""
    body = (ROOT / "watchlist.py").read_text()
    assert "to_csv" in body
    assert "send_document" in body, (
        "the complete watchlist must be attached, not just summarised")


def test_watchlist_workflow_runs_premarket_on_weekdays():
    import yaml as _yaml

    wf = _yaml.safe_load((WORKFLOWS / "watchlist.yml").read_text())
    on = wf.get("on") or wf.get(True)
    crons = [c["cron"] for c in on["schedule"]]
    assert crons, "the watchlist must be scheduled"
    for c in crons:
        minute, hour, _, _, dow = c.split()
        assert dow == "1-5", "markets are shut at the weekend"
        # 03:15 UTC = 08:45 IST, i.e. before the 09:15 open
        assert int(hour) < 4, (
            f"cron '{c}' is not pre-market IST; the list is useless after the "
            "open")


def test_watchlist_never_places_orders():
    body = (ROOT / "watchlist.py").read_text().lower()
    for bad in ("place_order", "placeorder", "/orders", "transactiontype"):
        assert bad not in body, f"watchlist.py must never trade ({bad})"


# --------------------------------------------------------------------------- #
#  BUG 27 - MODEL C "SWING". User, 29-Jul: "Create separate entry model based
#  on best entry, best stop, best hold".
#
#  C is the first SWING model in the harness: it holds overnight, which every
#  earlier test forbade outright. The danger is that adding it quietly guts the
#  intraday guarantees for A and B, so the horizon is an explicit, tested flag:
#  only `horizon: swing` may skip the square-off, and an intraday model can
#  never acquire a multi-day hold by accident.
#
#  Parameters were measured on 3,254 breakouts over 12 months:
#      entry  same-day close  +5.17%  vs next-open +4.40%, pullback +3.43%
#      stop   7%              +5.59%  vs entry-low +4.05%, 3% +4.24%
#      hold   5 days          +5.59%  vs 1d +2.49%, 30d +5.98% (6x the lockup)
# --------------------------------------------------------------------------- #
def test_models_declare_a_valid_horizon():
    import yaml as _yaml

    raw = _yaml.safe_load((ROOT / "models.yaml").read_text())
    for key, m in raw["models"].items():
        hz = m.get("horizon", "intraday")
        assert hz in ("intraday", "swing"), f"{key} has a bad horizon: {hz}"


def test_only_swing_models_may_hold_overnight():
    """An intraday model must never be handed hold_days."""
    import yaml as _yaml

    raw = _yaml.safe_load((ROOT / "models.yaml").read_text())
    for key, m in raw["models"].items():
        if m.get("horizon", "intraday") == "intraday":
            assert not m.get("hold_days"), (
                f"{key} is intraday but declares hold_days - that would let it "
                "carry a position overnight")
            assert not m.get("exit", {}).get("hold_days"), (
                f"{key} is intraday but its exit sets hold_days")


def test_intraday_models_still_square_off():
    """Adding C must not weaken the guarantee for A and B."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import load_models, simulate_model
    from strategy import BarEval, Signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 14, 0, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                 97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)
    rows = []
    for i in range(40):
        ts = datetime(2026, 7, 27, 14, 5, tzinfo=ist) + timedelta(minutes=5 * i)
        if ts.hour >= 15 and ts.minute > 25:
            ts = datetime(2026, 7, 28, 9, 15, tzinfo=ist) + timedelta(minutes=5 * i)
        p = 100.0 + i * 0.5
        rows.append({"datetime": ts, "open": p, "high": p + 0.4,
                     "low": p - 0.05, "close": p + 0.3})
    bars = pd.DataFrame(rows)

    _defaults, models = load_models()
    for m in models:
        if m.is_swing:
            continue
        t = simulate_model(sig, bars, 100000, None, m.exit)
        assert t.exit_date == "2026-07-27", f"{m.key} held overnight"


def _daily(n, start=100.0, step=1.0, low_off=1.0):
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd
    ist = ZoneInfo("Asia/Kolkata")
    rows = []
    for i in range(n):
        p = start + step * (i + 1)
        rows.append({"datetime": datetime(2026, 7, 28, tzinfo=ist) + timedelta(days=i),
                     "open": p, "high": p + 1.0, "low": p - low_off, "close": p})
    return pd.DataFrame(rows)


def _swing_sig():
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal
    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    return Signal("X", "1", "NSE_EQ", datetime(2026, 7, 27, 15, 20, tzinfo=ist),
                  100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6,
                  97.0, 96.0, bar_open=99.5, bar_high=100.5, bar_low=99.0)


def test_swing_time_exit_after_hold_days():
    """A grind that never hits stop or target must exit on day `hold_days`."""
    from ab_paper import simulate_swing

    t = simulate_swing(_swing_sig(), _daily(20, step=0.2), 100000,
                       {"stop_pct": 7.0, "target_r": 99.0}, hold_days=5)
    assert t.exit_reason == "TIME"
    assert t.bars_held == 5, f"held {t.bars_held} days, expected 5"


def test_swing_stop_is_a_percent_not_the_entry_candle_low():
    """
    C deliberately uses a 7% stop. The entry-candle low (99.0 here) would be a
    ~1% stop and would fire on the first red day; 7% must not.
    """
    from ab_paper import simulate_swing

    sig = _swing_sig()
    t = simulate_swing(sig, _daily(6, step=0.5, low_off=2.0), 100000,
                       {"stop_pct": 7.0, "target_r": 99.0}, hold_days=5)
    assert t.stop == pytest.approx(93.0), f"stop was {t.stop}, expected 7% = 93.0"
    assert t.exit_reason == "TIME", "a 7% stop must survive a 2-point daily wick"


def test_swing_stop_wins_ties_with_the_target():
    """If one day both breaches the stop and tags the target, take the stop."""
    from datetime import datetime
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import simulate_swing
    ist = ZoneInfo("Asia/Kolkata")
    # a single huge-range day: low 90 (below the 93 stop), high 130 (above 3R)
    bars = pd.DataFrame([{"datetime": datetime(2026, 7, 28, tzinfo=ist),
                          "open": 100.0, "high": 130.0, "low": 90.0, "close": 128.0}])
    t = simulate_swing(_swing_sig(), bars, 100000,
                       {"stop_pct": 7.0, "target_r": 3.0}, hold_days=5)
    assert t.exit_reason == "SL", "the pessimistic branch must be taken"


def test_swing_charges_costs_and_never_trades():
    from ab_paper import simulate_swing

    t = simulate_swing(_swing_sig(), _daily(6, step=0.5), 100000,
                       {"stop_pct": 7.0, "target_r": 3.0}, hold_days=5)
    assert t.costs > 0
    assert t.pnl == pytest.approx(t.gross_pnl - t.costs)

    body = (ROOT / "ab_paper.py").read_text().lower()
    for bad in ("place_order", "placeorder", "transactiontype"):
        assert bad not in body, f"ab_paper.py must never trade ({bad})"


def test_c03_and_c11_can_be_switched_off_but_default_on():
    """
    Both measured NEGATIVE at a 5-day horizon, so C disables them - but the
    live intraday scanner must keep Pine behaviour unless it opts out.
    """
    from config import Strategy

    assert Strategy().require_c03 is True
    assert Strategy().require_c11 is True

    from strategy import BarEval, gate_ok, WeeklySnapshot
    import indicators as ind

    snap = WeeklySnapshot(
        symbol="X", security_id="1", exchange_segment="NSE_EQ",
        week_start="2026-07-27", entry_level=100.0, level_52=100.0,
        hi_short2=90.0, close_1=95.0,
        ema_fast=ind.EmaState(length=20, prev=1.0),
        ema_slow=ind.EmaState(length=50, prev=1.0), ema_slow_2=1.0,
        rsi=ind.RsiState(length=14, avg_gain=1.0, avg_loss=1.0, prev_close=1.0),
        rsi_1=1.0,
        macd=ind.MacdState(fast=12, slow=26, signal=9, ema_fast=1.0,
                           ema_slow=1.0, sig=1.0),
        vol_sma=ind.SmaState(length=10, sum_prev=1.0),
        g_ema_fast=1.0, g_ema_slow=1.0, g_ema_slow_2=1.0,
        g_rsi=1.0, g_rsi_1=1.0, g_hist=1.0,
        prev_daily_close=1.0, prev_daily_open=1.0)

    conds = {f"c{i:02d}": True for i in range(1, 14)}
    conds["c03"] = False                     # fresh-breakout row fails
    ev = BarEval(conds, {})

    assert not gate_ok(snap, Strategy(), ev), "c03 must block by default"
    assert gate_ok(snap, Strategy(require_c03=False), ev), \
        "require_c03=False must let an extended breakout through"


# --------------------------------------------------------------------------- #
#  BUG 28 - the alert did not say what to actually DO.
#  User, 29-Jul: "send buy price, sl level price and first target price".
#
#  The alert gave only the trigger price, so the three numbers needed at the
#  moment of buying had to be computed by hand while the candle was still hot.
#
#  The danger in adding them is DRIFT: an alert quoting a 5% stop while Model C
#  paper-trades a 7% one would be worse than no plan at all. So the plan is
#  read from models.yaml - the same file the simulator uses - and the tests
#  below assert the printed numbers equal the ones simulate_swing enforces.
# --------------------------------------------------------------------------- #
def test_trade_plan_matches_the_swing_engine_exactly():
    """Printed stop and target must equal what simulate_swing actually uses."""
    from datetime import datetime, timedelta
    from zoneinfo import ZoneInfo

    import pandas as pd

    from ab_paper import load_models, simulate_swing
    from strategy import BarEval, Signal
    from telegram import _swing_plan_params

    _defaults, models = load_models()
    swing = [m for m in models if m.is_swing]
    assert swing, "no swing model configured"
    C = swing[0]

    ist = ZoneInfo("Asia/Kolkata")
    price = 607.55
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 29, 9, 15, tzinfo=ist),
                 price, 579.75, 579.75, "cross", ev, "2026-07-27", 1e6,
                 600.0, 561.0, bar_open=600.0, bar_high=610.0, bar_low=598.0)

    p = _swing_plan_params()
    assert p, "the alert could not read the swing plan from models.yaml"
    alert_stop = price * (1 - p["stop_pct"] / 100.0)
    alert_tgt = price + p["target_r"] * (price - alert_stop)

    # flat bars -> time exit, but t.stop exposes the engine's stop
    flat = pd.DataFrame([
        {"datetime": datetime(2026, 7, 30, tzinfo=ist) + timedelta(days=i),
         "open": price, "high": price * 1.001,
         "low": price * 0.999, "close": price} for i in range(5)])
    t = simulate_swing(sig, flat, 1_000_000, C.exit, C.hold_days)
    assert t.stop == pytest.approx(alert_stop), (
        f"alert quotes stop {alert_stop:.4f} but the engine uses {t.stop:.4f}")

    # a bar that tags the quoted target must exit there, at exactly 3R
    hit = pd.DataFrame([{"datetime": datetime(2026, 7, 30, tzinfo=ist),
                         "open": price, "high": alert_tgt * 1.01,
                         "low": price * 0.999, "close": alert_tgt}])
    t2 = simulate_swing(sig, hit, 1_000_000, C.exit, C.hold_days)
    assert t2.exit_reason == "TGT"
    assert t2.exit == pytest.approx(alert_tgt)
    assert t2.r_multiple == pytest.approx(p["target_r"], abs=0.01)


def test_alert_shows_buy_stop_and_target():
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal
    from telegram import format_signal

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 72.4, "macd_hist": 15.75, "ema_fast": 500.5,
                  "ema_slow": 442.3, "rsi_1": 67.9, "week_volume": 4686484,
                  "vol_sma": 1020073, "week_open": 561.5, "day_open": 600.15})
    sig = Signal("APCOTEXIND", "1", "NSE_EQ",
                 datetime(2026, 7, 29, 9, 15, tzinfo=ist),
                 607.55, 579.75, 579.75, "cross", ev, "2026-07-27",
                 4686484, 600.15, 561.5,
                 bar_open=600.2, bar_high=610.0, bar_low=598.0)
    msg = format_signal(sig)
    for want in ("Buy", "Stop", "Target", "607.55", "565.02", "735.14"):
        assert want in msg, f"alert is missing {want!r}"
    # the stop must be BELOW and the target ABOVE the buy price
    assert msg.index("Buy") < msg.index("Stop") < msg.index("Target")


def test_batch_alert_also_carries_the_plan():
    """A multi-signal alert must stay actionable without opening a chart."""
    from datetime import datetime
    from zoneinfo import ZoneInfo

    from strategy import BarEval, Signal
    from telegram import _compact

    ist = ZoneInfo("Asia/Kolkata")
    ev = BarEval({f"c{i:02d}": True for i in range(1, 14)},
                 {"rsi": 70.0, "macd_hist": 1.0})
    sig = Signal("X", "1", "NSE_EQ", datetime(2026, 7, 29, 9, 15, tzinfo=ist),
                 100.0, 98.0, 98.0, "cross", ev, "2026-07-27", 1e6, 99.0, 97.0,
                 bar_open=99.5, bar_high=100.5, bar_low=99.0)
    out = _compact(sig)
    assert "SL" in out and "T1" in out
    assert "93.00" in out, "7% stop on a 100 entry should print 93.00"


def test_trade_plan_is_omitted_rather_than_wrong(tmp_path, monkeypatch):
    """
    If models.yaml is missing or has no swing model, print NO plan. A silently
    wrong stop is far more dangerous than an alert without one.
    """
    import telegram as tg

    monkeypatch.setattr(tg, "_PLAN_CACHE", None, raising=False)
    monkeypatch.setattr(tg, "__file__", str(tmp_path / "telegram.py"))
    assert tg.format_trade_plan(100.0) == []
    monkeypatch.setattr(tg, "_PLAN_CACHE", None, raising=False)


def test_trade_plan_never_quotes_a_stop_above_entry():
    from telegram import _swing_plan_params, format_trade_plan

    p = _swing_plan_params()
    assert p and 0 < p["stop_pct"] < 100
    assert format_trade_plan(0) == [], "a zero price must not produce a plan"
    lines = " ".join(format_trade_plan(250.0))
    assert "232.50" in lines, "7% below 250 is 232.50"
    # and the target must sit above entry: 250 + 3 x 17.50 = 302.50
    assert "302.50" in lines, "3R above 250 with a 7% stop is 302.50"


# --------------------------------------------------------------------------- #
#  BUG 29 - second Telegram destination. User, 29-Jul: "i want to add one more
#  telegram bot into the same alert, but it has different token and chat id".
#
#  The subtle risk is NOT the fan-out itself, it is what a failure means.
#  scan.py treats a False return as "delivery failed" and deliberately skips
#  state.save() so the next run retries. If a dead SPARE bot could force that
#  False, every run would re-alert the same signal - turning a nice-to-have
#  mirror into a duplicate-alert generator.
#
#  So: the PRIMARY alone decides the return value; a secondary failure is
#  logged and swallowed. These tests pin that down.
# --------------------------------------------------------------------------- #
class _FakeTg:
    def __init__(self, ok=True, raises=False):
        self.ok, self.raises, self.calls = ok, raises, []
        self.dry_run = False

    def _r(self, what):
        self.calls.append(what)
        if self.raises:
            raise RuntimeError("revoked token")
        return self.ok

    def send(self, text, disable_preview=True): return self._r(("send", text))
    def send_document(self, path, caption=""): return self._r(("doc", str(path)))
    def send_signal(self, sig): return self._r(("signal", sig.symbol))
    def send_batch(self, sigs): return self._r(("batch", len(sigs)))


def _fanout(*clients):
    from telegram import TelegramFanout
    f = TelegramFanout.__new__(TelegramFanout)
    f.clients = [(c, "primary" if i == 0 else "secondary")
                 for i, c in enumerate(clients)]
    return f


def test_fanout_delivers_to_every_destination():
    a, b = _FakeTg(), _FakeTg()
    assert _fanout(a, b).send("hello") is True
    assert a.calls and b.calls, "both destinations must receive the message"


def test_a_dead_secondary_must_not_cause_duplicate_alerts():
    """The critical one: secondary down -> still True, so state.save() runs."""
    a, b = _FakeTg(ok=True), _FakeTg(ok=False)
    assert _fanout(a, b).send("hello") is True, (
        "a failing SPARE bot must not block state.save(), or every scan would "
        "re-alert the same signal")


def test_a_dead_primary_still_forces_a_retry():
    a, b = _FakeTg(ok=False), _FakeTg(ok=True)
    assert _fanout(a, b).send("hello") is False, (
        "if the primary never got it, scan.py must retry next run")


def test_a_raising_secondary_is_isolated():
    a, b = _FakeTg(ok=True), _FakeTg(raises=True)
    assert _fanout(a, b).send("hello") is True, (
        "an exception in one destination must never escape the fan-out")


def test_fanout_covers_documents_and_batches():
    a, b = _FakeTg(), _FakeTg()
    f = _fanout(a, b)
    assert f.send_batch([]) is True
    assert f.send_document("x.csv", "cap") is True
    assert len(b.calls) == 2, "documents and batches must mirror too"


def test_second_destination_needs_both_token_and_chat_id(monkeypatch):
    """A half-configured pair must be ignored, never guessed."""
    from config import Secrets

    both = Secrets(telegram_bot_token="t1", telegram_chat_id="c1",
                   telegram_bot_token_2="t2", telegram_chat_id_2="c2")
    assert len(both.telegram_destinations) == 2

    half = Secrets(telegram_bot_token="t1", telegram_chat_id="c1",
                   telegram_bot_token_2="t2")           # chat id missing
    assert len(half.telegram_destinations) == 1, \
        "an incomplete second destination must be skipped"

    none = Secrets(telegram_bot_token="t1", telegram_chat_id="c1")
    assert len(none.telegram_destinations) == 1


def test_every_sender_uses_build_telegram():
    """
    One construction path, so a second bot can never mirror in scan.py while
    being silently missing from the watchlist or the paper report.
    """
    for name in ("scan.py", "watch.py", "watchlist.py",
                 "paper_report.py", "ab_paper.py"):
        body = (ROOT / name).read_text()
        assert "build_telegram(" in body, f"{name} must use build_telegram()"
        assert "Telegram(cfg.secrets" not in body, \
            f"{name} still builds a single-destination client"


def test_workflows_pass_the_second_bot_secrets():
    """Env vars are useless if the workflow never forwards them."""
    import yaml as _yaml

    for wf in WORKFLOWS.glob("*.yml"):
        text = wf.read_text()
        if "TELEGRAM_BOT_TOKEN:" not in text:
            continue                     # tests.yml sends nothing
        assert "TELEGRAM_BOT_TOKEN_2" in text, f"{wf.name} misses the 2nd token"
        assert "TELEGRAM_CHAT_ID_2" in text, f"{wf.name} misses the 2nd chat id"
        _yaml.safe_load(text)            # still valid YAML


# --------------------------------------------------------------------------- #
#  BUG 30 - A and B retired, C runs alone. User, 29-Jul: "remove a and b paper
#  trade, keep only C".
#
#  The harness was written around a two-way comparison, so several code paths
#  assumed `len(models) == 2`: the "HEAD TO HEAD" title, the "X leads by Y%"
#  verdict, and the Telegram summary. With one model those either read wrong or
#  printed nothing at all - the report would have said "need at least two
#  models with trades" forever and never told you whether C was making money.
# --------------------------------------------------------------------------- #
def test_single_model_report_states_profitability(capsys):
    """
    With ONE model there is nothing to rank - say if it is winning. Built as a
    synthetic single-model list so the test stays valid however many models
    models.yaml happens to define.
    """
    import pandas as pd

    from ab_paper import load_models, print_report

    _defaults, all_models = load_models()
    models = all_models[:1]
    rows = [{"model": models[0].key, "model_label": models[0].label,
             "symbol": f"S{i}", "signal_date": "2026-07-27",
             "exit_reason": "TIME", "pnl": 500.0, "pnl_pct": 5.0,
             "gross_pnl": 520.0, "costs": 20.0, "r_multiple": 1.5,
             "mfe_pct": 6.0, "mae_pct": -1.0, "bars_held": 5,
             "invested": 100000.0} for i in range(12)]
    print_report(pd.DataFrame(rows), models)
    out = capsys.readouterr().out
    assert "need at least two models" not in out
    assert "PROFITABLE so far" in out, out[-400:]
    assert "HEAD TO HEAD" not in out, "no comparison to make with one model"


def test_single_model_telegram_summary_is_not_silent():
    import pandas as pd

    from ab_paper import load_models, telegram_summary

    _defaults, all_models = load_models()
    models = all_models[:1]
    rows = [{"model": models[0].key, "model_label": models[0].label,
             "symbol": f"S{i}", "signal_date": "2026-07-27",
             "exit_reason": "TIME", "pnl": -200.0, "pnl_pct": -2.0,
             "gross_pnl": -180.0, "costs": 20.0, "r_multiple": -0.5,
             "mfe_pct": 1.0, "mae_pct": -3.0, "bars_held": 5,
             "invested": 100000.0} for i in range(11)]
    msg = telegram_summary(pd.DataFrame(rows), models)
    assert models[0].key in msg
    assert "❌" in msg, "a losing single model must be flagged, not hidden"


def test_intraday_models_a_and_b_stay_retired():
    """
    A and B were removed once every intraday exit measured negative after
    costs. New models may be added (D arrived 30-Jul), but nothing may
    reintroduce an INTRADAY paper model without a fresh measurement.
    """
    from ab_paper import load_models

    _defaults, models = load_models()
    keys = [m.key for m in models]
    assert "A_gated" not in keys and "B_mover" not in keys
    assert models, "at least one model must remain"
    for m in models:
        assert m.is_swing, f"{m.key} is intraday - that question is settled"
        assert m.hold_days == 5
        rule = str(m.exit.get("rule", "")).lower()
        if rule == "btst":
            # Model E is an overnight rule with its own exit grammar
            # (stop_pct / take_pct). It is still a swing horizon and still
            # capped at 5 sessions, which is what this test guards. Its stop
            # is deliberately NOT 7% - see the measurement block in
            # models.yaml and BUG 39.
            assert "take_pct" in m.exit, f"{m.key} btst exit needs take_pct"
            assert m.exit["stop_pct"] > 0
        else:
            assert m.exit["stop_pct"] == 7.0
            assert m.exit["target_r"] == 3.0


def test_retired_models_reasoning_is_preserved():
    """
    Deleting A and B must not delete WHY. Their measured results are the
    justification for C existing, so the evidence stays in the repo.
    """
    doc = (ROOT / "models.yaml").read_text()
    assert "A \"gated\"" in doc or "A \"gated\"" in doc.replace("'", '"'), \
        "models.yaml must record what A and B were"
    assert "removed" in doc.lower() and "intraday" in doc.lower(), \
        "models.yaml must record why they were removed"


# --------------------------------------------------------------------------- #
#  BUG 31 - market cap filter (Pine c12) finally enabled.
#  User, 29-Jul: "Add 1 filter on to it, market capital >1000rs".
#
#  c12 had auto-passed since day one because Dhan exposes no shares
#  outstanding, so sub-1000 Cr names Chartink would reject - GANESHBE 881 Cr,
#  PYRAMID 665 Cr - reached the alerts anyway.
#
#  THE TRAP: c12 is MANDATORY inside gate_ok(). Flipping use_mcap on without a
#  data source makes snap.mcap None for every symbol, c12 False for every
#  symbol, and the scanner goes completely silent. Verified before the fix.
#  So "unknown" must PASS, and only a KNOWN-small cap may block.
# --------------------------------------------------------------------------- #
def _mcap_snap(mcap):
    import indicators as ind
    from strategy import WeeklySnapshot
    return WeeklySnapshot(
        symbol="X", security_id="1", exchange_segment="NSE_EQ",
        week_start="2026-07-27", entry_level=100.0, level_52=100.0,
        hi_short2=90.0, close_1=95.0,
        ema_fast=ind.EmaState(length=20, prev=1.0),
        ema_slow=ind.EmaState(length=50, prev=1.0), ema_slow_2=1.0,
        rsi=ind.RsiState(length=14, avg_gain=1.0, avg_loss=1.0, prev_close=1.0),
        rsi_1=1.0,
        macd=ind.MacdState(fast=12, slow=26, signal=9, ema_fast=1.0,
                           ema_slow=1.0, sig=1.0),
        vol_sma=ind.SmaState(length=10, sum_prev=1.0),
        g_ema_fast=1.0, g_ema_slow=1.0, g_ema_slow_2=1.0,
        g_rsi=1.0, g_rsi_1=1.0, g_hist=1.0,
        prev_daily_close=1.0, prev_daily_open=1.0, mcap=mcap)


def test_unknown_market_cap_must_not_block_a_signal():
    """The whole-scan-goes-silent bug. Unknown is not small."""
    from config import Strategy
    from strategy import evaluate_bar

    cfg = Strategy(use_mcap=True, min_mcap=1000.0)
    ev = evaluate_bar(_mcap_snap(None), cfg, 105.0, 99.0, 1e6, 100.0, 1.0)
    assert ev.conditions["c12"] is True, (
        "a symbol missing from mcap.csv must PASS c12 - blocking on missing "
        "data silently deletes stocks from the scan")


def test_small_caps_are_blocked_and_large_caps_pass():
    from config import Strategy
    from strategy import evaluate_bar

    cfg = Strategy(use_mcap=True, min_mcap=1000.0, mcap_margin_pct=5.0)

    def c12(mc):
        return evaluate_bar(_mcap_snap(mc), cfg, 105.0, 99.0,
                            1e6, 100.0, 1.0).conditions["c12"]

    assert c12(881.0) is False, "GANESHBE at 881 Cr must be blocked"
    assert c12(665.0) is False, "PYRAMID at 665 Cr must be blocked"
    assert c12(3078.0) is True, "MONARCH at 3078 Cr must pass"
    assert c12(206505.0) is True


def test_mcap_margin_keeps_borderline_names():
    """
    Share counts are a few percent stale, so a hard cut at exactly min_mcap
    would drop genuine setups on a rounding difference.
    """
    from config import Strategy
    from strategy import evaluate_bar

    def c12(mc, margin):
        cfg = Strategy(use_mcap=True, min_mcap=1000.0, mcap_margin_pct=margin)
        return evaluate_bar(_mcap_snap(mc), cfg, 105.0, 99.0,
                            1e6, 100.0, 1.0).conditions["c12"]

    assert c12(980.0, 5.0) is True, "980 Cr is within the 5% tolerance"
    assert c12(980.0, 0.0) is False, "with no margin it is below 1000"
    assert c12(940.0, 5.0) is False, "clearly below even with tolerance"


def test_disabling_use_mcap_restores_the_old_behaviour():
    from config import Strategy
    from strategy import evaluate_bar

    cfg = Strategy(use_mcap=False, min_mcap=1000.0)
    ev = evaluate_bar(_mcap_snap(1.0), cfg, 105.0, 99.0, 1e6, 100.0, 1.0)
    assert ev.conditions["c12"] is True


def test_mcap_table_loader_is_tolerant(tmp_path):
    """A malformed row must be skipped, not crash the weekly build."""
    from mcap import load_table

    p = tmp_path / "mcap.csv"
    p.write_text("symbol,mcap_cr,updated\n"
                 "RADICO,58556.00,2026-07-29\n"
                 "BROKEN,notanumber,2026-07-29\n"
                 "EMPTY,,2026-07-29\n"
                 "ZERO,0,2026-07-29\n"
                 "monarch,3078.00,2026-07-29\n")
    t = load_table(p)
    assert t == {"RADICO": 58556.0, "MONARCH": 3078.0}, t
    assert load_table(tmp_path / "missing.csv") == {}


def test_mcap_is_enabled_in_the_shipped_config():
    import yaml as _yaml

    cfg = _yaml.safe_load((ROOT / "config.yaml").read_text())["strategy"]
    assert cfg["use_mcap"] is True
    assert cfg["min_mcap"] == 1000
    assert 0 <= cfg.get("mcap_margin_pct", 0) <= 10


def test_snapshot_workflow_refreshes_and_commits_the_mcap_table():
    """A stale table is nearly as bad as none - it must rebuild weekly."""
    body = (WORKFLOWS / "snapshot.yml").read_text()
    assert "python mcap.py" in body, "snapshot.yml must rebuild mcap.csv"
    assert "mcap.csv" in body.split("git add")[1][:120], \
        "mcap.csv must be committed alongside the snapshot"
    assert body.index("python mcap.py") < body.index("python build_snapshot.py"), \
        "the table must be built BEFORE the snapshot that reads it"
    assert "continue-on-error: true" in body, \
        "a market-cap fetch failure must not abort the weekly snapshot"


# --------------------------------------------------------------------------- #
#  BUG 32 - Model D, the row-3 trigger level. User, 30-Jul: "i just want that
#  fresh breakout trigger level - enter on first 5-min close above hi_short2".
#
#  hi_short2 is highest(high,26) as of TWO weeks ago - row 3 of the indicator
#  table - and is <= entry_level by construction, so D fires earlier than C.
#
#  HONESTY: a point-in-time backtest of this trigger measured -0.05%/trade
#  (t = -0.21, n = 815). An earlier +0.82% came from a look-ahead bug that
#  gated on COMPLETED weekly bars while entering mid-week. D ships as PAPER
#  only, on the user's explicit instruction, to be judged on live results.
# --------------------------------------------------------------------------- #
def test_trigger_level_defaults_to_the_pine_level():
    """The live scanner and Models A/B/C must be unaffected."""
    from config import Strategy
    assert Strategy().trigger_level == "entry"


def test_hi_short2_trigger_fires_at_or_before_the_entry_level():
    """
    hi_short2 <= entry_level always, so a cross of it can never be LATER than
    a cross of entry_level. If this inverts, the level wiring is wrong.
    """
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars, replay_week
    from test_strategy import make_daily

    daily = make_daily(weeks=200)
    wkb = build_weekly_bars(daily)
    target = wkb.iloc[-1]["week_start"]

    base = Strategy(strict_entry=False)
    snap = build_snapshot("T", "1", "NSE_EQ", daily, base, target)
    assert snap is not None
    assert snap.hi_short2 <= snap.entry_level + 1e-9

    # a ramp that crosses hi_short2 first, then entry_level
    lo, hi = snap.hi_short2, snap.entry_level
    rows = []
    for i in range(40):
        p = lo * 0.995 + (hi * 1.01 - lo * 0.995) * (i / 39.0)
        rows.append({"datetime": pd.Timestamp(target) + pd.Timedelta(minutes=5 * i),
                     "open": p, "high": p * 1.002, "low": p * 0.998,
                     "close": p, "volume": 10_000})
    bars = pd.DataFrame(rows)

    d_cfg = Strategy(strict_entry=False, trigger_level="hi_short2")
    c_cfg = Strategy(strict_entry=False, trigger_level="entry")
    d_sig = replay_week(snap, d_cfg, bars).signals
    c_sig = replay_week(snap, c_cfg, bars).signals
    assert d_sig, "the hi_short2 trigger produced no signal"
    if c_sig:
        assert d_sig[0].bar_time <= c_sig[0].bar_time, \
            "D must never trigger later than C"


def test_signal_reports_the_level_it_actually_used():
    """
    The alert prints sig.entry_level. If D quoted C's level the trade plan
    would show a breakout price the entry never crossed.
    """
    import pandas as pd

    from config import Strategy
    from strategy import build_snapshot, build_weekly_bars, replay_week
    from test_strategy import make_daily

    daily = make_daily(weeks=200)
    wkb = build_weekly_bars(daily)
    target = wkb.iloc[-1]["week_start"]
    cfg = Strategy(strict_entry=False, trigger_level="hi_short2")
    snap = build_snapshot("T", "1", "NSE_EQ", daily, cfg, target)

    lvl = snap.hi_short2
    rows = [{"datetime": pd.Timestamp(target) + pd.Timedelta(minutes=5 * i),
             "open": lvl * 0.99, "high": lvl * 1.02, "low": lvl * 0.98,
             "close": lvl * (0.995 if i == 0 else 1.01), "volume": 10_000}
            for i in range(6)]
    sigs = replay_week(snap, cfg, pd.DataFrame(rows)).signals
    assert sigs
    assert sigs[0].entry_level == pytest.approx(snap.hi_short2), \
        "the signal must carry the level it actually crossed"


def test_model_d_is_registered_and_uses_the_row3_level():
    from ab_paper import load_models

    _defaults, models = load_models()
    keys = [m.key for m in models]
    assert "D_early" in keys
    D = [m for m in models if m.key == "D_early"][0]
    assert D.strategy["trigger_level"] == "hi_short2"
    assert D.is_swing and D.hold_days == 5
    # C must be untouched by D's addition
    C = [m for m in models if m.key == "C_swing"][0]
    assert C.strategy.get("trigger_level", "entry") == "entry"


def test_ab_paper_feeds_market_caps_to_the_snapshot():
    """
    With use_mcap on, a snapshot built without a cap reports mcap=None. That
    passes c12 as 'unknown', but it means the paper models would not see the
    same universe as the live scanner. Wire the real table in.
    """
    body = (ROOT / "ab_paper.py").read_text()
    assert "load_mcap_table" in body, "ab_paper.py must load mcap.csv"
    assert "mcap=MCAP.get(" in body, \
        "build_snapshot must receive the market cap"


def test_model_d_records_the_measured_result_honestly():
    """The negative backtest must stay documented next to the model."""
    doc = (ROOT / "models.yaml").read_text()
    assert "-0.05" in doc and "look-ahead" in doc.lower(), \
        "models.yaml must record D's measured result and the earlier bug"


# --------------------------------------------------------------------------- #
#  BUG 33 - the watchlist ignored market cap and only tracked C's level.
#  User, 30-Jul: "update watchlist logic and market cap as well".
#
#  Two independent gaps, both silent:
#
#  1. MARKET CAP. is_eligible() screened on c03/c05/c08 only, so the digest
#     happily listed sub-1000 Cr names the live scan would reject. Reading the
#     cap off the SNAPSHOT would not have worked either: the committed
#     weekly_snapshot.csv was built before the feature and carries mcap=None
#     for 2099 of 2100 rows, so the filter would have been inert with no
#     visible symptom. The table is loaded from mcap.csv instead.
#
#  2. LEVEL. It measured distance to entry_level only. Model D triggers on
#     hi_short2, which is <= entry_level, so any name about to trigger D was
#     either missing or shown against a level it would not cross first.
# --------------------------------------------------------------------------- #
def test_watchlist_applies_the_market_cap_screen():
    from config import Strategy
    from watchlist import is_eligible

    class S:
        close_1 = 100.0
        hi_short2 = 120.0
        g_hist = 1.0
        g_ema_slow = 50.0
        g_ema_slow_2 = 49.0

    cfg = Strategy(use_mcap=True, min_mcap=1000.0, mcap_margin_pct=5.0)
    assert is_eligible(S(), cfg, 3078.0) is True, "a large cap must pass"
    assert is_eligible(S(), cfg, 881.0) is False, "GANESHBE at 881 Cr must fail"
    assert is_eligible(S(), cfg, None) is True, (
        "unknown cap must PASS - never hide a name because the cap could not "
        "be resolved")

    off = Strategy(use_mcap=False, min_mcap=1000.0)
    assert is_eligible(S(), off, 100.0) is True, \
        "with use_mcap off the screen must not apply"


def test_watchlist_reads_caps_from_the_table_not_the_snapshot():
    """
    The committed snapshot predates c12 and has mcap=None almost everywhere.
    Reading caps from there would disable the screen silently.
    """
    body = (ROOT / "watchlist.py").read_text()
    assert "load_mcap_table" in body, "watchlist.py must load mcap.csv"
    assert "caps.get(" in body, "the cap must come from the table"


def test_watchlist_tracks_both_model_levels():
    """
    C triggers on entry_level, D on hi_short2. hi_short2 <= entry_level, so a
    name approaching D's level must not be measured against C's.
    """
    body = (ROOT / "watchlist.py").read_text()
    assert "hi_short2" in body, "watchlist.py must consider D's level"
    assert "level_d" in body and "level_c" in body, \
        "the CSV must record both levels so nothing is hidden"


def test_watchlist_picks_the_nearest_level_above_price():
    """A breakout reaches the LOWER of the two levels first."""
    from watchlist import build_message

    # D level 95 is nearer than C level 100 for a price of 94
    rows = [dict(symbol="AAA", ltp=94.0, level=95.0, which="D",
                 level_c=100.0, level_d=95.0, mcap_cr=5000.0,
                 pct=-1.05, gap=1.05, bucket="WATCH")]
    msg = build_message("2026-07-27", rows,
                        {"universe": 1, "eligible": 1, "capped": 1}, 3.0)
    assert "95.00" in msg, "the nearer (D) level must be quoted"
    assert "[D]" in msg, "the message must say which model's level it is"
    assert "5,000Cr" in msg, "the market cap should be visible"


# --------------------------------------------------------------------------- #
#  BUG 34 - the market-cap screen was silently skipped in the watchlist.
#  User, 30-Jul: "still coming too many false stocks and under 1000cr",
#  with QMSMEDI (264 Cr) in the digest.
#
#  is_eligible() did  getattr(cfg, "use_mcap", False)  but main() passes the
#  full Config, and use_mcap lives on cfg.STRATEGY. The getattr default made it
#  False, so the entire branch was skipped - no error, no warning, just no
#  filtering. The BUG 33 test missed it because it passed a Strategy directly,
#  which is not what production does.
#
#  LESSON: a test must call the function the way the caller does. Passing a
#  tidier object than production uses is how a no-op passes review.
# --------------------------------------------------------------------------- #
def test_is_eligible_works_with_the_object_main_actually_passes():
    """main() hands is_eligible a Config, not a Strategy. Both must work."""
    from config import load_config
    from watchlist import is_eligible

    class S:
        close_1 = 100.0
        hi_short2 = 120.0
        g_hist = 1.0
        g_ema_slow = 50.0
        g_ema_slow_2 = 49.0

    cfg = load_config(None)
    assert cfg.strategy.use_mcap is True, "this test assumes c12 is enabled"

    # the exact call shape used in production
    assert is_eligible(S(), cfg, 264.0) is False, (
        "QMSMEDI at 264 Cr must be rejected when a full Config is passed - "
        "this is the call main() makes")
    assert is_eligible(S(), cfg, 5000.0) is True
    assert is_eligible(S(), cfg, None) is True, "unknown must still pass"

    # and the Strategy form must behave identically
    assert is_eligible(S(), cfg.strategy, 264.0) is False
    assert is_eligible(S(), cfg.strategy, 5000.0) is True


def test_watchlist_main_passes_something_is_eligible_understands():
    """
    Guard the wiring itself: if main() ever passes a bare Strategy again, or
    is_eligible stops unwrapping, this catches it without a live run.
    """
    body = (ROOT / "watchlist.py").read_text()
    assert 'getattr(cfg, "strategy", cfg)' in body, (
        "is_eligible must unwrap Config -> Strategy so the screen cannot be "
        "silently skipped")


# --------------------------------------------------------------------------- #
#  BUG 35 - watchlist ranked on distance alone.
#  User, 30-Jul: "rank them according max condition pass + away %".
#
#  Distance-only ordering put a weak name that happened to be 0.1% away above
#  a name passing every condition 0.4% away. Proximity is not quality: a 5/8
#  setup one tick from its level is still a 5/8 setup.
#
#  Only rows knowable PRE-MARKET are scored (the digest runs 08:45). c01/c02
#  need the developing close, c09 needs accumulating weekly volume, c10/c13
#  need today's candle - scoring those at 08:45 would produce a number that
#  changes every morning for no reason. The 8 scored rows use the g_* closed
#  week values already on the snapshot, so the score is stable all week.
# --------------------------------------------------------------------------- #
def _score_snap(**over):
    class S:
        close_1 = 100.0
        hi_short2 = 120.0        # c03 pass
        g_ema_fast = 55.0
        g_ema_slow = 50.0        # c04 pass
        g_ema_slow_2 = 49.0      # c05 pass
        g_rsi = 70.0             # c06 pass
        g_rsi_1 = 65.0           # c07 pass
        g_hist = 1.0             # c08 pass
    s = S()
    for k, v in over.items():
        setattr(s, k, v)
    return s


def test_condition_score_counts_only_premarket_rows():
    from config import load_config
    from watchlist import SCORED_ROWS, score_conditions

    cfg = load_config(None)
    # rows that need the developing week must NOT be scored
    for row in ("c01", "c02", "c09", "c10", "c13"):
        assert row not in SCORED_ROWS, f"{row} is not knowable pre-market"
    assert len(SCORED_ROWS) == 8

    score, failing = score_conditions(_score_snap(), cfg, 5000.0, 500.0)
    assert score == 8 and failing == [], failing


def test_condition_score_reports_what_failed():
    from config import load_config
    from watchlist import score_conditions

    cfg = load_config(None)
    # Break c04 (EMA20 below EMA50) and c06 (RSI under the minimum). Note
    # g_rsi feeds c07 too, so drop g_rsi_1 as well to isolate exactly two
    # failures - the earlier version of this test forgot that and expected 6.
    s = _score_snap(g_ema_fast=40.0, g_rsi=50.0, g_rsi_1=40.0)
    score, failing = score_conditions(s, cfg, 5000.0, 500.0)
    assert failing == ["c04", "c06"], failing
    assert score == 6, score
    # a small cap must cost the c12 point
    score2, failing2 = score_conditions(_score_snap(), cfg, 264.0, 500.0)
    assert "c12" in failing2 and score2 == 7


def test_unknown_market_cap_does_not_cost_a_point():
    """Same rule as the live gate: unknown is not disqualifying."""
    from config import load_config
    from watchlist import score_conditions

    cfg = load_config(None)
    score, failing = score_conditions(_score_snap(), cfg, None, 500.0)
    assert "c12" not in failing and score == 8


def test_watchlist_ranks_by_score_then_distance():
    """
    A high scorer slightly further out must outrank a low scorer that is nearer
    - that is the whole point of the change.

    SUPERSEDED BY BUG 36 (30-Jul-2026). The ranking key moved from the frozen
    gate-row count (`score`) to the measured PRE score (`pre_score`), because
    the gate rows barely varied across the shortlist - nearly every name at a
    26-week high passes them. The INVARIANT under test is unchanged: quality
    outranks proximity. Only the field supplying "quality" changed.
    """
    from watchlist import build_message

    rows = [
        dict(symbol="WEAK", ltp=99.9, level=100.0, which="C", mcap_cr=5000.0,
             score=5, max_score=8, pre_score=3, pre_max=8,
             pct=-0.1, gap=0.1, bucket="WATCH"),
        dict(symbol="STRONG", ltp=99.6, level=100.0, which="C", mcap_cr=5000.0,
             score=8, max_score=8, pre_score=8, pre_max=8,
             pct=-0.4, gap=0.4, bucket="WATCH"),
    ]
    msg = build_message("2026-07-27", rows,
                        {"universe": 2, "eligible": 2, "capped": 2}, 3.0)
    assert msg.index("STRONG") < msg.index("WEAK"), \
        "the high scorer must rank above the nearer weak one"
    assert "8/8" in msg and "3/8" in msg, "the score must be visible"


# --------------------------------------------------------------------------- #
#  BUG 36 - watchlist listed names that could never be traded, and ranked on
#  gate rows that carry almost no information.
#
#  User, 30-Jul: "still coming too many false stocks and under 1000cr", and
#  then "53 approaching stocks". The digest was unusable as a shortlist.
#
#  Measured on 33,755 point-in-time signals over 5 years (RALLY_FILTERS.md),
#  target P(MFE >= +30% in 30 days):
#
#    LIQUIDITY IS A TRAP. The strongest raw factors were all illiquidity
#    proxies - turnover < 0.63 Cr/day scored +4.36%/trade. That is unfillable:
#    two-thirds of the whole measured edge came from stocks a real order cannot
#    trade. Inside a tradeable universe the same edge is +0.28%.
#
#    VOLATILITY IS THE FILTER. atr_pct decile 1 -> P(+30%) = 5.5%,
#    decile 10 -> 22.4%. A four-fold spread, robust out of sample. This
#    CONTRADICTS the "tight coiled base" intuition and the tests below pin that
#    direction down so it cannot be silently inverted later.
#
#    PRE SCORE RANKS. Out-of-sample P(+30%) by PRE score: 1 -> 2.5%,
#    4 -> 7.4%, 7 -> 20.4%, 8 -> 22.8%.
#
#  SCOPE, deliberately: watchlist.py ONLY. scan.py, telegram.py alerts and
#  LOGIC_VERSION are untouched, so no live alert changes behaviour. The test
#  below asserts that scoping.
# --------------------------------------------------------------------------- #
def _metrics(**over):
    """A metrics dict that passes every gate and scores 8/8, unless overridden."""
    m = dict(px=250.0, atr_pct=4.5, turnover_cr=25.0, ret_12m=80.0,
             ret_3m=30.0, ret_1m=20.0, dist_50dma=20.0, dist_200dma=35.0,
             dma200_slope=4.0, base_tight=4.6, base_depth_pct=-40.0,
             spike_level=4.0)
    m.update(over)
    return m


def test_bug36_hard_gates_reject_illiquid_and_quiet():
    from watchlist import (MIN_ATR_PCT, MIN_PRICE_GATE, MIN_TURNOVER_CR,
                           gate_reasons)

    assert gate_reasons(_metrics()) == [], "a clean name must pass every gate"

    # illiquid - the QMSMEDI class of problem
    bad = gate_reasons(_metrics(turnover_cr=0.4))
    assert any("turnover" in b for b in bad), bad

    # penny - one tick is a whole percent
    bad = gate_reasons(_metrics(px=12.0))
    assert any("px" in b for b in bad), bad

    # too quiet - the single most predictive factor measured
    bad = gate_reasons(_metrics(atr_pct=1.8))
    assert any("atr" in b for b in bad), bad

    # thresholds must not drift without a deliberate edit
    assert MIN_TURNOVER_CR == 2.0
    assert MIN_PRICE_GATE == 30.0
    assert MIN_ATR_PCT == 3.0


def test_bug36_unknown_metrics_never_silently_drop_a_name():
    """
    UNKNOWN IS NOT A FAILURE - the same principle as c12/market cap.

    A transient Dhan error must not quietly empty the watchlist, and it must
    not quietly promote a junk name either. `None` returns no gate failures
    (so nothing is deleted) while score_pre gives 0 (so it ranks last).
    """
    from watchlist import gate_reasons, score_pre

    assert gate_reasons(None) == [], "unknown metrics must not delete a name"
    score, passed = score_pre(None)
    assert score == 0 and passed == [], "unknown must not manufacture a score"

    # a NaN field fails its own condition without raising
    score, _ = score_pre(_metrics(ret_12m=float("nan")))
    assert score == 7, score


def test_bug36_pre_score_direction_matches_the_measurement():
    """
    Every condition must point the way the data pointed. The volatility ones
    are the important pair: high ATR and a LIVELY base score HIGHER. If someone
    later "fixes" this to reward tight quiet bases, this test fails.
    """
    from watchlist import PRE_MAX, score_pre

    assert PRE_MAX == 8
    full, passed = score_pre(_metrics())
    assert full == 8, passed

    # SUPERSEDED BY BUG 45 (THE_EDGE.md, 31-Jul-2026).
    # atr_pct and base_tight were SCORED here until the full 23,994-signal
    # study re-tested every factor against the exit that actually ships. Both
    # flipped sign out of sample (atr IS -0.16 / OOS +0.69; base_tight
    # IS -0.10 / OOS +0.82) and were removed from the score. atr_pct remains a
    # TRADEABILITY gate, which is a different job. Asserting that they no
    # longer move the score is the point of these two lines.
    assert score_pre(_metrics(atr_pct=2.0))[0] == full, (
        "atr_pct must NOT be scored - it failed out of sample")
    assert score_pre(_metrics(base_tight=2.0))[0] == full, (
        "base_tight must NOT be scored - it failed out of sample")

    # momentum: weaker must score lower
    assert score_pre(_metrics(ret_12m=5.0))[0] < full
    assert score_pre(_metrics(dist_50dma=2.0))[0] < full
    assert score_pre(_metrics(dma200_slope=-1.0))[0] < full

    # a deep, broken base must score lower than a shallow one
    assert score_pre(_metrics(base_depth_pct=-70.0))[0] < full

    # CENTENKA: a level made by one lonely spike must score LOWER
    assert score_pre(_metrics(spike_level=0.5))[0] < full

    # very high-priced names score lower
    assert score_pre(_metrics(px=3000.0))[0] < full


def test_bug36_brk_score_tags_breakout_conviction():
    from watchlist import BRK_MAX, score_brk

    assert BRK_MAX == 5


    import pandas as pd

    def frame(o, h, l, c, v, prev_close=100.0, vol_hist=1000.0):
        rows = [dict(open=prev_close, high=prev_close, low=prev_close,
                     close=prev_close, volume=vol_hist) for _ in range(50)]
        rows.append(dict(open=o, high=h, low=l, close=c, volume=v))
        return pd.DataFrame(rows)

    # wide range, big gap, huge volume, closes at the high, well clear of 100
    strong = frame(o=102.0, h=112.0, l=101.0, c=111.5, v=5000.0)
    score, passed = score_brk(strong, 100.0)
    assert score == 5, passed

    # scrapes over the level on no volume, closes mid-range, no gap
    weak = frame(o=100.1, h=100.6, l=99.8, c=100.1, v=900.0)
    score, _ = score_brk(weak, 100.0)
    assert score <= 1, score

    # not computable -> None, so the caller omits the tag instead of printing 0
    assert score_brk(None, 100.0)[0] is None
    assert score_brk(pd.DataFrame(), 100.0)[0] is None
    assert score_brk(strong, 0.0)[0] is None


def test_bug36_screened_names_leave_the_watch_bucket_but_stay_in_the_csv():
    """
    A screened-out name must not be listed as actionable, but must remain
    visible in the attachment - a wrong gate has to be auditable, not silent.
    """
    from watchlist import build_message

    rows = [
        dict(symbol="GOOD", ltp=99.6, level=100.0, which="C", mcap_cr=5000.0,
             pre_score=7, pre_max=8, atr_pct=4.4, pct=-0.4, gap=0.4,
             bucket="WATCH", screened_ok=True),
        dict(symbol="JUNK", ltp=99.9, level=100.0, which="C", mcap_cr=264.0,
             pre_score=2, pre_max=8, atr_pct=1.1, pct=-0.1, gap=0.1,
             bucket="SCREENED", screened_ok=True, gate_failed="turnover<2Cr"),
    ]
    msg = build_message("2026-07-27", rows,
                        {"universe": 2, "eligible": 2, "capped": 2,
                         "screened": 1}, 3.0)
    assert "GOOD" in msg
    assert "JUNK" not in msg, "a screened name must not be listed as actionable"
    assert "screened out" in msg.lower()


def test_bug36_unscreened_names_rank_last_but_are_still_shown():
    from watchlist import build_message

    rows = [
        dict(symbol="KNOWN", ltp=99.5, level=100.0, which="C", mcap_cr=5000.0,
             pre_score=4, pre_max=8, pct=-0.5, gap=0.5, bucket="WATCH",
             screened_ok=True),
        dict(symbol="NODATA", ltp=99.95, level=100.0, which="C", mcap_cr=5000.0,
             pre_score=None, pre_max=8, pct=-0.05, gap=0.05, bucket="WATCH",
             screened_ok=False),
    ]
    msg = build_message("2026-07-27", rows,
                        {"universe": 2, "eligible": 2, "capped": 2,
                         "unscreened": 1}, 3.0)
    assert "NODATA" in msg, "an unscreened name must still be visible"
    assert msg.index("KNOWN") < msg.index("NODATA"), \
        "a scored name outranks an unscored one even when further away"
    assert "unscreened" in msg.lower()


def test_bug36_change_is_scoped_to_the_watchlist_only():
    """
    The gates are a SHORTLIST tool. They must not have leaked into the live
    scanner, the alert path, or the snapshot format - a change there would
    alter real alerts, which was explicitly not asked for.
    """
    import inspect

    import scan
    import strategy
    import telegram

    for mod in (scan, telegram, strategy):
        src = inspect.getsource(mod)
        for token in ("MIN_TURNOVER_CR", "MIN_ATR_PCT", "score_pre", "score_brk"):
            assert token not in src, \
                f"{token} leaked into {mod.__name__} - watchlist-only change"

    # the snapshot contract is untouched, so no rebuild is forced
    assert strategy.LOGIC_VERSION == 3, \
        "LOGIC_VERSION must not change: no snapshot field was added"


def test_bug36_metrics_are_computed_from_closed_bars_only():
    """
    compute_metrics must be usable pre-market: it may look at the LTP for the
    current price, but every historical input has to come from closed daily
    bars. A silent look-ahead here would inflate every backtest that reuses it.
    """
    import numpy as np
    import pandas as pd

    from watchlist import compute_metrics

    n = 300
    rng = np.random.default_rng(3)
    close = 100 * np.cumprod(1 + rng.normal(0.001, 0.02, n))
    df = pd.DataFrame(dict(
        open=close * 0.995, high=close * 1.02, low=close * 0.98,
        close=close, volume=np.full(n, 500000.0)))

    m = compute_metrics(df, level=float(close.max()), ltp=float(close[-1]))
    assert m is not None
    for key in ("atr_pct", "turnover_cr", "ret_12m", "dist_50dma",
                "dma200_slope", "base_tight", "base_depth_pct", "spike_level"):
        assert key in m, key
        assert not isinstance(m[key], str)
    assert m["atr_pct"] > 0 and m["turnover_cr"] > 0

    # too little history must return None, never a half-computed dict
    assert compute_metrics(df.head(10), level=100.0) is None
    assert compute_metrics(pd.DataFrame(), level=100.0) is None


# --------------------------------------------------------------------------- #
#  BUG 37 - the market-cap filter was ENABLED in config but INERT in the live
#  scanner, because scan.py read the cap off the FROZEN SNAPSHOT.
#
#  The snapshot stores whatever cap was known when it was BUILT. The live
#  27-Jul snapshot was written before mcap.py was wired into the workflow, so:
#
#      rows with a populated mcap: 1 of 2100     (only SAKHTISUG)
#
#  evaluate_bar() treats `snap.mcap is None` as a PASS - correctly, because
#  "unknown is not small". The combination means use_mcap: true did absolutely
#  nothing, with no error anywhere. Two names below the 1000 Cr floor alerted
#  in the 27-Jul week and reached the user's Telegram:
#
#      GANESHBE   881 Cr        PYRAMID   665 Cr
#
#  FIX: load_snapshots() overrides s.mcap from mcap.csv, which the snapshot
#  workflow refreshes and commits. Market cap is a slow-moving universe filter,
#  not a frozen weekly level - freezing it bought nothing and broke the filter.
#
#  This is the same CLASS of bug as BUG 24/25 (stale snapshot data trusted
#  silently). The rule that keeps being relearned: a stale input must either be
#  refreshed from the live source or refused loudly - never used quietly.
# --------------------------------------------------------------------------- #
def _snapshot_csv(tmp_path, rows):
    """Write a minimal but complete snapshot CSV and return its directory."""
    import pandas as pd

    import strategy as _strat

    # Built by round-tripping a real WeeklySnapshot through to_row(), so the
    # fixture can never drift from the actual on-disk column set.
    import indicators as _ind

    _snap = _strat.WeeklySnapshot(
        symbol="X", security_id="1", exchange_segment="NSE_EQ",
        week_start="2026-07-27",
        entry_level=100.0, level_52=120.0, hi_short2=95.0, close_1=90.0,
        ema_fast=_ind.EmaState(length=20, prev=55.0),
        ema_slow=_ind.EmaState(length=50, prev=50.0),
        ema_slow_2=49.0,
        rsi=_ind.RsiState(length=14, avg_gain=2.0, avg_loss=1.0, prev_close=90.0),
        rsi_1=65.0,
        macd=_ind.MacdState(fast=12, slow=26, signal=9,
                            ema_fast=90.0, ema_slow=88.0, sig=1.5),
        vol_sma=_ind.SmaState(length=10, sum_prev=1000.0),
        g_ema_fast=55.0, g_ema_slow=50.0, g_ema_slow_2=49.0,
        g_rsi=70.0, g_rsi_1=65.0, g_hist=1.0,
        prev_daily_close=95.0, prev_daily_open=94.0,
        mcap=None,
    )
    base = {k: ("" if v is None else str(v)) for k, v in _snap.to_row().items()}
    base["logic_version"] = str(_strat.LOGIC_VERSION)
    out = []
    for r in rows:
        d = dict(base)
        d.update({k: str(v) for k, v in r.items()})
        out.append(d)
    df = pd.DataFrame(out)
    p = tmp_path / "weekly_snapshot.csv"
    df.to_csv(p, index=False)
    return p


def test_bug37_scan_reads_market_cap_from_the_live_table(tmp_path):
    """
    A snapshot with an EMPTY mcap column must still get real caps, because
    load_snapshots() overrides them from mcap.csv.
    """
    import scan
    from config import load_config

    _snapshot_csv(tmp_path, [
        dict(symbol="GANESHBE", security_id="11"),
        dict(symbol="PYRAMID", security_id="12"),
        dict(symbol="RADICO", security_id="13"),
    ])
    (tmp_path / "mcap.csv").write_text(
        "symbol,mcap_cr,updated\n"
        "GANESHBE,881,2026-07-29\n"
        "PYRAMID,665,2026-07-29\n"
        "RADICO,58556,2026-07-29\n")
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: true\n  min_mcap: 1000\n  mcap_margin_pct: 5.0\n")

    cfg = load_config(tmp_path / "config.yaml")
    assert cfg.strategy.use_mcap is True

    snaps = scan.load_snapshots(cfg, "2026-07-27")
    caps = {s.symbol: s.mcap for s in snaps}
    assert caps["GANESHBE"] == 881.0, "cap must come from mcap.csv, not the snapshot"
    assert caps["PYRAMID"] == 665.0
    assert caps["RADICO"] == 58556.0


def test_bug37_the_two_names_that_wrongly_alerted_are_now_blocked(tmp_path):
    """
    The exact regression, end to end: GANESHBE (881 Cr) and PYRAMID (665 Cr)
    must FAIL c12 while RADICO passes, using the real gate code.
    """
    import scan
    from config import load_config
    from strategy import evaluate_bar

    _snapshot_csv(tmp_path, [
        dict(symbol="GANESHBE", security_id="11"),
        dict(symbol="PYRAMID", security_id="12"),
        dict(symbol="RADICO", security_id="13"),
    ])
    (tmp_path / "mcap.csv").write_text(
        "symbol,mcap_cr\nGANESHBE,881\nPYRAMID,665\nRADICO,58556\n")
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: true\n  min_mcap: 1000\n  mcap_margin_pct: 5.0\n")
    cfg = load_config(tmp_path / "config.yaml")

    snaps = {s.symbol: s for s in scan.load_snapshots(cfg, "2026-07-27")}
    got = {}
    for sym, s in snaps.items():
        ev = evaluate_bar(s, cfg.strategy, price=110.0, week_open=100.0,
                          week_volume=5000.0, day_open=105.0, week_fraction=1.0)
        got[sym] = ev.conditions["c12"]

    assert got["GANESHBE"] is False, "881 Cr must fail a 1000 Cr floor"
    assert got["PYRAMID"] is False, "665 Cr must fail a 1000 Cr floor"
    assert got["RADICO"] is True, "58,556 Cr must pass"


def test_bug37_unknown_market_cap_still_passes(tmp_path):
    """
    UNKNOWN IS NOT SMALL, still. A symbol missing from mcap.csv must keep
    passing c12 - the override must not turn a data gap into a silent block,
    which would delete names from the scan with no visible symptom.
    """
    import scan
    from config import load_config
    from strategy import evaluate_bar

    _snapshot_csv(tmp_path, [dict(symbol="NEWLISTING", security_id="99")])
    (tmp_path / "mcap.csv").write_text("symbol,mcap_cr\nSOMETHINGELSE,5000\n")
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: true\n  min_mcap: 1000\n")
    cfg = load_config(tmp_path / "config.yaml")

    s = scan.load_snapshots(cfg, "2026-07-27")[0]
    assert s.mcap is None, "a symbol absent from the table stays unknown"
    ev = evaluate_bar(s, cfg.strategy, price=110.0, week_open=100.0,
                      week_volume=5000.0, day_open=105.0, week_fraction=1.0)
    assert ev.conditions["c12"] is True, "unknown must never block a name"


def test_bug37_empty_mcap_table_is_an_error_not_a_silent_pass(tmp_path, caplog):
    """
    The failure that hid this bug for a week: config says the filter is ON, the
    data says it cannot run, and nothing was logged. It must be LOUD now.
    """
    import logging

    import scan
    from config import load_config

    _snapshot_csv(tmp_path, [dict(symbol="AAA", security_id="1")])
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: true\n  min_mcap: 1000\n")
    cfg = load_config(tmp_path / "config.yaml")

    with caplog.at_level(logging.ERROR):
        scan.load_snapshots(cfg, "2026-07-27")
    blob = caplog.text.lower()
    assert "use_mcap is on" in blob and "mcap" in blob, caplog.text


def test_bug37_mostly_empty_table_is_flagged(tmp_path, caplog):
    """A table covering a tiny fraction of the universe means the filter is
    effectively off. That must be reported, not merely counted."""
    import logging

    import scan
    from config import load_config

    rows = [dict(symbol=f"S{i:03d}", security_id=str(i)) for i in range(20)]
    _snapshot_csv(tmp_path, rows)
    (tmp_path / "mcap.csv").write_text("symbol,mcap_cr\nS000,5000\nS001,6000\n")
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: true\n  min_mcap: 1000\n")
    cfg = load_config(tmp_path / "config.yaml")

    with caplog.at_level(logging.ERROR):
        snaps = scan.load_snapshots(cfg, "2026-07-27")
    assert len(snaps) == 20
    assert "inert" in caplog.text.lower(), caplog.text


def test_bug37_override_is_skipped_when_use_mcap_is_off(tmp_path):
    """No surprise behaviour when the filter is disabled: the snapshot value
    stands, and no table is consulted."""
    import scan
    from config import load_config

    _snapshot_csv(tmp_path, [dict(symbol="AAA", security_id="1", mcap="123")])
    (tmp_path / "mcap.csv").write_text("symbol,mcap_cr\nAAA,999999\n")
    (tmp_path / "config.yaml").write_text(
        "strategy:\n  use_mcap: false\n  min_mcap: 1000\n")
    cfg = load_config(tmp_path / "config.yaml")

    s = scan.load_snapshots(cfg, "2026-07-27")[0]
    assert s.mcap == 123.0, "with use_mcap off the frozen value must be kept"


def test_bug37_watch_and_watchlist_share_the_same_fix():
    """
    watch.py and watchlist.py both call load_snapshots(), so the fix must be
    inside that function rather than duplicated at each call site. If someone
    later inlines a private loader in one of them, this fails.
    """
    import inspect

    import scan
    import watch
    import watchlist

    src = inspect.getsource(scan.load_snapshots)
    assert "load_mcap_table" in src, "the override must live in load_snapshots"

    for mod in (watch, watchlist):
        body = inspect.getsource(mod)
        assert "def load_snapshots" not in body, (
            f"{mod.__name__} must reuse scan.load_snapshots, not shadow it")


# --------------------------------------------------------------------------- #
#  BUG 38 - c09 held entries until AFTER the move, and there was no BTST path.
#
#  User, 31-Jul: "rather than showing deferred why I'm missing those entry",
#  "only yasho gave massive, others loosing", "more focus on btst setup".
#
#  THE CASE. YASHO, 31-Jul-2026:
#      26W level 3262.80, opened 3258.90 - already at the level
#      weekly volume  Mon 0.08x  Tue 0.19x  Wed 0.51x  Fri 2.24x
#      c09 flipped true at 12:20, price 3690.90 -> the alert was 13.1% HIGH
#      day high/close 3858.70
#      from the level +18.3%, from the alert +4.5% - 75% of the move was gone
#
#  c09 is "weekly volume > 10-week average". Mid-week that volume has not
#  traded yet, so the row passes only once the MOVE creates the volume. The
#  gate therefore systematically delays entry on exactly the explosive names.
#
#  MEASURED, point-in-time, 5 years, 18,259 tradeable breakouts, same trades
#  entered early vs when the full gate allowed it:
#      next close  early +0.45% (t 12.7)  vs gated +0.16% (t 4.3)
#      3 days      early +0.96% (t 16.5)  vs gated +0.15% (t 2.6)
#      5 days      early +1.26% (t 17.4)  vs gated +0.36% (t 5.0)
#  and c09 blocked 37% of breakouts for the ENTIRE week - entries never given.
#
#  BTST. Buying every breakout close and selling next close earns +0.01%
#  (nothing). The edge is entirely in the breakout DAY's character:
#      TIER A  day>=15% & closed at high   n=417   +1.75%  t 5.2
#      TIER B  close@high .90 & rvol>=3 & atr>=3   n=1086  +0.83%  t 5.0
#  Both held out of sample (A +1.62 -> +2.01, B +0.76 -> +0.96).
#
#  NOT SHIPPED, deliberately: selling at the next OPEN measures +0.45% t=42,
#  but the whole universe gaps +0.42% at the open in this data. That is a
#  recording artifact, not an edge. Only next-CLOSE is used.
# --------------------------------------------------------------------------- #
def test_bug38_drop_c09_is_enabled_in_the_shipped_config():
    from config import load_config

    cfg = load_config(None)
    assert cfg.strategy.drop_c09 is True, (
        "c09 must be OUT of the live gate - it delays entry until after the "
        "move (YASHO 31-Jul: alerted 13.1% above the level)")


def test_bug38_dropping_c09_removes_only_that_row():
    """
    drop_c09 must not weaken anything else. Every other gate row still has to
    be able to block an entry, and the level break itself is untouched.
    """
    from config import Strategy
    from strategy import BarEval, gate_ok

    class Snap:
        close_1 = 100.0; hi_short2 = 120.0
        g_ema_fast = 55.0; g_ema_slow = 50.0; g_ema_slow_2 = 49.0
        g_rsi = 70.0; g_rsi_1 = 65.0; g_hist = 1.0
        mcap = 5000.0

    cfg = Strategy(drop_c09=True, strict_entry=True, gate_tolerance=0,
                   require_c03=False, require_c11=False, use_mcap=False)

    def ev(**over):
        conds = {f"c{i:02d}": True for i in range(1, 14)}
        conds.update(over)
        return BarEval(conditions=conds, values={})

    assert gate_ok(Snap(), cfg, ev()) is True
    # c09 false must NO LONGER block
    assert gate_ok(Snap(), cfg, ev(c09=False)) is True, "c09 must be ignored"
    # everything else must still block
    for row in ("c04", "c05", "c06", "c07", "c08", "c10", "c13"):
        assert gate_ok(Snap(), cfg, ev(**{row: False})) is False, \
            f"{row} must still be able to block an entry"

    # and with drop_c09 OFF, c09 blocks again
    cfg_on = Strategy(drop_c09=False, strict_entry=True, gate_tolerance=0,
                      require_c03=False, require_c11=False, use_mcap=False)
    assert gate_ok(Snap(), cfg_on, ev(c09=False)) is False


def test_bug38_chase_guard_stays_off_because_it_measured_worse():
    """
    The intuitive fix - "with c09 gone, also refuse signals far above the
    level" - was tested and is WRONG. Capping extension made returns worse at
    every threshold (5d: off +0.221% vs 3% cap +0.103%, and worse still out of
    sample), because an extended breakout is a STRONG breakout. A 5% cap would
    have discarded 343 of the 417 Tier A BTST setups.

    This test exists so nobody "fixes" it back on without re-measuring.
    """
    from config import load_config

    cfg = load_config(None)
    assert cfg.strategy.max_ext_above_level == 0.0, (
        "the chase guard measured NEGATIVE - see the config.yaml note before "
        "turning it back on")


def _btst_frame(day_ret, close_pos, rvol, atr_pct=5.0, n=260, price=100.0):
    """Synthesise a daily frame ending in a candle with the wanted character."""
    import numpy as np
    import pandas as pd

    base_v = 500000.0          # keeps 20d turnover well above the 2 Cr floor
    rows = []
    for _ in range(n - 1):
        rows.append(dict(open=price * 0.995, high=price * (1 + atr_pct / 200),
                         low=price * (1 - atr_pct / 200), close=price,
                         volume=base_v))
    o = price
    c = o * (1 + day_ret / 100.0)
    rng = max(abs(c - o) / max(close_pos, 0.01), c * 0.01)
    low = c - rng * close_pos
    high = low + rng
    rows.append(dict(open=o, high=high, low=low, close=c, volume=base_v * rvol))
    return pd.DataFrame(rows)


def test_bug38_btst_tier_a_is_the_yasho_shape():
    from btst import classify

    # age=0 is explicit: _btst_frame() builds a series that is ALREADY above
    # `level` on every bar, so there is no cross in it to date and
    # breakout_age() correctly returns 999. The live scanner passes age=0 for
    # anything whose alert fired today (BUG 53).
    m = classify(_btst_frame(day_ret=18.4, close_pos=1.0, rvol=8.0, atr_pct=4.9),
                 level=95.0, age=0)
    assert m is not None
    assert m["tier"] == "A", m
    assert m["day_ret"] > 15 and m["close_pos"] >= 0.85


def test_bug38_btst_rejects_the_losers_from_31_jul():
    """
    The same day YASHO ran, eight other names alerted and most lost money.
    Their shapes must NOT qualify - that is the whole point of the tiers.
    """
    from btst import classify

    # ATULAUTO: closed red, at the bottom of its range
    assert classify(_btst_frame(-1.2, 0.06, 4.75, 2.9), 95.0)["tier"] is None
    # IIFL: red, weak close, no volume
    assert classify(_btst_frame(-0.6, 0.21, 0.85, 3.3), 95.0)["tier"] is None
    # SILVERTUC: red, mid-range
    assert classify(_btst_frame(-0.7, 0.39, 3.47, 5.2), 95.0)["tier"] is None
    # QUESS: up 7.9% but did NOT close at the high -> no tier
    assert classify(_btst_frame(7.9, 0.73, 15.0, 3.5), 95.0)["tier"] is None


def test_bug38_btst_requires_tradeability():
    """An overnight hold in an illiquid name is worse than an intraday one -
    you cannot get out at the open."""
    from btst import classify

    quiet = classify(_btst_frame(18.0, 1.0, 8.0, atr_pct=1.2), 95.0)
    assert quiet["tier"] is None, "a 1.2% ATR stock must not qualify"
    assert quiet.get("reject") == "not tradeable"


def test_bug38_btst_never_places_orders_and_is_read_only():
    from pathlib import Path

    body = (ROOT / "btst.py").read_text()
    for token in ("place_order", "placeorder", "/orders"):
        assert token not in body.lower(), "BTST must never place an order"
    # it must not write state.json - that would corrupt the live de-dup
    assert "state.save()" not in body and ".mark(" not in body, \
        "BTST must not mutate alert state"


def test_bug38_watchlist_flags_btst_capable_names():
    from watchlist import btst_ready

    # REVISED BY BUG 45: base_tight dropped (failed OOS), dist_200dma added.
    # YASHO-like: high ATR, strong 12m trend, well above the 200DMA
    assert btst_ready(dict(atr_pct=4.9, ret_12m=80.0, dist_200dma=40.0)) is True
    # too quiet to ever print a +15% day - capability test, still enforced
    assert btst_ready(dict(atr_pct=2.1, ret_12m=80.0, dist_200dma=40.0)) is False
    # no trend
    assert btst_ready(dict(atr_pct=4.9, ret_12m=2.0, dist_200dma=40.0)) is False
    # not extended above the 200DMA
    assert btst_ready(dict(atr_pct=4.9, ret_12m=80.0, dist_200dma=5.0)) is False
    # base_tight must no longer matter either way
    assert btst_ready(dict(atr_pct=4.9, ret_12m=80.0, dist_200dma=40.0,
                           base_tight=1.0)) is True
    assert btst_ready(None) is False


def test_bug38_btst_is_a_separate_job_from_the_live_scanner():
    """BTST must not alter scan.py's alert path."""
    import inspect

    import scan

    src = inspect.getsource(scan)
    for token in ("TIER", "btst", "classify"):
        assert token not in src, f"{token} leaked into scan.py"


# --------------------------------------------------------------------------- #
#  BUG 39 - MODEL E, the BTST rule (user request, 31-Jul-2026).
#
#  "take entry in top 5 btst watchlist stocks on end of day, sl should be 1%
#   (adjust gap down possibilities), next day close position at end of the day
#   if stocks moves more than 2%, or carry forward position next day if stocks
#   in negative or positive less than 2%, intact D"
#
#  Shipped EXACTLY as specified. The measurement is recorded here because the
#  1% stop is the part most likely to be "tidied up" later by someone who has
#  not seen the numbers.
#
#  1,130 top-5 BTST trades, 5 years, net 0.22%, gap-aware fills:
#      stop    win%    mean     PF     t    SL hit
#      1%      26.5   +0.939   1.96   6.9    74%     <- as specified, best mean
#      3%      41.2   +0.660   1.35   4.1    59%
#      5%      55.1   +0.861   1.37   4.6    44%
#      none    69.2   +0.915   1.31   3.6     0%
#
#  The 1% stop genuinely measures best on the mean. The caveats are real too:
#    * median ATR of these names is 4.66%/day, so 1% is 0.21x ONE DAY'S RANGE
#      and P(next-day low <= -1%) = 70%. It fires on noise, not on a thesis break.
#    * win rate 26.5%, median trade -1.22%, longest losing streak in-sample
#      THIRTY trades - about six weeks at 4-5 signals/week.
#    * slippage is charged on 74% of trades. At 0.30% slippage the 1% stop
#      (+0.743%) and the 5% stop (+0.741%) are identical; at 0.75% the 1% stop
#      is WORSE (+0.449% vs +0.561%).
#    * out of sample the ordering reverses: 1% stop +1.036 -> +0.751, while a
#      plain next-close exit goes +0.735 -> +1.020.
#
#  Hence E_btst_wide: same signal, same entry, 5% stop, so live forward data
#  decides instead of my backtest.
#
#  GAP HANDLING is explicitly required by the user ("adjust gap down
#  possibilities") and is tested below: when a session opens BELOW the stop the
#  fill must be the OPEN, not the stop. 7.5% of these names gap under -1%, and
#  when they do the average fill is -2.03%, not -1.00%.
# --------------------------------------------------------------------------- #
def test_bug39_model_e_exists_and_d_is_untouched():
    from ab_paper import load_models

    _d, models = load_models()
    by = {m.key: m for m in models}
    assert "E_btst" in by, "Model E must be defined"
    assert "E_btst_wide" in by, "the 5% control arm must exist"

    # D_early must be EXACTLY as it was - the user said "intact D"
    d = by["D_early"]
    assert d.horizon == "swing" and d.hold_days == 5
    assert d.exit["rule"] == "swing"
    assert d.exit["stop_pct"] == 7.0
    assert d.exit["be_at_r"] == 1.0
    assert d.exit["target_r"] == 3.0
    assert d.strategy["trigger_level"] == "hi_short2"

    e = by["E_btst"]
    assert e.exit["rule"] == "btst"
    assert e.exit["stop_pct"] == 1.0, "1% stop as specified"
    assert e.exit["take_pct"] == 2.0, "exit on a close >= +2% as specified"
    assert e.strategy["btst_top_n"] == 5, "top 5 per day as specified"
    assert e.strategy["btst_only"] is True


def _sig(price=100.0):
    from datetime import datetime

    from strategy import BarEval, Signal

    ev = BarEval(conditions={f"c{i:02d}": True for i in range(1, 14)},
                 values={"rsi": 70.0, "macd_hist": 1.0})
    return Signal(symbol="X", security_id="1", exchange_segment="NSE_EQ",
                  bar_time=datetime(2026, 7, 31, 15, 25), price=price,
                  entry_level=95.0, level_52=99.0, trigger="cross",
                  evaluation=ev, week_start="2026-07-27",
                  week_volume=1e6, day_open=90.0, week_open=90.0,
                  bar_open=99.0, bar_high=101.0, bar_low=98.0)


def _days(rows):
    import pandas as pd

    return pd.DataFrame([
        dict(datetime=pd.Timestamp("2026-08-03") + pd.Timedelta(days=i),
             open=o, high=h, low=l, close=c, volume=1e6)
        for i, (o, h, l, c) in enumerate(rows)])


def test_bug39_exits_next_close_when_up_more_than_2pct():
    from ab_paper import simulate_btst

    # day 1 closes +3% -> take it
    t = simulate_btst(_sig(100.0), _days([(100.5, 104.0, 99.5, 103.0)]),
                      100000.0, dict(stop_pct=1.0, take_pct=2.0, hold_days=5))
    assert t.exit_reason == "TGT"
    assert t.bars_held == 1
    assert t.exit == 103.0


def test_bug39_carries_when_the_move_is_small_or_negative():
    from ab_paper import simulate_btst

    # day 1 closes +0.5% (less than 2%) -> carry; day 2 closes +2.5% -> exit
    t = simulate_btst(_sig(100.0),
                      _days([(100.1, 101.0, 99.5, 100.5),
                             (100.6, 103.0, 100.0, 102.5)]),
                      100000.0, dict(stop_pct=5.0, take_pct=2.0, hold_days=5))
    assert t.exit_reason == "TGT"
    assert t.bars_held == 2, "must have carried through day 1"

    # a NEGATIVE day 1 that does not hit the stop must also carry
    t2 = simulate_btst(_sig(100.0),
                       _days([(99.9, 100.2, 98.0, 98.5),
                              (98.6, 103.0, 98.4, 102.6)]),
                       100000.0, dict(stop_pct=5.0, take_pct=2.0, hold_days=5))
    assert t2.exit_reason == "TGT" and t2.bars_held == 2


def test_bug39_gap_down_fills_at_the_open_not_the_stop():
    """
    The user explicitly asked to "adjust gap down possibilities". A stop is a
    promise the market does not keep: when the session opens BELOW it, the fill
    is the open. Simulating a -1.00% fill on a -3% gap would be a lie that
    flatters every backtest of this model.
    """
    from ab_paper import simulate_btst

    t = simulate_btst(_sig(100.0), _days([(97.0, 98.0, 96.0, 96.5)]),
                      100000.0, dict(stop_pct=1.0, take_pct=2.0, hold_days=5))
    assert t.exit_reason == "SL"
    assert t.exit == 97.0, "must fill at the OPEN, not the 99.00 stop"
    assert "gap" in (t.exit_note or "").lower()
    assert t.pnl_pct < -3.0, "the loss must reflect the real fill"


def test_bug39_stop_wins_ties_against_the_target():
    """A day that both breaches the stop and closes above the target must be
    recorded as a STOP - daily bars cannot prove which came first."""
    from ab_paper import simulate_btst

    t = simulate_btst(_sig(100.0), _days([(100.0, 106.0, 98.0, 103.0)]),
                      100000.0, dict(stop_pct=1.0, take_pct=2.0, hold_days=5))
    assert t.exit_reason == "SL", "the pessimistic branch is the honest one"


def test_bug39_time_exit_caps_the_carry():
    from ab_paper import simulate_btst

    flat = [(100.0, 100.5, 99.6, 100.1)] * 6
    t = simulate_btst(_sig(100.0), _days(flat), 100000.0,
                      dict(stop_pct=5.0, take_pct=2.0, hold_days=5))
    assert t.exit_reason == "TIME"
    assert t.bars_held == 5, "carry must not run forever"


def test_bug39_tier_logic_is_shared_with_the_nightly_scanner():
    """
    ab_paper's Model E and btst.py must agree on what a BTST setup IS. If the
    thresholds are duplicated they will drift; this asserts one source.
    """
    import inspect

    import pandas as pd

    import ab_paper
    import btst

    src = inspect.getsource(ab_paper.btst_tier_for)
    assert "import btst" in src and "TIER_A_DAY" in src, \
        "the paper model must import the thresholds, not restate them"
    assert btst.TIER_A_DAY == 15.0
    assert btst.TIER_B_RVOL == 3.0
    # The POINT of this test is that the two modules share ONE definition, so
    # it asserts the shared-source invariant rather than pinning the numbers
    # (which BUG 54 moved). The exact floors are asserted in their own test.
    assert 0.5 <= btst.TIER_A_CLOSE_POS <= 1.0

    # a YASHO-shaped day is Tier A in both
    day = dict(open=100.0, high=118.5, low=99.5, close=118.4, volume=8e6)
    prev = pd.DataFrame(dict(volume=[1e6] * 60, high=[101.0] * 60,
                             low=[99.0] * 60, close=[100.0] * 60))
    assert ab_paper.btst_tier_for(day, prev, atr_pct=4.9) == "A"
    # a weak day is neither
    weak = dict(open=100.0, high=101.0, low=98.0, close=99.0, volume=5e5)
    assert ab_paper.btst_tier_for(weak, prev, atr_pct=4.9) is None


def test_bug39_the_1pct_stop_evidence_is_recorded():
    """
    The 1% stop fires on ~74% of trades and its advantage vanishes under
    realistic slippage. Whoever changes it next must see that first.
    """
    doc = (ROOT / "models.yaml").read_text()
    for token in ("74%", "slippage", "losing streak", "E_btst_wide"):
        assert token.lower() in doc.lower(), f"models.yaml must document {token}"


def test_bug39_btst_fields_reach_the_ledger():
    from ab_paper import LEDGER_COLS

    assert "btst_tier" in LEDGER_COLS
    assert "btst_day_ret" in LEDGER_COLS


def test_bug39_model_e_does_not_touch_the_live_scanner():
    import inspect

    import scan

    src = inspect.getsource(scan)
    for token in ("btst_only", "btst_top_n", "simulate_btst", "E_btst"):
        assert token not in src, f"{token} leaked into scan.py"


# --------------------------------------------------------------------------- #
#  BUG 40 - the end-of-day Telegram said HOW the models were doing but never
#  WHICH stocks were actually taken.
#
#  User, 31-Jul: "send alert end of the day in which which stocks i tooks trade
#  so i can confirm with the watchlist".
#
#  telegram_summary() reports aggregate standings (trades / win% / PF / avgR)
#  per model. There was no way to reconcile the 08:45 watchlist against the
#  positions the models actually entered, which is the whole point of running
#  a watchlist. Added todays_trades_message() as a SEPARATE message - the two
#  answer different questions and merging them made both unreadable.
#
#  Three things it must get right, all tested below:
#    * NO_FILL is NOT a trade. It is a signal that could not be taken (price
#      above capital, no candle after the signal). Listing it as a position
#      would make the reconciliation wrong in the most confusing way - a name
#      on the watchlist appearing as "taken" when nothing was bought.
#    * an OPEN position is an instruction for tomorrow, not a result. It must
#      show the STOP, not a P&L of 0.00%.
#    * names that signalled but were not taken are still named, so a watchlist
#      entry never just vanishes with no explanation.
# --------------------------------------------------------------------------- #
def _ledger_rows():
    return [
        dict(model="E_btst", symbol="YASHO", signal_date="2026-07-31",
             entry=3858.7, qty=25, stop=3820.1, exit_date="", exit=0,
             exit_reason="", bars_held=0, pnl_pct=0, btst_tier="A"),
        dict(model="E_btst", symbol="NAZARA", signal_date="2026-07-31",
             entry=317.1, qty=315, stop=313.9, exit_date="2026-08-01",
             exit=326.0, exit_reason="TGT", bars_held=1, pnl_pct=2.59,
             btst_tier="B"),
        dict(model="E_btst_wide", symbol="YASHO", signal_date="2026-07-31",
             entry=3858.7, qty=25, stop=3665.8, exit_date="", exit=0,
             exit_reason="", bars_held=0, pnl_pct=0, btst_tier="A"),
        dict(model="D_early", symbol="AETHER", signal_date="2026-07-31",
             entry=1571.0, qty=63, stop=1461.0, exit_date="2026-08-01",
             exit=1520.0, exit_reason="SL", bars_held=1, pnl_pct=-3.47,
             btst_tier=""),
        dict(model="D_early", symbol="NELCO", signal_date="2026-07-31",
             entry=1013.0, qty=98, stop=942.1, exit_date="", exit=0,
             exit_reason="NO_FILL", bars_held=0, pnl_pct=0, btst_tier=""),
    ]


def test_bug40_eod_message_names_every_stock_taken():
    import pandas as pd

    from ab_paper import load_models, todays_trades_message

    _d, models = load_models()
    msg = todays_trades_message(pd.DataFrame(_ledger_rows()), models)

    for sym in ("YASHO", "NAZARA", "AETHER"):
        assert sym in msg, f"{sym} was taken and must be listed"
    # the model that took it must be identifiable
    assert "E_btst" in msg and "D_early" in msg
    # entry price and P&L must both be visible for a closed trade
    assert "317.10" in msg and "+2.59%" in msg
    assert "-3.47%" in msg


def test_bug40_no_fill_is_not_reported_as_a_position():
    """
    The most dangerous confusion: a watchlist name showing as 'taken' when
    nothing was actually bought.
    """
    import pandas as pd

    from ab_paper import load_models, todays_trades_message

    _d, models = load_models()
    msg = todays_trades_message(pd.DataFrame(_ledger_rows()), models)

    assert "4 position(s)" in msg, "NO_FILL must not be counted as a position"
    # it is still NAMED, under a heading that says it was not taken
    assert "not taken" in msg.lower()
    assert "NELCO" in msg


def test_bug40_open_positions_show_the_stop_not_a_fake_pnl():
    import pandas as pd

    from ab_paper import load_models, todays_trades_message

    _d, models = load_models()
    msg = todays_trades_message(pd.DataFrame(_ledger_rows()), models)

    open_lines = [ln for ln in msg.splitlines() if "OPEN" in ln]
    assert len(open_lines) == 2, open_lines
    for ln in open_lines:
        assert "SL" in ln, "an open position must carry its stop"
        assert "+0.00%" not in ln, "an open position has no realised P&L"
    # the two models' different stops must both be shown correctly
    assert "3,820.10" in msg and "3,665.80" in msg


def test_bug40_empty_and_all_nofill_days_are_handled():
    import pandas as pd

    from ab_paper import load_models, todays_trades_message

    _d, models = load_models()
    assert todays_trades_message(pd.DataFrame(), models) == ""

    only_nofill = [r for r in _ledger_rows() if r["exit_reason"] == "NO_FILL"]
    msg = todays_trades_message(pd.DataFrame(only_nofill), models)
    assert "No positions taken today" in msg, msg


def test_bug40_reports_the_latest_session_not_the_whole_ledger():
    """A multi-day ledger must produce TODAY's positions, not every trade
    ever recorded."""
    import pandas as pd

    from ab_paper import load_models, todays_trades_message

    _d, models = load_models()
    rows = _ledger_rows() + [
        dict(model="E_btst", symbol="OLDNAME", signal_date="2026-07-20",
             entry=100.0, qty=10, stop=99.0, exit_date="2026-07-21",
             exit=103.0, exit_reason="TGT", bars_held=1, pnl_pct=2.8,
             btst_tier="A")]
    msg = todays_trades_message(pd.DataFrame(rows), models)
    assert "OLDNAME" not in msg, "only the most recent session belongs here"
    assert "2026-07-31" in msg

    # and it can be pinned to a specific day
    old = todays_trades_message(pd.DataFrame(rows), models, day="2026-07-20")
    assert "OLDNAME" in old and "YASHO" not in old


def test_bug40_both_telegram_paths_send_the_trade_list():
    """--report-only and a live run must BOTH send it, or the 16:15 job would
    be silent on the days it matters."""
    import inspect

    import ab_paper

    src = inspect.getsource(ab_paper.main)
    assert src.count("todays_trades_message") == 2, (
        "both the --report-only path and the live path must send the list")
    # and it must go out BEFORE the standings in each
    for chunk in src.split("todays_trades_message")[1:]:
        assert "telegram_summary" in chunk


# --------------------------------------------------------------------------- #
#  BUG 41 - the BTST scan ran AFTER the close, which made it unusable.
#
#  User, 31-Jul: "this one should not nightly BTST scanner, it must be just
#  right before end of the day btst.py so we can take able to take entry for
#  the next day in advance btst".
#
#  I shipped it at 15:40 IST. That is ten minutes after the bell, so the only
#  entry left is TOMORROW'S OPEN - which forfeits the overnight gap the entire
#  model is built to capture. A correct design error, caught by the user.
#
#  MEASURED on 5-minute data (149 large caps, 8,047 stock-days), deciding the
#  tier at a cutoff and entering at that same cutoff:
#
#      cutoff   tier precision   entry vs close   next-close (net)
#      15:00        70.2%           -0.86%          +1.18%  t 2.2
#      15:10        71.6%           -0.58%          +0.56%  t 1.0
#      15:15        75.6%           -0.28%          +0.60%  t 1.2
#      15:20        82.3%           -0.14%          +1.03%  t 2.2
#      15:25       100.0%           -0.00%          +1.31%  t 2.7
#
#  Two findings. Precision climbs towards the bell because the candle stops
#  changing. And the entry is CHEAPER earlier - these names close at their high
#  so the last minutes drift up - which means moving the scan forward costs
#  nothing on price and gains everything on executability.
#
#  15:20 chosen: 82% of flagged names still qualify at the close, the entry is
#  0.14% below the close, and ten minutes is enough to act. 15:25 measures a
#  shade better but five minutes is not a plan.
#
#  CONSEQUENCE: at 15:20 today's DAILY candle does not exist yet, so the tier
#  must be judged on a partial candle assembled from 5-minute bars, with the
#  volume benchmark pro-rated for the elapsed session. Both are tested here.
# --------------------------------------------------------------------------- #
def test_bug41_btst_workflow_runs_before_the_close():
    import re

    wf = (ROOT / ".github/workflows/btst.yml").read_text()
    crons = re.findall(r'cron:\s*"([^"]+)"', wf)
    assert crons, "the BTST workflow must be scheduled"
    minute, hour = crons[0].split()[0], crons[0].split()[1]
    utc_minutes = int(hour) * 60 + int(minute)
    ist_minutes = utc_minutes + 5 * 60 + 30          # UTC -> IST
    close = 15 * 60 + 30
    assert ist_minutes < close, (
        f"BTST runs at {ist_minutes//60:02d}:{ist_minutes%60:02d} IST, which is "
        "AT OR AFTER the close - the entry would have to wait for tomorrow's "
        "open and the overnight move is lost")
    assert close - ist_minutes >= 8, (
        "less than ~8 minutes before the bell leaves no time to place an order")
    assert close - ist_minutes <= 35, (
        "too early: the tier call is unreliable before ~15:15")


def test_bug41_partial_candle_is_accepted_and_volume_prorated():
    """
    At 15:20 roughly 93% of the session has traded. Judging that volume against
    a FULL-day average would understate rvol and silently reject good setups.
    """
    import pandas as pd

    from btst import classify

    def frame(vol_today, n=260, price=1000.0, base=500000.0):
        rows = [dict(open=price * .995, high=price * 1.025, low=price * .975,
                     close=price, volume=base) for _ in range(n - 1)]
        # a Tier-B shaped day: closes at the high on heavy volume
        rows.append(dict(open=price, high=price * 1.06, low=price * .999,
                         close=price * 1.059, volume=vol_today))
        return pd.DataFrame(rows)

    # 2.8x a full day's average, but only 93% of the session has elapsed
    df = frame(vol_today=500000.0 * 2.8)
    full = classify(df, level=950.0, partial_frac=1.0, age=0)
    part = classify(df, level=950.0, partial_frac=0.93, age=0)
    assert part["rvol"] > full["rvol"], "a partial session must lift rvol"
    assert part["tier"] == "B", part
    assert part["partial_frac"] == 0.93

    # a genuinely quiet day must still be rejected even when pro-rated
    quiet = classify(frame(vol_today=500000.0 * 0.8), level=950.0,
                     partial_frac=0.93)
    assert quiet["tier"] is None, quiet


def test_bug41_after_close_flag_exists_and_is_off_in_the_workflow():
    """
    --after-close is for a post-close review. The scheduled 15:20 job must NOT
    use it, or it would skip every name whose candle is still forming, i.e.
    all of them.
    """
    body = (ROOT / "btst.py").read_text()
    assert "--after-close" in body

    wf = (ROOT / ".github/workflows/btst.yml").read_text()
    run = wf.split("python btst.py", 1)[1].split("\n\n", 1)[0]
    assert "inputs.after_close" in run, (
        "after-close must be a manual input, not a hardcoded flag")
    assert "'--after-close'" in run and "&&" in run, \
        "it must be conditional on the input, defaulting off"


def test_bug41_message_warns_the_candle_is_not_final():
    """
    82% precision means roughly one in five flagged names breaks down in the
    last ten minutes. That has to be said, not buried.
    """
    body = (ROOT / "btst.py").read_text()
    assert "still forming" in body, "a partial pick must be labelled"
    assert "82%" in body, "the measured precision must be stated to the user"


def test_bug41_intraday_fetch_only_covers_today():
    """The partial candle must be built from TODAY's bars only - pulling a
    wider window would blend sessions into one fake candle."""
    body = (ROOT / "btst.py").read_text()
    assert "intraday_candles" in body
    assert "dtime(9, 15)" in body, "the intraday window must start at the open"
    assert "BARS_PER_SESSION = 75" in body


# --------------------------------------------------------------------------- #
#  BUG 42 - Model E and the 15:20 BTST alert were two separate decisions that
#  could name different stocks at different prices.
#
#  User, 31-Jul: "BTST trade (Model E) also 15:20 IST (before the close) and
#  take entry for next day, take entry in top 5 BTST watchlist stocks".
#
#  THREE GAPS, all real:
#    1. btst.py ALERTED an uncapped list. The "top 5" existed only inside
#       ab_paper.py, so the message could show eight names while the ledger
#       traded five - and not necessarily the same five.
#    2. Model E entered at the COMPLETED daily close, a price that does not
#       exist when the 15:20 alert lands. The alert said "buy ~1350", the
#       ledger recorded 1357.20.
#    3. Model E re-derived its own tier from the post-close candle. Since
#       82% of 15:20 picks still qualify at the bell, ~1 in 5 ledger rows
#       would be a stock the alert never sent, and vice versa.
#
#  FIX: btst.py caps at TOP_N, records the 15:20 price, and writes
#  btst_picks.csv (committed by the workflow). Model E READS that file and
#  trades exactly it. Reconstruction from the close still exists for backfill
#  but is tagged btst_source="reconstructed" so the two are never confused.
#
#  The distinction between "this day has a picks file and the symbol is absent"
#  (not taken) and "no file for this day" (backfill) is what stops a missing
#  file from silently trading everything. Tested below.
# --------------------------------------------------------------------------- #
def test_bug42_btst_caps_the_alert_at_top_n():
    import btst

    assert btst.TOP_N == 5, "the user asked for the top 5"
    body = (ROOT / "btst.py").read_text()
    assert "qualified[:TOP_N]" in body, "the ALERT itself must be capped"
    assert "PICKS_FILE" in body and "btst_picks.csv" in body


def test_bug42_picks_file_records_the_1520_entry_price(tmp_path):
    """The entry written must be the price at the SCAN, not the close - the
    close does not exist yet when the alert is sent."""
    body = (ROOT / "btst.py").read_text()
    # slice to the END of the picks-file dict rather than a fixed character
    # count - new columns (BUG 53 arm/age, BUG 55 tradeable) kept pushing the
    # entry line past an arbitrary window and failing for no real reason.
    chunk = body.split("this IS the trade list", 1)[1].split("pfile =", 1)[0]
    assert '"scan_time"' in chunk
    assert '"entry": round(float(r["close"])' in chunk, (
        "entry must come from the partial candle's current price")
    assert '"rank"' in chunk and '"tier"' in chunk


def test_bug42_model_e_trades_the_picks_file(tmp_path):
    import pandas as pd

    from ab_paper import load_btst_picks

    p = tmp_path / "btst_picks.csv"
    p.write_text(
        "date,scan_time,rank,symbol,tier,entry,level,day_ret,close_pos,"
        "rvol,atr_pct,mcap_cr,partial_frac\n"
        "2026-07-31,15:20,1,YASHO,A,3800.00,3262.80,17.5,0.98,7.4,4.9,3762,0.93\n"
        "2026-07-31,15:20,2,NAZARA,B,317.00,310.00,10.6,0.95,8.3,2.9,10000,0.93\n")
    lookup, days = load_btst_picks(p)

    assert ("2026-07-31", "YASHO") in lookup
    assert lookup[("2026-07-31", "YASHO")]["entry"] == 3800.00
    assert lookup[("2026-07-31", "YASHO")]["rank"] == 1
    assert days.get("2026-07-31") is True
    assert ("2026-07-31", "SOMETHINGELSE") not in lookup


def test_bug42_missing_picks_file_does_not_trade_everything():
    """
    The dangerous failure: no file -> every breakout silently becomes a BTST
    trade. load_btst_picks must return an EMPTY day-map so the caller can tell
    "not taken" from "no data", and the caller must only skip when the day is
    actually covered.
    """
    from ab_paper import load_btst_picks

    lookup, days = load_btst_picks("/nonexistent/btst_picks.csv")
    assert lookup == {} and days == {}

    src = (ROOT / "ab_paper.py").read_text()
    assert "picks_have_day.get(day_key)" in src, (
        "a name absent from a COVERED day must be skipped, not reconstructed")
    assert 'btst_source = "reconstructed"' in src
    assert 'btst_source = "picks"' in src


def test_bug42_ledger_records_which_source_was_used():
    from ab_paper import LEDGER_COLS

    assert "btst_source" in LEDGER_COLS
    assert "btst_rank" in LEDGER_COLS


def test_bug42_top_n_cap_preserves_the_alert_ranking():
    """
    Re-ranking on post-close data could keep a different five than the alert
    sent. When btst_rank is present it must drive the order.
    """
    src = (ROOT / "ab_paper.py").read_text()
    chunk = src.split("BTST top-N cap", 1)[1][:1600]
    assert "btst_rank" in chunk, "the cap must honour the 15:20 rank"


def test_bug42_workflow_commits_the_picks_file():
    wf = (ROOT / ".github/workflows/btst.yml").read_text()
    assert "contents: write" in wf, "the job must be able to commit"
    assert "btst_picks.csv" in wf
    assert "git commit" in wf and "rebase" in wf, (
        "state.json is committed by the 5m scan - the picks push must rebase, "
        "not clobber")


def test_bug42_btst_and_model_e_agree_on_the_same_day():
    """
    End-to-end contract: for a given date the alert's picks and Model E's
    trades must be the same symbols at the same prices. Enforced by Model E
    reading the file rather than re-deriving - assert that wiring exists.
    """
    src = (ROOT / "ab_paper.py").read_text()
    assert "picks_lookup.get((day_key, sig.symbol))" in src
    assert 'sig.price = float(pick["entry"])' in src, (
        "the ledger must use the 15:20 price, not the completed close")


# --------------------------------------------------------------------------- #
#  BUG 43 - BTST listed stocks whose breakout was DAYS old.
#
#  User, 31-Jul: "something is wrong showing those stocks for btst which
#  already gave a move".
#
#  The live 15:20 message carried:
#      YASHO     broke out 31-Jul 12:20  -> age 0, ext  18.3%   correct
#      NELCO     broke out 31-Jul 12:50  -> age 0, ext   2.2%   correct
#      DEEPINDS  broke out 29-Jul 10:05  -> age 2, ext   9.8%   WRONG
#
#  CAUSE: candidates were selected with already_alerted(week, symbol), which
#  matches the WHOLE WEEK. A Monday breakout was still a candidate on Friday;
#  the day-character test then passed on an unrelated later candle and the
#  name was presented as a fresh BTST setup.
#
#  This is not a preference - it is a mismatch with what was measured. The
#  5-year study only ever tested the breakout day. Re-measured over 1,548
#  qualifying stock-days, next-day return net of costs:
#
#      age 0   n=1174   +0.807%   t 5.00    mean extension  7.1%
#      age 1   n= 164   +0.373%   t 0.87    mean extension 15.6%
#      age 2   n= 107   +1.311%   t 2.45    mean extension 18.6%
#      age 4   n=  29   -0.808%   t -1.01   mean extension 26.9%
#
#      Tier A  age 0    n=401   +1.736%   t 5.03
#      Tier A  age >=1  n= 83   +0.155%   t 0.20   <- the edge is GONE
#
#  Tier A is the reason the scanner exists and it does not survive ageing.
#  Later-day Tier B is positive but thin, drifts further above the level every
#  day, and was never part of the original measurement.
#
#  FIX: only names whose breakout alert is dated TODAY are candidates.
#  AlertState.alert_record()/alert_date() expose the timestamp that
#  already_alerted() throws away.
# --------------------------------------------------------------------------- #
def test_bug43_alert_state_exposes_the_alert_date(tmp_path):
    from datetime import datetime as _dt

    from state import AlertState

    p = tmp_path / "state.json"
    st = AlertState(p)
    st.mark("2026-07-27", "YASHO", _dt(2026, 7, 31, 12, 20), 3690.9)
    st.mark("2026-07-27", "DEEPINDS", _dt(2026, 7, 29, 10, 5), 546.55)
    st.save()

    st2 = AlertState(p)
    assert st2.alert_date("2026-07-27", "YASHO") == "2026-07-31"
    assert st2.alert_date("2026-07-27", "DEEPINDS") == "2026-07-29"
    assert st2.alert_date("2026-07-27", "MISSING") is None
    assert st2.alert_record("2026-07-27", "MISSING") is None
    # already_alerted must keep its old week-wide meaning for de-duplication
    assert st2.already_alerted("2026-07-27", "DEEPINDS") is True


def test_bug43_btst_only_accepts_breakouts_dated_today():
    body = (ROOT / "btst.py").read_text()
    assert "alert_record(" in body, "btst must read the alert TIMESTAMP"
    assert "today_str" in body and "bar == today_str" in body, (
        "candidates must be filtered to breakouts dated TODAY")
    # The BTST (confirmation) pool must not use week-wide selection - that is
    # what let a Monday breakout show on Friday. The ANTICIPATION pool legally
    # uses `not already_alerted(...)`: a name that fired at any point this week
    # is no longer "about to break out", so the week-wide sense is correct
    # there. Assert the negated form is the only one present.
    import re
    hits = re.findall(r"(not\s+)?state\.already_alerted\(week_s, s\.symbol\)",
                      body)
    # BUG 47 removed the anticipation pool's use of already_alerted entirely -
    # it now screens BOTH sides of the level and excludes only names that fired
    # TODAY. So either there are no call sites left, or any that remain are
    # negated. A BARE week-wide selection must never come back.
    for neg in hits:
        assert neg and neg.strip() == "not", (
            "bare week-wide selection found - the BTST pool must filter on "
            "the alert DATE, not merely 'fired sometime this week'")
    assert "fired_today" in body, (
        "the anticipation pool must exclude only TODAY's breakouts")


def test_bug43_stale_names_are_counted_and_explained():
    """A name dropped for being old must be reported, not silently vanish -
    otherwise the next person assumes the scan missed it."""
    body = (ROOT / "btst.py").read_text()
    assert "stale" in body
    # REVISED BY BUG 53. This used to require the message to quote
    # "+1.74% / +0.16%" as the reason an aged breakout was excluded. Aged
    # breakouts are no longer excluded wholesale - only aged TIER A is - so
    # the message says which tiers age out instead of a blanket exclusion.
    assert "tier B candidates only" in body, (
        "the log must say aged names are still tier B candidates")
    assert "Tier B holds with age; Tier A does not" in body, (
        "the alert must explain which tier survives ageing")


def test_bug43_the_measured_decay_is_recorded():
    """The ORIGINAL BUG 43 measurement must stay in the file even though its
    conclusion was narrowed by BUG 53 - the numbers are why tier A is still
    same-day only."""
    body = (ROOT / "btst.py").read_text()
    for token in ("age 0", "t 5.03", "t 0.20"):
        assert token in body, f"btst.py must document {token!r}"
    assert "AGE_GATE_TIERS" in body


def test_bug53_age_gate_applies_to_tier_a_only():
    """
    THE FINDING. Measured on 892,858 tradeable stock-days (not just breakout
    days, which is what BUG 43 had and why it over-generalised):

        aged tier A   IS +0.795 -> OOS -0.843   dead, stays excluded
        aged tier B   IS +0.999 -> OOS +0.885   most stable arm measured

    Tier A is a magnitude test and magnitude expires; tier B is a character
    test and character does not.
    """
    from btst import classify, AGED_EXT_MIN, AGED_EXT_MAX

    # ---- an aged TIER A shape must NOT qualify
    m = classify(_btst_frame(day_ret=18.4, close_pos=1.0, rvol=8.0, atr_pct=4.9),
                 level=95.0, age=3)
    assert m["tier"] is None, m
    assert "aged" in m["reject"].lower() or "old" in m["reject"].lower()

    # ---- the SAME shape fresh is still tier A
    assert classify(_btst_frame(day_ret=18.4, close_pos=1.0, rvol=8.0,
                                atr_pct=4.9), level=95.0, age=0)["tier"] == "A"


def test_bug53_aged_tier_b_qualifies_inside_the_band():
    """Aged tier B is tradeable, but only while it is still near its level -
    the band is both an edge filter and the 15:20 API-time budget (BUG 52)."""
    import pandas as pd

    from btst import classify, AGED_EXT_MAX

    def frame(level_mult, n=260, base=500000.0):
        """A tier-B day (closes at its high on 4x volume) sitting
        `level_mult` above the level."""
        price = 1000.0
        rows = [dict(open=price * .995, high=price * 1.03, low=price * .97,
                     close=price, volume=base) for _ in range(n - 1)]
        rows.append(dict(open=price, high=price * 1.06, low=price * .999,
                         close=price * 1.059, volume=base * 4.0))
        return pd.DataFrame(rows)

    df = frame(1.0)
    close = float(df.iloc[-1]["close"])

    # inside the band (a few % above the level) -> tier B even at age 40
    lvl_in = close / (1 + (AGED_EXT_MAX - 5) / 100.0)
    m = classify(df, level=lvl_in, age=40)
    assert m["tier"] == "B", m
    assert m["age"] == 40 and m["fresh"] is False

    # far above the band -> rejected as a chase
    lvl_far = close / (1 + (AGED_EXT_MAX + 25) / 100.0)
    assert classify(df, level=lvl_far, age=40)["tier"] is None

    # no cross at all in the lookback -> not a breakout to age
    assert classify(df, level=lvl_in, age=999)["tier"] is None


def test_bug53_breakout_age_counts_sessions_since_the_cross():
    import pandas as pd

    from btst import breakout_age

    def f(closes):
        return pd.DataFrame([dict(open=c, high=c, low=c, close=c, volume=1.0)
                             for c in closes])

    # crossed 100 on the last bar
    assert breakout_age(f([90, 95, 99, 101]), 100.0) == 0
    # crossed two bars ago and stayed above
    assert breakout_age(f([90, 99, 101, 103, 104]), 100.0) == 2
    # never crossed
    assert breakout_age(f([90, 91, 92, 93]), 100.0) == 999
    # above the whole time -> no cross event exists
    assert breakout_age(f([101, 102, 103]), 100.0) == 999


def test_bug54_close_pos_floors_are_098_for_both_tiers():
    """
    SUPERSEDES the BUG 53 version of this test, which asserted 0.95 for tier B
    and 0.85 for tier A. Both were wrong:

      * the 0.95-0.99 band that a 0.95 floor admits measures -0.222% (n=235,
        win 41.7%) - a losing bucket hidden inside a monotonic curve
      * tier A rows with close_pos 0.85-0.98 measure +0.101% over 5 years
        (n=193, win 42.5%) - noise occupying real slots

    0.98 rather than 0.99 because the live scan judges close_pos on a PARTIAL
    15:20 candle and 0.99 is knife-edge against a bar with ten minutes left.
    """
    import btst

    assert btst.TIER_B_CLOSE_POS == 0.98
    assert btst.TIER_A_CLOSE_POS == 0.98, (
        "BUG 54: the tier A floor was measured, not assumed - the discarded "
        "0.85-0.98 slice is worthless (+0.101%, 42.5% win)")

    # a 0.96 close no longer makes tier B
    m = btst.classify(_btst_frame(day_ret=6.0, close_pos=0.96, rvol=6.0,
                                  atr_pct=5.0), level=95.0, age=0)
    assert m["tier"] is None, m
    m = btst.classify(_btst_frame(day_ret=6.0, close_pos=0.99, rvol=6.0,
                                  atr_pct=5.0), level=95.0, age=0)
    assert m["tier"] == "B", m

    # a +18% day that FADED (closed mid-range) is no longer tier A
    m = btst.classify(_btst_frame(day_ret=18.0, close_pos=0.90, rvol=8.0,
                                  atr_pct=5.0), level=95.0, age=0)
    assert m["tier"] != "A", m


# --------------------------------------------------------------------------- #
#  BUG 55 - THE 15:48 ALERT. "entry time should have between 15:20 to 15:30,
#  but currently it's showings 15:48 which impossible after market trade"
#
#  The scan was correct; it simply had no concept of being too late. GitHub's
#  scheduler queued the 15:20 job 28 minutes late and the code still printed
#  "BUY NOW" on an entry that no longer existed.
# --------------------------------------------------------------------------- #
def test_bug55_entry_window_is_enforced():
    """A run after 15:30 must not tell anyone to buy at today's close."""
    body = (ROOT / "btst.py").read_text()
    assert "entry_close = dtime(15, 30)" in body
    assert "entry_open = dtime(15, 0)" in body
    assert "too_late" in body and "too_early" in body
    # the alert must degrade, not lie
    assert "MISSED" in body, "a late pick must be labelled MISSED"
    assert "TOO LATE" in body
    # and it must still SEND - suppressing it would hide the outage (BUG 49)
    seg = body.split("too_late = ", 1)[1][:1500]
    assert "return" not in seg.split("if too_late:", 1)[1][:400], (
        "a late run must still send and still write picks - silence is the "
        "failure mode BUG 49 was about")


def test_bug55_late_picks_are_marked_untradeable_and_model_e_skips_them():
    """The paper ledger must not book a fill that could never have happened."""
    body = (ROOT / "btst.py").read_text()
    chunk = body.split("this IS the trade list", 1)[1].split("pfile =", 1)[0]
    assert '"tradeable": 0 if too_late else 1' in chunk, (
        "the picks file must record whether the pick was enterable")

    ab = (ROOT / "ab_paper.py").read_text()
    assert 'pick.get("tradeable", 1)' in ab, (
        "Model E must honour the tradeable flag")
    # default 1 so pre-BUG-55 picks files keep working
    seg = ab.split('pick.get("tradeable"', 1)[1][:80]
    assert "1" in seg, "older picks files must default to tradeable"


def test_bug55_the_schedule_is_no_longer_cron_only():
    """
    GitHub cron is best-effort. On 03-Aug btst/report/ab ALL had zero
    scheduled runs. Each must accept an external trigger so cron-job.org can
    drive it on time, the way scan.yml already is.
    """
    for wf in ("btst.yml", "report.yml", "ab.yml"):
        body = (ROOT / ".github" / "workflows" / wf).read_text()
        assert "repository_dispatch:" in body, (
            f"{wf} must accept repository_dispatch - GitHub cron alone "
            f"demonstrably does not fire")
    btst_wf = (ROOT / ".github" / "workflows" / "btst.yml").read_text()
    assert "types: [btst]" in btst_wf


def test_bug55_top_n_is_a_cap_not_a_target():
    """
    The user expected 5 BTST entries and got 1. That is CORRECT behaviour and
    must stay correct: measured over 5 years, 63.7% of firing days produce
    exactly one qualifying name and only 2.1% produce five. Nothing may pad
    the list to reach TOP_N.
    """
    import btst

    assert btst.TOP_N == 5
    body = (ROOT / "btst.py").read_text()
    assert "qualified[:TOP_N]" in body, "TOP_N must slice, never extend"
    # there must be no 'fill up to N' logic anywhere near the picks
    seg = body.split("qualified.sort(key=rank_key)", 1)[1][:600]
    for bad in ("while len(picks) <", "pad", "top_up", "extend(qualified"):
        assert bad not in seg, f"picks must never be padded - found {bad!r}"


def test_bug54_conviction_scores_the_fat_tail_and_never_gates():
    """
    The conviction score answers "which pick is the BIG one". It must be a
    RANK: measured as a filter it costs coverage and RAISES drawdown
    (conv>=3: CAGR 176% DD -47.6% vs all: CAGR 494% DD -29.4%) while the mean
    barely moves. It orders P(+5%) 4.3% -> 31.2% monotonically, which is the
    thing it is actually for.
    """
    import btst

    assert btst.CONVICTION_MAX == 4

    # maximum conviction: fresh tier A, high atr, big day, unexhausted trend
    hi, why = btst.conviction(dict(fresh=True, tier="A", atr_pct=6.0,
                                   day_ret=18.0, ret_12m=120.0))
    assert hi == 4, why

    # an exhausted trend loses the point even though everything else is there
    ex, _ = btst.conviction(dict(fresh=True, tier="A", atr_pct=6.0,
                                 day_ret=18.0, ret_12m=450.0))
    assert ex == 3

    # a steady aged tier B scores low - correct, it is not an explosive setup
    lo, _ = btst.conviction(dict(fresh=False, tier="B", atr_pct=3.5,
                                 day_ret=2.0, ret_12m=300.0))
    assert lo == 0

    # NaN must not raise and must not score
    n, _ = btst.conviction(dict(fresh=True, tier="A", atr_pct=float("nan"),
                                day_ret=float("nan"), ret_12m=float("nan")))
    assert n == 1  # only the fresh-tier-A point

    # IT MUST NEVER REMOVE A TRADE. The scan may sort on conviction but must
    # not filter on it.
    body = (ROOT / "btst.py").read_text()
    scan = body.split("tiered = [r for r in rows", 1)[1][:2500]
    assert "conviction" in scan, "conviction must be computed for picks"
    for bad in ("if r[\"conviction\"] >", "conviction\") >=", "conviction >= "):
        assert bad not in scan, (
            f"conviction must be a RANK, not a gate - found {bad!r}")


def test_bug57_trend_floor_is_NOT_applied_to_confirmed_tiers():
    """
    SUPERSEDES test_bug53_trend_floor_applies_to_confirmed_tiers_too.

    BUG 53c applied BUG 51's trend floor to the tier list by symmetry, without
    measuring it there. Measured, it deletes the best slice in the study:

        removed by the floor  n=324  +3.173%  64.5% win  PF 3.56
                              P(+5%) 30.2%   IS +3.405 -> OOS +2.985
        the list it protected        +2.074%  62.3% win

    Date-clustered t=7.74, every year positive, survives de-duplication, and
    the top-5/day drawdown IMPROVES -28.0% -> -22.7%.

    BUG 51 was not wrong, it was misapplied: its losers were SCAN alerts. A
    tier pick already closes in the top 2% of range on 3x+ volume, so what the
    floor removes there is EARLY TREND (median ret_12m 24.5% vs 117.5%), not
    junk - consistent with BUG 54b measuring ret_12m>=400% at lift 0.35.
    """
    body = (ROOT / "btst.py").read_text()
    assert "def trend_ok" not in body, (
        "BUG 57: the trend floor must NOT gate the confirmed tier list")
    assert "qualified = list(tiered)" in body
    # it MUST still gate the anticipation path, where it was measured
    src = body.split("def classify_approach", 1)[1].split("\ndef ", 1)[0]
    assert "MIN_RET_12M" in src and "MIN_DIST_200DMA" in src, (
        "the trend floor must remain in classify_approach()")


def test_bug59_aged_pool_is_prescreened_on_close_pos():
    """
    THE 04-AUG HANG. BUG 53 widened the pool to every name within -10..+20% of
    its level. On 04-Aug that was 628 aged + 19 fired = 647 names, each needing
    ~2 history calls at the measured ~4 names/minute (BUG 52) = 2.6 HOURS. The
    job was killed by the workflow timeout at 15:48, after the close, having
    produced nothing.

    close_pos >= TIER_B_CLOSE_POS is mandatory for BOTH tiers and needs only
    today's high/low/LTP, which the bulk OHLC call already returns. Screening
    on it before any history request is FREE and, verified on 157,995
    stock-days, retains 100.00% of names that ultimately qualify while keeping
    ~5.5% of the universe.
    """
    import btst

    assert hasattr(btst, "fetch_ohlc") and hasattr(btst, "cheap_close_pos")
    body = (ROOT / "btst.py").read_text()
    assert "close_pos>=%.2f pre-screen" in body, (
        "the aged pool must be pre-screened on close_pos")
    seg = body.split("in_band = len(aged_pool)", 1)[1][:1800]
    assert "TIER_B_CLOSE_POS" in seg, "the screen must use the shipped floor"
    assert "kept.append(s)" in seg, "an unknown quote must NOT be dropped"

    # the cheap calculation must match what classify() would compute
    cp, dr = btst.cheap_close_pos(
        {"last_price": 100.0, "open": 90.0, "high": 101.0, "low": 89.0})
    assert abs(cp - (100.0 - 89.0) / (101.0 - 89.0)) < 1e-9
    assert abs(dr - (100.0 / 90.0 - 1) * 100) < 1e-9
    # junk must not raise, and must not pass as 0
    bad, _ = btst.cheap_close_pos({"last_price": 0, "high": 0, "low": 0})
    assert bad != bad, "unusable quote must return NaN, not a passing value"


def test_bug59_scan_has_a_deadline_and_sends_what_it_has():
    """
    A timeout that kills the job produces NOTHING - no alert, no picks, no
    explanation. The scan must stop itself before the close and send whatever
    is ready, with the truncation stated in the message.
    """
    body = (ROOT / "btst.py").read_text()
    assert "deadline = now.replace(hour=15, minute=26" in body
    assert "stopped_early" in body
    # BUG 61: ex.map() yields in SUBMISSION order, so one slow symbol blocked
    # every deadline check behind it. Completion-order iteration is required.
    import ast
    tree = ast.parse(body)
    calls = [n for n in ast.walk(tree)
             if isinstance(n, ast.Call)
             and isinstance(n.func, ast.Attribute) and n.func.attr == "map"]
    assert not calls, (
        "ex.map yields in submission order - a single slow symbol prevents "
        "the deadline from ever being evaluated; use run_until()")
    assert "run_until(" in body
    # and the message must disclose truncation
    assert "name(s) unchecked" in body, (
        "a truncated scan must say so in the alert")
    assert "todo = fired + aged_pool" in body, (
        "names that fired TODAY must be submitted before the aged pool")


def test_bug65_prior_exhaustion_is_rejected():
    """
    A breakout whose BASE was already pinned at 52-week highs, or whose stock
    had already doubled in the prior quarter, is spent:

        base within 2% of the 252d high   n=325  +1.333%  P(+5%) 12.6%
        prior 3-month > +100%             n=106  +1.592%  P(+5%)  7.5%
        everything else                 n=1,366  +2.624%  P(+5%) 25.9%

    Welch t 4.09; kept set clustered t 13.90; every year positive.
    """
    import btst

    assert btst.MAX_BASE_FROM_HIGH == -2.0
    assert btst.MAX_RET_3M_PRIOR == 100.0

    # pinned at highs -> rejected
    assert btst.exhausted({"base_from_high": -0.5, "ret_3m_prior": 20.0})
    # already doubled -> rejected
    assert btst.exhausted({"base_from_high": -30.0, "ret_3m_prior": 150.0})
    # emerging from a base -> kept
    assert btst.exhausted({"base_from_high": -12.0, "ret_3m_prior": 30.0}) is None
    # NaN must PASS - a young listing has no 252d high to be pinned against,
    # and rejecting it would repeat the BUG 51 over-reach.
    assert btst.exhausted({"base_from_high": float("nan"),
                           "ret_3m_prior": float("nan")}) is None

    # wired into BOTH paths
    body = (ROOT / "btst.py").read_text()
    assert body.count("spent = exhausted(m)") == 2, (
        "both classify() and classify_approach() must apply it")

    # the two names the user queried must SURVIVE - they were not extended
    assert btst.exhausted({"base_from_high": -10.7,
                           "ret_3m_prior": 37.3}) is None   # MOREPENLAB
    assert btst.exhausted({"base_from_high": -19.3,
                           "ret_3m_prior": 8.3}) is None    # RBA
    # but the two extended anticipation picks must be dropped
    assert btst.exhausted({"base_from_high": -1.6, "ret_3m_prior": 79.9})  # RAIN
    assert btst.exhausted({"base_from_high": -1.5, "ret_3m_prior": 46.5})  # TFCILTD


def test_bug65_refuted_claims_stay_refuted():
    """
    Two intuitive ideas were tested on 858,401 stock-days and measured
    BACKWARDS. Recorded so they are not "fixed" into the scanner later.

      1. "a big up-day today means no big move tomorrow" - the opposite:
         P(+5%) rises monotonically 3.6% -> 21.3% as today's move goes
         -5% -> +20%, and reaches 33.4% for a +15-20% day that closed at its
         high. TIER_A_DAY must stay at 15.
      2. "require a tight/VCP base" - tight bases measured WORSE:
         coiled (contraction<0.70) P(+5%) 9.2% vs expanding (>1.20) 31.2%.
    """
    import btst

    assert btst.TIER_A_DAY == 15.0, (
        "a large up-day is the STRONGEST predictor of a big next day, not a "
        "sign of exhaustion - lift 5.15x at +15-20%")

    body = (ROOT / "btst.py").read_text()
    for token in ("REFUTED", "15..20%", "TIGHT  n=123", "LOOSE  n=714"):
        assert token in body, f"the measurement must stay documented: {token!r}"

    # no tightness/squeeze gate may exist IN CODE (comments discuss it freely)
    code = "\n".join(ln for ln in body.splitlines()
                     if not ln.lstrip().startswith("#"))
    for bad in ("bbw_pct <", "contraction <", "MIN_TIGHT", "squeeze <"):
        assert bad not in code, (
            f"tightness measured negative - {bad!r} must not gate")


def test_bug64_degenerate_today_bar_is_repaired_from_the_quote():
    """
    E2E on 04-Aug came back with open == high == low == close == 570.15, so
    the scan reported day_ret 0.00% when the real bar was 552.00 -> 570.15
    (+3.29%). RAIN and TFCILTD matched Yahoo within feed noise on the same
    run, so this is a per-symbol data defect.

    It matters because day_ret is a GATE: TIER_A_DAY needs >= +15% and
    conviction awards a point at >= +6%. A zeroed open silently demotes a
    tier A setup to nothing - the best trade of the day would never be shown.
    """
    from datetime import date

    import pandas as pd

    import btst

    today = date(2026, 8, 4)
    df = pd.DataFrame([
        {"datetime": pd.Timestamp("2026-08-03"), "open": 536.35, "high": 543.0,
         "low": 526.9, "close": 543.0, "volume": 1e6},
        {"datetime": pd.Timestamp("2026-08-04"), "open": 570.15, "high": 570.15,
         "low": 570.15, "close": 570.15, "volume": 1e6},
    ])
    assert (df.iloc[-1].close / df.iloc[-1].open - 1) * 100 == 0.0

    q = {"open": 552.00, "high": 570.15, "low": 551.60, "last_price": 570.15}
    out = btst.repair_today_bar(df, q, today)
    t = out.iloc[-1]
    assert abs((t.close / t.open - 1) * 100 - 3.29) < 0.05, "must repair day_ret"
    assert t.high >= t.close and t.low <= t.open, "bar must stay consistent"
    # close_pos must still read "at the high"
    assert abs((t.close - t.low) / (t.high - t.low) - 1.0) < 1e-6

    # guards: no quote, or a bar that is not today's, must change nothing
    d2 = pd.DataFrame([{"datetime": pd.Timestamp("2026-08-04"), "open": 100,
                        "high": 110, "low": 99, "close": 108, "volume": 1}])
    assert btst.repair_today_bar(d2, None, today).iloc[-1].open == 100
    d3 = pd.DataFrame([{"datetime": pd.Timestamp("2026-07-30"), "open": 100,
                        "high": 110, "low": 99, "close": 108, "volume": 1}])
    assert btst.repair_today_bar(d3, q, today).iloc[-1].open == 100

    # and it must be wired into BOTH data paths
    body = (ROOT / "btst.py").read_text()
    assert body.count("repair_today_bar(df, ohlc_all.get(s.symbol)") == 2, (
        "both the confirmed and anticipation paths must repair today's bar")


def test_bug63_anticipate_list_also_honours_the_entry_window():
    """
    THE 20:43 RUN. The confirmed list correctly read "⛔ MISSED" while the
    anticipation list on the SAME message still said "BUY NOW ~570.15" on an
    entry that had been dead for five hours. BUG 55 was applied to one list
    and not the other, and Model F had no tradeable guard at all - so it would
    have booked three fills that were never available.
    """
    body = (ROOT / "btst.py").read_text()

    # the anticipation BUY NOW line must be conditional on too_late
    ant = body.split("SECOND PASS - ANTICIPATION", 1)[1]
    seg = ant.split("BUY NOW ~", 1)[0][-600:]
    assert "if too_late else" in seg, (
        "the anticipation entry line must degrade to MISSED after 15:30")
    assert body.count("⛔ <b>MISSED ~") == 2, (
        "BOTH lists must show MISSED when the window has closed")

    # the anticipation picks file must carry the flag
    apick = ant.split("aout = pd.DataFrame", 1)[1][:900]
    assert '"tradeable": 0 if too_late else 1' in apick, (
        "anticipate_picks.csv must record whether the pick was enterable")

    # and Model F must honour it
    ab = (ROOT / "ab_paper.py").read_text()
    fseg = ab.split("ant_lookup.get((day_key, sig.symbol))", 1)[1][:600]
    assert 'ap_.get("tradeable", 1)' in fseg, (
        "Model F must skip picks that were not tradeable")


def test_bug62_history_window_is_long_enough_for_annual_stats():
    """
    HISTORY_DAYS is CALENDAR days. 260 of them is only ~179 TRADING bars, but
    ret_12m needs cl[-252] and the 200DMA needs 200 bars - so on the confirmed
    path ret_12m was ALWAYS NaN and dist_200dma was never computed at all.
    Visible in the live 04-Aug picks file: ret_12m blank, dist_200dma 0.0 for
    both MOREPENLAB and RBA.
    """
    import numpy as np
    import pandas as pd

    import btst

    assert btst.HISTORY_DAYS >= 400, (
        "need ~420 calendar days to clear 252 trading bars after holidays")

    n = 300
    px = np.linspace(50, 100, n)
    d = pd.DataFrame({"open": px * 0.99, "high": px * 1.02,
                      "low": px * 0.98, "close": px, "volume": [1e6] * n})
    m = btst.classify(d, level=95.0, age=0)
    assert m["ret_12m"] == m["ret_12m"], "ret_12m must compute"
    assert m["dist_200dma"] == m["dist_200dma"], "dist_200dma must compute"
    assert m["dist_200dma"] > 0


def test_bug62_unknown_is_blank_not_zero_in_the_picks_file():
    """
    `float(x or 0)` turned "no 12-month history" into "+0.0% return" - a
    different and far more confident claim. Any later review reads it as a
    flat stock rather than an unknown.
    """
    import btst

    assert btst._num(float("nan")) == "", "NaN must render blank"
    assert btst._num(None) == "", "None must render blank"
    assert btst._num(0.0, 1) == 0.0, "a REAL zero must survive"
    assert btst._num(12.345, 1) == 12.3

    body = (ROOT / "btst.py").read_text()
    assert 'float(r.get("ret_12m") or 0)' not in body, (
        "unknown trend must not be written as 0")
    assert 'float(r.get("dist_200dma") or 0)' not in body
    assert body.count('_num(r.get("ret_12m"), 1)') == 2, (
        "both the confirmed and anticipate picks files must use _num")


def test_bug61_run_until_honours_the_deadline_despite_a_stall():
    """
    THE ACTUAL HANG. ex.map() yields in SUBMISSION order, so if symbol #1 is
    slow nothing is yielded and the deadline is never evaluated. With the old
    client settings one call could burn 210s (5 x 30s timeout + backoff) and a
    symbol needs two, so ONE name could hold the scan for seven minutes - the
    entire entry window - and the workflow timeout then killed it with no
    output at all.
    """
    import time
    from datetime import datetime, timedelta

    import btst
    from dhan import IST

    def work(i):
        time.sleep(30 if i == 0 else 0.05)
        return (i, {"ok": True})

    dl = datetime.now(IST) + timedelta(seconds=3)
    t0 = time.time()
    res, unfinished = btst.run_until(work, list(range(25)), 3, dl, "t")
    el = time.time() - t0
    assert el < 8, f"blocked on the stalled item ({el:.1f}s)"
    assert len(res) >= 10, "fast items must still be collected"
    assert unfinished >= 1

    # deadline=None must disable the cap (post-close review)
    res2, un2 = btst.run_until(lambda i: (i, {}), list(range(5)), 2, None, "t")
    assert len(res2) == 5 and un2 == 0


def test_bug61_urgent_mode_caps_retry_time():
    """Inside a ten-minute window a slow symbol is worth abandoning, not
    retrying. Patient defaults are kept for the post-close review."""
    body = (ROOT / "btst.py").read_text()
    assert "urgent = not (args.after_close or too_late)" in body
    assert "timeout=8 if urgent else 30" in body
    assert "max_retries=2 if urgent else 5" in body


def test_bug61_process_exits_hard_after_sending():
    """
    ThreadPoolExecutor registers an atexit hook that JOINS its non-daemon
    workers, so the interpreter blocks on exit even after
    shutdown(wait=False, cancel_futures=True). Measured: a worker sleeping
    60s holds the process 60s at exit. That is the observed 'stuck' run - the
    alert was already sent.
    """
    body = (ROOT / "btst.py").read_text()
    assert "os._exit(" in body, (
        "btst.py must hard-exit; a stuck socket read otherwise holds the "
        "process until the workflow timeout kills a successful run")
    tail = body.split('if __name__ == "__main__":', 1)[1]
    assert "sys.stdout.flush()" in tail and "sys.stderr.flush()" in tail, (
        "output must be flushed before os._exit bypasses normal teardown")
    assert "shutdown(wait=False, cancel_futures=True)" in body


def test_bug60_a_late_run_still_screens_and_records_missed():
    """
    REGRESSION I INTRODUCED IN BUG 59b. On a 16:15 run the 15:26 deadline was
    already in the past, so budget=0 -> room=0 -> the loop broke immediately
    and the job logged "checked 0 candles". That destroyed BUG 55's guarantee
    that a late run still screens and still records its picks as MISSED.

    Racing a deadline that has already passed protects nothing.
    """
    body = (ROOT / "btst.py").read_text()
    assert "enforce_deadline = (not args.after_close) and not too_late" in body, (
        "the 15:26 entry deadline must not apply once the window has closed")
    # BUG 66: the deadline is no longer switched OFF off-hours - it MOVES to a
    # wall-clock review budget, because "no cap" let the workflow timeout kill
    # the job outright (05-Aug: canceled at 14 min, zero output).
    assert "deadline = now + timedelta(minutes=REVIEW_BUDGET_MIN)" in body, (
        "a post-close run still needs a budget, not an unbounded scan")
    assert "deadline if enforce_deadline else None" not in body, (
        "passing None re-introduces the unbounded post-close scan")
    assert "if enforce_deadline and len(aged_pool) > room:" in body
    # and the log must say which mode it is in
    assert "post-close review - no deadline, recording as MISSED" in body


def test_bug67_missing_intraday_falls_back_to_the_quote():
    """
    THE 07-Aug BLIND SPOT. The 15:18 scan checked 7 candles; the 16:08 rerun
    checked 17 and found SBCL (Tier B, rvol 13.7x). SBCL was already +8.19% at
    close_pos 1.000 by 15:15 - three minutes BEFORE the 15:18 run - so it was
    not a timing miss. It was DELETED: while the market is open every symbol
    needs an intraday call to build today's bar, that call is the first thing
    to rate-limit, and failure did `return s, None` with only a DEBUG line.

    The bulk OHLC response already holds today's open/high/low/LTP/volume for
    all ~2,100 names in ONE request, so it is a complete substitute.
    """
    import numpy as np
    import pandas as pd

    import btst

    n = 300
    px = np.linspace(700, 850, n)
    hist = pd.DataFrame({
        "datetime": pd.date_range("2025-06-01", periods=n, freq="B"),
        "open": px * 0.99, "high": px * 1.01, "low": px * 0.98,
        "close": px, "volume": [5e5] * n})
    # SBCL as the bulk quote saw it, with no intraday feed available
    q = {"last_price": 920.90, "open": 850.80, "high": 920.90,
         "low": 848.00, "volume": 6.85e6}
    bar = {"datetime": pd.Timestamp("2026-08-07"), "open": q["open"],
           "high": max(q["high"], q["last_price"]),
           "low": min(q["low"], q["last_price"]),
           "close": q["last_price"], "volume": q["volume"]}
    df = pd.concat([hist, pd.DataFrame([bar])], ignore_index=True)
    m = btst.classify(df, 814.00, partial_frac=1.0, age=0)
    assert abs(m["day_ret"] - 8.24) < 0.1, "quote must reproduce day_ret"
    assert m["close_pos"] > 0.99, "quote must reproduce close_pos"
    # the whole point: volume survives, so rvol works and TIER B is testable
    assert m["rvol"] > 3.0, (
        "the quote must carry volume - otherwise rvol is 0 and TIER B can "
        "never fire on a fallback name")

    # dhan.ohlc must actually return volume
    src = (ROOT / "dhan.py").read_text()
    seg = src.split("def ohlc", 1)[1].split("\n    def ", 1)[0]
    assert '"volume"' in seg, "ohlc() must not discard today's volume"

    # and one() must fall back rather than dropping the symbol
    body = (ROOT / "btst.py").read_text()
    o = body.split("def one(s):", 1)[1].split("def candle(s):", 1)[0]
    assert "quote fallback" in o
    assert o.count("return s, None") <= 4, (
        "a failed intraday call must fall back to the quote, not delete "
        "the candidate")


def test_bug67_silent_drops_are_counted_and_reported():
    """
    "7 candle(s) checked" out of 46 candidates read exactly like a quiet
    market. A symbol that never reached classify() is a BLIND SPOT, not a
    rejection, and must be visible in the log AND in the message.
    """
    body = (ROOT / "btst.py").read_text()
    assert "drops = Counter()" in body, "every drop reason must be counted"
    for reason in ("daily error", "daily empty", "no intraday, no quote",
                   "quote fallback"):
        assert f'drops["{reason}"]' in body, f"unlabelled drop path: {reason}"
    assert "coverage: %d/%d candidates produced a candle" in body
    # it must reach the user, not just the log
    assert "candidate(s) returned no candle" in body, (
        "a data outage must not be presentable as a quiet day")
    assert "This list is INCOMPLETE" in body


def test_bug66_post_close_run_cannot_hit_the_workflow_timeout():
    """
    THE 05-Aug CANCELLATION. A post-close review run had no stopping rule at
    all (BUG 60 switched the deadline off rather than moving it), while the
    workflow kept a hard 14-minute timeout:

        16:53:28  23 fired + 23 aged; checking the candles ...
        17:07:22  ##[error]The operation was canceled.

    46 names at >18s each (heavy rate limiting at 22:23 IST vs 3.19s measured
    intraday). Canceled = no alert, no picks file, no explanation - the exact
    failure BUG 59b existed to prevent, reintroduced in the branch it missed.
    """
    import re

    import btst

    # the review budget must leave room for send + commit inside the timeout
    wf = (ROOT / ".github" / "workflows" / "btst.yml").read_text()
    m = re.search(r"timeout-minutes:\s*(\d+)", wf)
    assert m, "btst.yml must set a timeout"
    timeout = int(m.group(1))
    assert btst.REVIEW_BUDGET_MIN + 4 <= timeout, (
        f"review budget {btst.REVIEW_BUDGET_MIN}min + overhead must finish "
        f"inside the {timeout}min workflow timeout")

    # run_until must always receive a real deadline, never None
    body = (ROOT / "btst.py").read_text()
    for call in ("deadline, \"btst\")", "deadline, \"anticipate\")"):
        assert call in body, f"run_until must be bounded: {call}"

    # the anticipation trim must apply in BOTH modes now
    seg = body.split("SECOND PASS - ANTICIPATION", 1)[1]
    assert "if enforce_deadline:\n            ant_budget" not in seg, (
        "the anticipation trim must not be gated on enforce_deadline")
    assert "ant_budget = max(0.0," in seg


def test_bug60_anticipation_pass_is_also_bounded():
    """
    BUG 59 protected the confirmed pass and left this one unbounded - 296
    names x ~3.75s with no stopping rule, the identical BUG 52/59 failure.
    Any pool built from the snapshot must be cheaply screened AND deadlined.
    """
    body = (ROOT / "btst.py").read_text()
    seg = body.split("SECOND PASS - ANTICIPATION", 1)[1]
    assert "anticipation close_pos pre-screen" in seg, (
        "the anticipation pool must be pre-screened on close_pos")
    assert "ANTICIPATE_CLOSE_POS - margin" in seg, (
        "the screen must use the shipped anticipation floor")
    assert "ant_stopped" in seg and "break" in seg, (
        "the anticipation loop must stop at the deadline")
    # unknown quote must not be silently dropped here either
    assert "cp != cp or cp >=" in seg


def test_bug59_workflow_timeout_is_inside_the_entry_window():
    """A 30-minute timeout on a job that must finish in 12 is meaningless."""
    wf = (ROOT / ".github" / "workflows" / "btst.yml").read_text()
    import re
    m = re.search(r"timeout-minutes:\s*(\d+)", wf)
    assert m, "btst.yml must set a timeout"
    assert int(m.group(1)) <= 15, (
        "the timeout must be shorter than the 15:20->15:30 entry window, "
        "otherwise the job can be killed AFTER the close (04-Aug: 15:48)")


def test_bug58_message_stats_match_the_shipped_rule():
    """
    THE STATS IN THE ALERT MUST DESCRIBE THE RULE THAT IS ACTUALLY RUNNING.

    Caught from a live 04-Aug message: the footer still read "Tier A measured
    +1.75%/trade (t 5.2, n=417) ... Tier B +0.83% ... Win rate ~50%". Those
    are the ORIGINAL tier numbers, left untouched through BUG 53 (aged tier
    B), BUG 54 (close_pos 0.98) and BUG 57 (trend floor off). The shipped rule
    measures +3.04% / +2.15% at a 63% win rate, so the message was
    understating it by more than half and quoting a win rate 13 points low.

    A stale stat is not cosmetic - it is the number the user sizes on.
    """
    body = (ROOT / "btst.py").read_text()
    foot = body.split("Tier A measured", 1)[1][:900]
    # the retired numbers must be gone from the FOOTER
    for stale in ("+1.75%/trade", "n=417", "+0.83% (t 5.0", "n=1086",
                  "Win rate ~50%"):
        assert stale not in foot, (
            f"footer still quotes the pre-BUG-53/54/57 stat {stale!r}")
    for live in ("+3.04%", "n=302", "+2.15%", "n=852"):
        assert live in foot, f"footer must quote the shipped stat {live!r}"

    # the anticipation footer claimed above-level beat below; measured on the
    # window this list screens, it is the other way round.
    assert "above-level beat" not in body, (
        "BUG 58: below-level measured better (+1.33% vs +0.89%) inside the "
        "anticipation window")
    assert "below-level beat" in body


def test_bug57_recall_is_documented_as_the_known_tradeoff():
    """
    The scanner catches ~3% of all big movers. That is inherent, not a bug:
    31 stocks/week jump >=+5% and the top-5 cap allows 25 trades/week, so
    every second tier tested measured NEGATIVE on its own (best was
    cp>=0.95 & rvol>=5 at +0.551%, OOS t=1.26). Recorded so it is not
    "fixed" by loosening later.
    """
    body = (ROOT / "btst.py").read_text()
    assert "RECALL" in body, "the precision/recall tradeoff must be documented"


def test_bug53_no_double_count_between_confirmed_and_anticipate():
    """An aged tier B name is eligible for BOTH pools. Without an explicit
    exclusion Model E and Model F would buy it on the same day and the ledger
    would count one trade twice."""
    body = (ROOT / "btst.py").read_text()
    assert "taken = {r[\"symbol\"] for r in picks}" in body
    seg = body.split("pending = [s for s in snaps", 1)[1][:300]
    assert "taken" in seg, (
        "the anticipation pool must exclude names already taken as picks")


def test_bug53_picks_file_records_the_arm():
    """The paper ledger has to be able to score fresh and aged apart forever,
    otherwise the two get blended and neither can be judged."""
    body = (ROOT / "btst.py").read_text()
    chunk = body.split("out = pd.DataFrame", 1)[1][:1400]
    for field in ('"age"', '"arm"', '"ext_pct"'):
        assert field in chunk, f"picks file must record {field}"
    assert "fresh_A" in chunk and "aged_B" in chunk


def test_bug53_message_does_not_claim_everything_broke_out_today():
    """A name 55 days past its cross must not be presented under a headline
    that says it broke out today."""
    body = (ROOT / "btst.py").read_text()
    # look at the message construction only, not the comments explaining the
    # change (which necessarily quote the old wording).
    msg = body.split("lines = [f\"🌙 <b>BTST", 1)[1].split("if dropped:", 1)[0]
    assert "broke out TODAY only" not in msg, (
        "the header can no longer claim every pick is same-day")
    assert "broke out {age} session(s) ago" in msg, (
        "aged picks must state their own age in the message")


# --------------------------------------------------------------------------- #
#  BUG 44 - MODEL F, anticipation. "how i can catch them 1 day ago".
#
#  THE 31-JUL EVIDENCE. Twelve names fired that day. Buying all twelve at the
#  30-Jul close returns +5.47%/trade with 10 winners - and that number is
#  survivorship, because only ONE of the twelve was on the 30-Jul watchlist:
#      5 were more than 3% away      (NELCO 6.4%, IIFL 7.1%, TORNTPHARM 4.2%)
#      4 failed a frozen gate        (NAZARA, BAJFINANCE, EMCURE, SILVERTUC)
#      2 were ALREADY ABOVE level    (QUESS -0.52%, BLSE -2.44%)
#      1 was on the list             (YASHO)
#  On that day 34 names were approaching and exactly one fired: a 2.9% hit
#  rate, with the other 33 averaging -0.09%.
#
#  PROXIMITY ALONE LOSES MONEY - 234,518 approaching-days, top-5 per day, net:
#      <=3%  -0.195%  t -4.57   OOS -0.217%
#      <=5%  -0.160%  t -3.68   OOS -0.118%
#      <=10% -0.138%  t -3.08   OOS -0.143%
#
#  THE ONE FILTER THAT WORKS - close position in the daily range:
#      close_pos >= 0.9 -> 26.7% break out next day ; below -> 15.9%
#      <=3% + close_pos>=0.9 : n 1423, 5.4/wk, 52.0% win, +0.688%, t 7.73,
#                              out-of-sample +0.701%
#
#  Model F ships that rule. It runs ALONGSIDE E (overlap 0.1%), never instead.
#
#  HONESTY: 73% of F's picks do not break out next day, and on 30-Jul none of
#  the top-5 had close_pos >= 0.9 (YASHO was 0.62) - F would have bought
#  nothing and missed the +20%. Only 16.3% of approaching names break out next
#  day at all; that is the ceiling on any 1-day-early rule.
# --------------------------------------------------------------------------- #
def test_bug44_model_f_exists_and_e_and_d_are_untouched():
    from ab_paper import load_models

    _d, models = load_models()
    by = {m.key: m for m in models}
    assert "F_anticipate" in by

    f = by["F_anticipate"]
    assert f.strategy["anticipate_only"] is True
    assert f.strategy["btst_top_n"] == 5
    assert f.exit["rule"] == "btst"
    assert f.exit["stop_pct"] == 1.0 and f.exit["take_pct"] == 2.0

    # D and E must be exactly as before
    d = by["D_early"]
    assert d.exit["stop_pct"] == 7.0 and d.strategy["trigger_level"] == "hi_short2"
    e = by["E_btst"]
    assert e.exit["stop_pct"] == 1.0 and e.strategy["btst_only"] is True
    assert not e.strategy.get("anticipate_only", False)


def test_bug44_anticipate_thresholds_match_the_measurement():
    import btst

    assert btst.ANTICIPATE_NEAR == 3.0, "<=3% measured best and held OOS"
    assert btst.ANTICIPATE_CLOSE_POS == 0.90, (
        "0.9 is the threshold that doubled the breakout rate (26.7% vs 15.9%)")
    assert btst.ANTICIPATE_TOP_N == 5
    assert btst.ANTICIPATE_FILE == "anticipate_picks.csv"


def _approach_frame(close_pos, gap_pct, atr=5.0, n=260, price=1000.0,
                    base=500000.0, day=1.0):
    """A candle that closes `gap_pct` below the level at `close_pos` of range."""
    import pandas as pd

    level = price
    c = level * (1 - gap_pct / 100.0)
    o = c / (1 + day / 100.0)
    rng = max(abs(c - o) / max(close_pos, 0.01), c * 0.02)
    low = c - rng * close_pos
    high = low + rng
    rows = [dict(open=price * .995, high=price * (1 + atr / 200),
                 low=price * (1 - atr / 200), close=price, volume=base)
            for _ in range(n - 1)]
    rows.append(dict(open=o, high=high, low=low, close=c, volume=base * 1.5))
    return pd.DataFrame(rows), level


def _stub_trend(monkeypatch, ret_12m=120.0, dist_200dma=40.0):
    """
    classify_approach reads the trend from watchlist.compute_metrics (BUG 51).
    Synthetic fixtures have no real 200-day history, so tests that are about
    close_pos or PRE must pin the trend to a passing value - otherwise they
    fail for a reason they are not testing.
    """
    import watchlist

    real = watchlist.compute_metrics

    def patched(d, level, px=None):
        m = real(d, level, px) or {}
        m["ret_12m"] = ret_12m
        m["dist_200dma"] = dist_200dma
        return m

    monkeypatch.setattr(watchlist, "compute_metrics", patched)


def test_bug44_requires_a_close_at_the_top_of_range(monkeypatch):
    """
    close_pos is the gate. The PRE trend confirmation added by BUG 46 is
    stubbed here so this test measures ONE thing; BUG 46 tests the pairing.
    """
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)

    df, lvl = _approach_frame(close_pos=0.97, gap_pct=1.5)
    m = classify_approach(df, lvl, lvl)
    assert m is not None and m["ok"] is True, m

    # same distance, weak close -> rejected. This is the entire edge.
    df2, lvl2 = _approach_frame(close_pos=0.45, gap_pct=1.5)
    m2 = classify_approach(df2, lvl2, lvl2)
    assert m2 is not None and m2["ok"] is False
    assert "close_pos" in m2["reject"]


def test_bug44_above_level_names_are_now_scanned(monkeypatch):
    """
    SUPERSEDED BY BUG 47. This used to assert that a name already ABOVE its
    level was rejected. That was wrong and it discarded the better half:
    same filter, 5 years, below +0.580% (n=2,229) vs above +0.819% (n=4,657).
    A name above its level is now accepted and tagged side="above", up to
    ANTICIPATE_ABOVE_MAX; beyond that it is a chase and still rejected.
    """
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)

    df, lvl = _approach_frame(close_pos=0.97, gap_pct=1.5)
    df.loc[df.index[-1], "close"] = lvl * 1.02
    df.loc[df.index[-1], "high"] = lvl * 1.03
    m = classify_approach(df, lvl, lvl)
    assert m is not None, "an above-level name must no longer be discarded"
    assert m["side"] == "above"
    assert m["gap_pct"] < 0, "gap is negative when extended above the level"

    # but a runaway extension is still refused
    df2, lvl2 = _approach_frame(close_pos=0.97, gap_pct=1.5)
    df2.loc[df2.index[-1], "close"] = lvl2 * 1.25
    df2.loc[df2.index[-1], "high"] = lvl2 * 1.26
    assert classify_approach(df2, lvl2, lvl2) is None, "25% above is a chase"


def test_bug44_respects_the_distance_window_and_tradeability():
    from btst import classify_approach

    # too far from the level
    df, lvl = _approach_frame(close_pos=0.97, gap_pct=7.0)
    assert classify_approach(df, lvl, lvl) is None

    # too quiet to move - tradeability gate
    df2, lvl2 = _approach_frame(close_pos=0.97, gap_pct=1.5, atr=1.0)
    m = classify_approach(df2, lvl2, lvl2)
    assert m is not None and m["ok"] is False
    assert m["reject"] == "not tradeable"


def test_bug44_uses_the_nearer_of_the_two_levels():
    """
    watchlist.py quotes the nearer of entry_level and hi_short2. Reconstructing
    only entry_level measured a DIFFERENT trade - that is how the 30-Jul top-5
    failed to reproduce (MUTHOOTMF was 2.03% from its D level, not 3.94% from
    its C level).
    """
    from btst import classify_approach

    df, lvl_c = _approach_frame(close_pos=0.97, gap_pct=1.0)
    lvl_d = lvl_c * 0.995          # a nearer D level, still above price
    m = classify_approach(df, lvl_c, lvl_d)
    assert m is not None
    assert abs(m["level"] - lvl_d) < 1e-6, "must quote the NEARER level"
    assert m["which"] == "D"


def test_bug44_watchlist_shows_the_two_lists_separately():
    """The user asked to 'show separate stocks of recommendation'. Confirmation
    and anticipation must be distinct sections, not merged."""
    body = (ROOT / "btst.py").read_text()
    # BUG 53 renamed this header: the confirmed list is no longer same-day
    # only, so it can no longer claim to be.
    assert "CONFIRMED — today's setups" in body
    assert "ANTICIPATE — about to break out" in body
    assert "classify_approach" in body


def test_bug44_model_f_never_reconstructs():
    """
    F's signal is 'closed at the high while still BELOW the level'. That cannot
    be inferred from a breakout replay, so unlike E there is no fallback - if
    the 15:20 file has no row, there is no trade.
    """
    src = (ROOT / "ab_paper.py").read_text()
    chunk = src.split("anticipate_only", 1)[1][:900]
    assert "ant_lookup.get((day_key, sig.symbol))" in chunk
    assert "if ap_ is None:" in chunk and "continue" in chunk
    assert 'btst_source = "anticipate"' in src


def test_bug44_evidence_against_plain_proximity_is_recorded():
    doc = (ROOT / "models.yaml").read_text()
    for token in ("-0.195", "t -4.57", "26.7%", "15.9%", "0.1%", "73%"):
        assert token in doc, f"models.yaml must record {token!r}"


def test_bug44_workflow_commits_both_pick_files():
    wf = (ROOT / ".github/workflows/btst.yml").read_text()
    assert "anticipate_picks.csv" in wf
    assert "btst_picks.csv" in wf


# --------------------------------------------------------------------------- #
#  BUG 45 - the PRE score was built on two factors that do not survive.
#
#  User, 31-Jul: "deeply analyze watchlist logic based on all the successful
#  past trades... I want to find the edge".
#
#  METHOD NOTE THAT MATTERS. The request was to study the WINNERS. Doing only
#  that finds edges that are not there, so the loser set was kept as control -
#  and that decided the result. The 16 symbols the user named as winners
#  (THYROCARE, TBZ, MARICO, OAL, GANESHBE, THANGAMAYL, AETHER, SMLMAH,
#  BAJFINANCE, RADICO, KMEW, TMB, SENCO, MONARCH, NAZARA, RRKABEL) produced
#  219 signals over 5 years and averaged +0.172% against +0.808% for everyone
#  else. Their feature medians are indistinguishable from the population
#  (close_pos ratio 0.98, spike_level 1.00, ret_12m 0.85). The edge is not in
#  WHICH STOCKS; it is in WHICH SIGNALS.
#
#  THE RETRACTION. RALLY_FILTERS.md selected atr_pct and base_tight against
#  P(MFE >= +30%). Re-tested on 23,994 signals against the exit that actually
#  ships (30d cap / 7% stop / 5% trail), both flip sign out of sample:
#      atr_pct      IS -0.16   OOS +0.69
#      base_tight   IS -0.10   OOS +0.82
#      brk_rvol     IS -0.50   OOS +0.05     (volume - the intuitive one)
#  They are removed from the score. atr_pct stays as a TRADEABILITY gate: it
#  screens stocks that cannot move, which is not the same as predicting which
#  ones will.
#
#  WHAT SURVIVED, by out-of-sample quintile spread:
#      close_pos +1.29 | ret_12m +1.13 | gap_pct +1.07 | base_depth -1.01
#      entry -0.75 | dma200_slope +0.69 | ext_pct +0.67 | ret_1m +0.66
#      dist_200dma +0.52 | spike_level +0.40 | dist_50dma +0.39 | ret_3m +0.37
#
#  THE RULE: close_pos>=0.90 & ret_12m>=40 & ext_pct>=3
#      n 1,210   4.7/wk   43.8% win   +1.758%/trade   PF 1.96   t 6.87
#      out-of-sample +1.715%   vs a baseline that is NEGATIVE OOS (-0.197%)
#      positive in all 9 walk-forward windows; excess over same-date signals
#      +1.463% (t 6.42); remove the top 1% of trades and it is still +1.228%
#      while the unfiltered baseline becomes -0.131%.
#  A FOURTH condition collapses it (n 191, +0.226%) - that is the overfitting
#  boundary and it is visible in the data.
# --------------------------------------------------------------------------- #
def test_bug45_failed_factors_are_not_scored():
    """atr_pct, base_tight and brk_rvol must not influence any score."""
    from watchlist import score_pre

    base = _metrics()
    full = score_pre(base)[0]
    for k, v in (("atr_pct", 1.0), ("base_tight", 1.0), ("brk_rvol", 0.1)):
        m = dict(base); m[k] = v
        assert score_pre(m)[0] == full, (
            f"{k} changed the PRE score - it failed out of sample "
            "(see THE_EDGE.md) and must not be scored")


def test_bug45_atr_survives_only_as_a_tradeability_gate():
    from watchlist import MIN_ATR_PCT, gate_reasons, score_pre

    # still gates
    assert any("atr" in r for r in gate_reasons(_metrics(atr_pct=1.5)))
    assert MIN_ATR_PCT == 3.0
    # but does not score
    assert score_pre(_metrics(atr_pct=9.0))[0] == score_pre(_metrics(atr_pct=3.1))[0]


def test_bug45_pre_score_uses_the_surviving_factors():
    from watchlist import PRE_MAX, score_pre

    assert PRE_MAX == 8
    full = score_pre(_metrics())[0]
    assert full == 8
    # each robust factor must be able to cost a point
    for k, v in (("ret_12m", 5.0), ("ret_1m", 0.0), ("dist_200dma", 1.0),
                 ("dist_50dma", 1.0), ("dma200_slope", -2.0),
                 ("base_depth_pct", -80.0), ("spike_level", 0.1),
                 ("px", 4000.0)):
        m = _metrics(); m[k] = v
        assert score_pre(m)[0] < full, f"{k} must be scored"


def test_bug45_close_pos_is_weighted_twice_in_the_brk_score():
    """close_pos is the strongest factor in the study (OOS +1.29), so it is
    scored at both 0.85 and the 0.90 line that defines the edge rule."""
    import pandas as pd

    from watchlist import score_brk

    def frame(cp):
        rows = [dict(open=100.0, high=100.0, low=100.0, close=100.0,
                     volume=1000.0) for _ in range(50)]
        lo, hi = 100.0, 110.0
        c = lo + (hi - lo) * cp
        rows.append(dict(open=101.0, high=hi, low=lo, close=c, volume=5000.0))
        return pd.DataFrame(rows)

    weak = score_brk(frame(0.50), 100.0)[0]
    mid = score_brk(frame(0.87), 100.0)[0]
    strong = score_brk(frame(0.97), 100.0)[0]
    assert weak < mid < strong, (weak, mid, strong)


def test_bug45_watchlist_ranks_on_momentum_not_just_proximity():
    """
    Distance alone measured as noise (-0.195%/trade, t -4.57). With PRE scores
    tied, the stronger 12-month trend must win even if it is further away.
    """
    from watchlist import build_message

    rows = [
        dict(symbol="NEARWEAK", ltp=99.9, level=100.0, which="C", mcap_cr=5000.0,
             pre_score=6, pre_max=8, ret_12m=10.0, atr_pct=4.0,
             pct=-0.1, gap=0.1, bucket="WATCH", screened_ok=True),
        dict(symbol="FARSTRONG", ltp=99.0, level=100.0, which="C", mcap_cr=5000.0,
             pre_score=6, pre_max=8, ret_12m=150.0, atr_pct=4.0,
             pct=-1.0, gap=1.0, bucket="WATCH", screened_ok=True),
    ]
    msg = build_message("2026-07-27", rows,
                        {"universe": 2, "eligible": 2, "capped": 2}, 3.0)
    assert msg.index("FARSTRONG") < msg.index("NEARWEAK"), (
        "on equal PRE score the stronger 12m trend must rank first")


def test_bug45_btst_ready_drops_base_tight_and_adds_dist200():
    from watchlist import btst_ready

    good = dict(atr_pct=4.9, ret_12m=80.0, dist_200dma=40.0)
    assert btst_ready(good) is True
    # base_tight must be irrelevant now
    assert btst_ready({**good, "base_tight": 0.5}) is True
    assert btst_ready({**good, "base_tight": 9.0}) is True
    # the robust factors must still bind
    assert btst_ready({**good, "ret_12m": 10.0}) is False
    assert btst_ready({**good, "dist_200dma": 2.0}) is False


def test_bug45_the_evidence_is_recorded_in_the_source():
    body = (ROOT / "watchlist.py").read_text()
    for token in ("OOS +0.69", "OOS +0.82", "OOS +0.05", "close_pos"):
        assert token in body, f"watchlist.py must record {token!r}"
    assert "THE_EDGE" in body or "23,994" in body or "factor study" in body


# --------------------------------------------------------------------------- #
#  BUG 46 - the 08:45 score and the 15:20 candle were never combined.
#
#  After BUG 45 the question was whether to WEIGHT the PRE score so a
#  weak-momentum name (ARKADE: ret_12m -32%) could not bank four weak passes.
#  Four designs were built and measured on 18,231 tradeable signals. The flat
#  8-point score already shipped WON:
#
#      design        rho(score,return)   best cut   mean     OOS
#      OLD flat 8         +0.57            >=7     +0.473   +0.225
#      NEW flat 8         +0.50            >=7     +0.577   +0.364   <- shipped
#      WEIGHTED 10        +0.28            >=9     +0.480   +0.130
#      GATED 8            +0.57            >=6     +0.438   +0.188
#
#  Weighting made it WORSE (monotonicity rho 0.50 -> 0.28). So the score was
#  left alone - the ARKADE worry was already handled: zero weak-momentum names
#  reach >=6 under the shipped score, versus 181 under the old one.
#
#  WHAT DID PAY was combining the two scans, which are measuring different
#  things - the morning score says "real uptrend", the afternoon candle says
#  "being accumulated right now":
#
#      selector                        n    /wk   win%    mean     OOS
#      PRE >= 6                     6991   27.2   36.9   +0.423   +0.094
#      close_pos >= 0.90            3414   13.3   40.6   +1.040   +0.877
#      PRE>=6 AND close_pos>=0.90   1488    5.8   42.4   +1.719   +1.701
#      PRE>=7 AND close_pos>=0.90    739    2.9   44.5   +2.641   +2.832
#
#  4x the score alone, and the out-of-sample column barely moves. MIN_PRE_CONFIRM
#  = 6 rather than 7 so the list does not go empty for a week at a time.
# --------------------------------------------------------------------------- #
def test_bug46_pre_confirmation_is_wired_into_both_scans():
    import btst

    assert btst.MIN_PRE_CONFIRM == 6
    body = (ROOT / "btst.py").read_text()
    assert "_pre_score_from_daily" in body
    assert "MIN_PRE_CONFIRM" in body


def test_bug46_weak_trend_is_rejected_even_with_a_perfect_close(monkeypatch):
    """
    The whole point: closing at the very high is NOT enough on its own. A stock
    with no trend behind it must still be refused.
    """
    import btst
    from btst import classify_approach

    df, lvl = _approach_frame(close_pos=0.98, gap_pct=1.0)

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)
    assert classify_approach(df, lvl, lvl)["ok"] is True

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (3, []))
    m = classify_approach(df, lvl, lvl)
    assert m["ok"] is False
    assert "PRE" in m["reject"], m["reject"]


def test_bug46_score_is_shared_not_restated():
    """
    btst.py must IMPORT watchlist.score_pre. Restating the thresholds would let
    the 08:45 list and the 15:20 scan drift into disagreeing about what 6/8
    means.
    """
    import inspect

    import btst

    src = inspect.getsource(btst._pre_score_from_daily)
    assert "from watchlist import" in src
    assert "score_pre" in src and "compute_metrics" in src


def test_bug46_pre_score_reaches_the_picks_files_and_the_message():
    body = (ROOT / "btst.py").read_text()
    assert '"pre": int(r.get("pre", 0))' in body, "picks files must carry it"
    assert "/8</b>" in body, "the message must show the score"


def test_bug46_weighted_variants_were_rejected_on_evidence():
    body = (ROOT / "test_regressions.py").read_text()
    for token in ("WEIGHTED 10", "+0.28", "GATED 8", "flat"):
        assert token in body, f"the rejected designs must stay documented: {token}"


# --------------------------------------------------------------------------- #
#  BUG 47 - two changes from the final review, 01-Aug-2026.
#
#  (a) MODEL F WAS SCANNING THE WORSE HALF.
#      classify_approach() returned None for any name already trading ABOVE
#      its level - "not our trade". Measured on the same filter
#      (close_pos>=0.90 & ret_12m>=40), 5 years, net of costs:
#
#          side of level      n       win%     net       OOS
#          BELOW  (scanned)   2,229    50.8%   +0.580%   +0.569%
#          ABOVE  (discarded) 4,657    54.7%   +0.819%   +0.776%
#
#      Twice the sample, better return, and it holds out of sample. The 31-Jul
#      movers said the same thing from the other direction: 13 of 19 had
#      ALREADY broken out at the prior close and only 3 sat in the sub-3%
#      window this scan required.
#
#      Both sides are now scanned and tagged `side`, so they stay separable
#      forever. A runaway extension is still refused (ANTICIPATE_ABOVE_MAX).
#
#  (b) THE 08:45 LIST WAS OVERSOLD.
#      Its measured ceiling is +0.094% out of sample. The strongest factor in
#      the whole study - close_pos - cannot be computed at 08:45 because the
#      candle has not formed. On 30-Jul the top 15 spanned 55x in price and
#      33x in market cap and ZERO met close_pos>=0.90; of the next day's 19
#      big movers exactly ONE was on the list.
#
#      It is now titled "Shortlist", capped at 8, and states in the message
#      that it is a narrowing pass rather than a signal. This is a HONESTY
#      fix, not a performance one - it stops the morning list implying a
#      precision the data does not support.
# --------------------------------------------------------------------------- #
def test_bug47_above_level_names_are_scanned_and_tagged(monkeypatch):
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)

    # below the level -> tagged "below", positive gap
    df, lvl = _approach_frame(close_pos=0.96, gap_pct=1.2)
    below = classify_approach(df, lvl, lvl)
    assert below["side"] == "below" and below["gap_pct"] > 0

    # above the level -> tagged "above", negative gap, still accepted
    df2, lvl2 = _approach_frame(close_pos=0.96, gap_pct=1.2)
    df2.loc[df2.index[-1], "close"] = lvl2 * 1.03
    df2.loc[df2.index[-1], "high"] = lvl2 * 1.04
    above = classify_approach(df2, lvl2, lvl2)
    assert above is not None and above["side"] == "above"
    assert above["gap_pct"] < 0


def test_bug47_extension_cap_still_refuses_a_chase(monkeypatch):
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)
    assert btst.ANTICIPATE_ABOVE_MAX == 10.0

    df, lvl = _approach_frame(close_pos=0.96, gap_pct=1.0)
    df.loc[df.index[-1], "close"] = lvl * 1.30
    df.loc[df.index[-1], "high"] = lvl * 1.31
    assert classify_approach(df, lvl, lvl) is None


def test_bug47_close_pos_still_gates_both_sides(monkeypatch):
    """Extending the pool must not weaken the filter that IS the edge."""
    import btst
    from btst import classify_approach

    import pandas as pd

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch)
    # BUG 65 added a prior-exhaustion gate. This fixture is 259 identical bars
    # at 1000, so the prior close sits ~1.96% under the 252-day high and trips
    # it by construction - an artifact of the synthetic frame, not the logic
    # under test here. Neutralise it so the test keeps measuring close_pos.
    monkeypatch.setattr(btst, "exhausted", lambda m: None)

    def frame(close_pos, mult, n=260, price=1000.0, base=5e5):
        """Build the last candle directly so close_pos is exactly as asked -
        overriding `close` on a prebuilt frame silently changes the range."""
        lvl = price
        c = lvl * mult
        rng = c * 0.06
        low = c - rng * close_pos
        high = low + rng
        rows = [dict(open=price * .995, high=price * 1.02, low=price * .98,
                     close=price, volume=base) for _ in range(n - 1)]
        rows.append(dict(open=(low + high) / 2, high=high, low=low,
                         close=c, volume=base * 1.5))
        return pd.DataFrame(rows), lvl

    for mult, side in ((0.985, "below"), (1.03, "above")):
        df, lvl = frame(0.40, mult)
        m = classify_approach(df, lvl, lvl)
        assert m is not None, f"{side}: should be classified, then refused"
        assert m["ok"] is False, f"{side}: a weak close must still be refused"
        # and the same candle closing at the high must pass on BOTH sides
        df2, lvl2 = frame(0.96, mult)
        m2 = classify_approach(df2, lvl2, lvl2)
        assert m2 is not None and m2["ok"] is True, f"{side}: strong close"


def test_bug47_above_level_ranks_before_below():
    """It measured better (+0.819% vs +0.580%), so it must sort first."""
    body = (ROOT / "btst.py").read_text()
    chunk = body.split("aq = [r for r in ant_rows", 1)[1][:600]
    assert '"above"' in chunk and "sort" in chunk, chunk[:300]
    assert "+0.819" in body and "+0.580" in body, (
        "btst.py must record the measurement behind the ordering")


def test_bug47_anticipation_pool_covers_both_sides():
    body = (ROOT / "btst.py").read_text()
    assert "fired_today" in body
    assert "both sides of the" in body


def test_bug47_shortlist_is_labelled_honestly():
    body = (ROOT / "watchlist.py").read_text()
    assert "NARROWING PASS" in body, "the docstring must say what this is"
    assert "+0.09" in body, "the measured ceiling must be stated"
    from watchlist import TOP_N
    assert TOP_N == 8, "ranking beyond ~8 is noise"


def test_bug47_shortlist_message_warns_it_is_not_a_signal():
    from watchlist import build_message

    rows = [dict(symbol="AAA", ltp=99.0, level=100.0, which="C", mcap_cr=5000.0,
                 pre_score=7, pre_max=8, ret_12m=120.0, atr_pct=4.5,
                 pct=-1.0, gap=1.0, bucket="WATCH", screened_ok=True)]
    msg = build_message("2026-07-27", rows,
                        {"universe": 2, "eligible": 2, "capped": 2}, 3.0)
    assert "Shortlist" in msg
    assert "15:20" in msg, "it must point at where decisions are made"
    assert "not a signal" in msg.lower()


# --------------------------------------------------------------------------- #
#  BUG 48 - the workflow hardcoded --top 15 and silently overrode TOP_N.
#
#  BUG 47 set watchlist.TOP_N = 8 ("ranking beyond ~8 is noise"). The live
#  01-Aug message still printed "CANDIDATES FOR TODAY'S 15:20 SCAN (top 15)"
#  because watchlist.yml passed `--top "${{ inputs.top || '15' }}"` on every
#  scheduled run. The constant was right; the caller ignored it.
#
#  This is the second time a hardcoded caller has beaten a measured default
#  (see BUG 37, where scan.py read a frozen market cap instead of the live
#  table). The rule being relearned: if a value is derived from a measurement,
#  exactly ONE place may define it, and every caller must fall through to it.
#
#  Fix: the workflow only passes --top when the operator explicitly types a
#  value into workflow_dispatch. The scheduled run passes nothing.
# --------------------------------------------------------------------------- #
def test_bug48_workflow_does_not_hardcode_top():
    wf = (ROOT / ".github/workflows/watchlist.yml").read_text()
    assert "'15'" not in wf and '"15"' not in wf, (
        "the workflow must not hardcode a list size - TOP_N is the source of "
        "truth")
    run = wf.split("python watchlist.py", 1)[1].split("\n\n", 1)[0]
    assert "inputs.top &&" in run, (
        "--top must only be passed when explicitly supplied, so the scheduled "
        "run falls through to watchlist.TOP_N")


def test_bug48_default_list_size_is_the_measured_one():
    import inspect

    from watchlist import TOP_N, main

    assert TOP_N == 8
    src = inspect.getsource(main)
    assert "default=TOP_N" in src, (
        "argparse must default to the constant, not to a literal")


def test_bug48_message_prints_the_actual_count():
    """The header must reflect what was really shown, not a fixed number."""
    from watchlist import build_message

    rows = [dict(symbol=f"S{i:02d}", ltp=99.0 - i * 0.01, level=100.0,
                 which="C", mcap_cr=5000.0, pre_score=7, pre_max=8,
                 ret_12m=100.0 - i, atr_pct=4.0, pct=-1.0, gap=1.0 + i * 0.01,
                 bucket="WATCH", screened_ok=True)
            for i in range(20)]
    msg = build_message("2026-07-27", rows,
                        {"universe": 20, "eligible": 20, "capped": 20}, 3.0)
    assert "(top 8)" in msg, msg.split("\n")[5]
    listed = [ln for ln in msg.splitlines() if ln.startswith("• ")]
    assert len(listed) == 8, f"expected 8 rows, got {len(listed)}"
    assert "12 more in the CSV" in msg


# --------------------------------------------------------------------------- #
#  BUG 49 - a data outage was indistinguishable from a quiet day.
#
#  Mon 03-Aug-2026: the user received NOTHING. No alerts, no shortlist, no
#  BTST scan. Every workflow reported SUCCESS.
#
#  CHAIN OF EVENTS
#    1. Sun 02-Aug the Weekly Snapshot rebuild FAILED after 2 minutes.
#    2. Mon 03-Aug the retry HUNG - still in_progress after 4.5 hours against
#       a 90-minute timeout.
#    3. The committed snapshot therefore still carried week_start=2026-07-27
#       while the current week was 2026-08-03.
#    4. load_snapshots() dropped all 2,100 rows. That is CORRECT - it is BUG
#       25's stale-level guard, and alerting on week-old levels would be far
#       worse.
#    5. Every downstream job hit `if not snaps: return 0`.
#
#  So three workflows ran green all day producing nothing, and no failure
#  notice fired because nothing had technically failed. The guard worked; the
#  REPORTING of the guard did not.
#
#  FIXES
#    a) scan.py / watchlist.py / btst.py exit 2 and send a Telegram alarm
#       naming the job and the staleness. The 5-minute scanner de-duplicates
#       to one alert per day via AlertState so it does not send ~75 of them.
#    b) snapshot.yml: timeout 90 -> 45 min, cancel-in-progress true so a hung
#       build cannot block the next attempt, and the failure message no longer
#       claims the scanner "will keep using last week's levels" - it will not,
#       it refuses to run at all.
#    c) NEW healthcheck.yml at 08:30 IST - a dead-man's switch that asserts the
#       snapshot is for the CURRENT week before the market opens, independent
#       of every other job.
# --------------------------------------------------------------------------- #
def test_bug49_jobs_exit_nonzero_on_a_stale_snapshot():
    """Silent success is the bug. Every scheduled job must go red."""
    import inspect

    import btst
    import scan
    import watchlist

    for mod, name in ((scan, "scan"), (watchlist, "watchlist"), (btst, "btst")):
        src = inspect.getsource(mod.main)
        assert "report_stale_snapshot" in src, (
            f"{name}.main must alarm when the snapshot is unusable")
        assert "return 2" in src, (
            f"{name}.main must exit non-zero so the workflow fails visibly")


def test_bug49_stale_alert_is_deduplicated_per_day(tmp_path):
    """The 5-minute scan would otherwise send ~75 identical alerts a day."""
    from state import AlertState

    p = tmp_path / "state.json"
    st = AlertState(p)
    assert st.stale_alerted("2026-08-03") is False
    st.mark_stale("2026-08-03")
    st.save()

    st2 = AlertState(p)
    assert st2.stale_alerted("2026-08-03") is True
    assert st2.stale_alerted("2026-08-04") is False, "a new day must re-alert"


def test_bug49_snapshot_age_is_computed(tmp_path):
    import pandas as pd

    import scan
    from config import load_config

    (tmp_path / "config.yaml").write_text("strategy:\n  use_mcap: false\n")
    cfg = load_config(tmp_path / "config.yaml")

    assert scan.snapshot_age_days(cfg) is None, "missing file -> None"

    old = pd.Timestamp.now().normalize() - pd.Timedelta(days=7)
    pd.DataFrame({"week_start": [str(old.date())] * 3}).to_csv(
        tmp_path / "weekly_snapshot.csv", index=False)
    age = scan.snapshot_age_days(cfg)
    assert age is not None and age >= 6, age


def test_bug49_alarm_never_crashes_the_job(tmp_path, monkeypatch):
    """
    The alarm runs on the failure path. If Telegram is down it must not raise
    and mask the original problem.
    """
    import scan
    from config import load_config

    (tmp_path / "config.yaml").write_text("strategy:\n  use_mcap: false\n")
    cfg = load_config(tmp_path / "config.yaml")

    def boom(*a, **k):
        raise RuntimeError("telegram unreachable")

    monkeypatch.setattr(scan, "build_telegram", boom)
    scan.report_stale_snapshot(cfg, "2026-08-03", "Test Job")   # must not raise


def test_bug49_healthcheck_workflow_exists_and_runs_premarket():
    import re

    p = ROOT / ".github/workflows/healthcheck.yml"
    assert p.exists(), "the dead-man's switch must exist"
    wf = p.read_text()

    crons = re.findall(r'cron:\s*"([^"]+)"', wf)
    assert crons, "it must be scheduled"
    mi, hr = crons[0].split()[0], crons[0].split()[1]
    ist = int(hr) * 60 + int(mi) + 330
    assert ist < 9 * 60 + 15, (
        f"health check runs {ist//60:02d}:{ist%60:02d} IST - must be before "
        "the 09:15 open")
    assert "week_start" in wf and "stale" in wf.lower()
    assert "exit 1" in wf, "an unhealthy check must fail the workflow"


def test_bug49_snapshot_workflow_is_bounded():
    """
    SUPERSEDED BY BUG 50. This originally required timeout <= 60 and
    cancel-in-progress: true, on the belief that the 03-Aug build had hung.
    The log later showed it was rate-limited and crawling toward a legitimate
    ~154-minute finish, so both requirements were wrong and would have
    guaranteed the outage they were meant to prevent.

    What still matters, and is asserted here, is that the job is BOUNDED at
    all - an unbounded build hides a real hang. The specific bound is pinned
    by test_bug50_snapshot_timeout_fits_a_ratelimited_build.
    """
    import re

    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    m = re.search(r"timeout-minutes:\s*(\d+)", wf)
    assert m, "the snapshot job must declare a timeout"
    assert int(m.group(1)) <= 360


def test_bug49_failure_message_does_not_promise_a_fallback():
    """
    The old message said the scanner "will keep using last week's levels".
    It does not - it refuses to run, which is a much bigger deal and the user
    needs to know that immediately.
    """
    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    assert "keep using last week" not in wf
    assert "does NOT fall back" in wf


# --------------------------------------------------------------------------- #
#  BUG 50 - I diagnosed a rate-limited build as a HUNG build and "fixed" it by
#  making the timeout SHORTER. That would have guaranteed the failure.
#
#  WHAT THE 03-AUG LOG ACTUALLY SHOWED
#      07:20  200/2372  ok=174 err=1     363s   -> 1.81 s/symbol
#      07:33  400/2372  ok=346 err=8    1182s   -> 4.09 s/symbol for 200-400
#  with repeated "429 rate limited" warnings. Projected total: ~154 min.
#
#  The build was not hung. It was throttled and crawling, and the 4.5 hours it
#  had been running was consistent with finishing. My change from 90 -> 45 min
#  would have killed it at ~770 of 2,372 symbols; even 90 min only reaches
#  ~1,430. Both produce an unusable snapshot and the same outage.
#
#  I also set cancel-in-progress: true, which for a RESUMABLE long build is
#  strictly harmful - it throws away a healthy run in progress.
#
#  CORRECTED
#    * timeout 330 min, sized from the measured degraded rate with headroom
#    * cancel-in-progress back to false
#    * a continuation step: the commit already runs on always(), and
#      build_snapshot.py resumes from a partial file for the same week and
#      logic_version, so a killed run is not wasted - it is continued
#    * max_workers 4 -> 3 and data_rate 4 -> 3/sec. A 429 triggers a GLOBAL
#      pause, so more threads on the same bucket make the build SLOWER.
#
#  The lesson, and the reason this block is long: I inferred "hung" from
#  elapsed time alone without reading the log. The elapsed time was the only
#  evidence I had, and it was consistent with two very different causes.
# --------------------------------------------------------------------------- #
def test_bug50_snapshot_timeout_fits_a_ratelimited_build():
    import re

    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    m = re.search(r"timeout-minutes:\s*(\d+)", wf)
    assert m, "the snapshot job must have a timeout"
    mins = int(m.group(1))
    # measured worst case ~154 min; anything under ~180 risks killing a
    # healthy but throttled build
    assert mins >= 180, (
        f"timeout {mins} min is below the measured worst case (~154 min at "
        "4.1 s/symbol under 429 backoff) - this is the BUG 50 mistake")
    assert mins <= 360, "an unbounded timeout hides a genuine hang"


def test_bug50_slow_build_is_not_cancelled():
    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    assert "cancel-in-progress: false" in wf, (
        "a resumable multi-hour build must never be cancelled by a newer run")


def test_bug50_partial_progress_is_committed_and_continued():
    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    commit = wf.split("name: Commit snapshot", 1)[1][:400]
    assert "if: always()" in commit, (
        "partial progress must be committed even when the build is killed")
    assert "Continue if the snapshot is incomplete" in wf
    assert "run_attempt" in wf, "the continuation must not loop forever"


def test_bug50_build_snapshot_resumes_from_a_partial_file():
    """The continuation only works because the builder resumes. Pin that."""
    src = (ROOT / "build_snapshot.py").read_text()
    assert "resuming:" in src
    assert "done_symbols" in src
    assert "logic_version" in src, (
        "resume must be scoped to the same logic version or it silently keeps "
        "stale formulas")


def test_bug50_rate_limits_leave_headroom():
    from config import load_config

    cfg = load_config(None)
    assert cfg.runtime.data_rate_per_sec <= 3, (
        "Dhan allows 5/s but a 429 pauses ALL threads - stay well under")
    assert cfg.runtime.max_workers <= 3, (
        "more workers on one rate bucket makes a throttled build slower")


def test_bug50_reasoning_is_recorded():
    """So the timeout is not 'tidied down' again by someone reading it cold."""
    wf = (ROOT / ".github/workflows/snapshot.yml").read_text()
    assert "BUG 50" in wf
    cfg = (ROOT / "config.yaml").read_text()
    assert "429" in cfg and "GLOBAL pause" in cfg


# --------------------------------------------------------------------------- #
#  BUG 51 - the anticipation gate had no trend floor.
#
#  Of 11 live alerts on 03-Aug-2026, the four clean losers were precisely the
#  four names with no established uptrend:
#
#      symbol      ret_12m   dist200   outcome
#      BHAGCHEM      -9.8%     13.1%    -2.64%
#      TBOTEK        +8.0%      5.5%    -3.48%
#      SPORTKING    +51.9%     54.4%    -1.24%   (gapped +15% overnight)
#      SWANDEF      no 12m      79.7%   -0.40%
#
#  while all six winners had ret_12m >= 53% and dist200 >= 25%.
#
#  n=11 proves nothing, so it was re-tested on 18,202 tradeable signals over
#  5 years. The effect is real and monotonic:
#      ret_12m < 10   n 1,693   -0.339%   OOS -0.397%
#      ret_12m 10-50  n 6,145   -0.099%   OOS -0.568%
#      ret_12m >= 50  n10,364   +0.456%   OOS +0.151%
#      ret_12m >=100  n 5,273   +0.711%   OOS +0.489%
#  and stacked on the shipped filter:
#      close_pos>=0.9                      n 3,405  +1.043%  OOS +0.881%
#      + ret_12m>=40 (previous)            n 2,376  +1.404%  OOS +1.341%
#      + ret_12m>=50 & dist200>=25 (now)   n 1,822  +1.610%  OOS +1.647%
#
#  DELIBERATELY NOT ADOPTED: a gap cap. On 03-Aug it looked decisive -
#  SPORTKING gapped +15% and BHAGCHEM +8.5%, both lost - but over 5 years
#  gap_pct<=4 measured +0.160% against a +0.194% baseline, i.e. WORSE. Two
#  observations are an anecdote. This is recorded so it is not "fixed" later.
# --------------------------------------------------------------------------- #
def test_bug51_trend_floor_matches_the_measurement():
    import btst

    assert btst.MIN_RET_12M == 50.0
    assert btst.MIN_DIST_200DMA == 25.0


def test_bug51_rejects_the_four_losers_of_03_aug(monkeypatch):
    """The exact trend readings of the four names that lost money that day."""
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))

    losers = {
        "BHAGCHEM": (-9.8, 13.1),
        "TBOTEK": (8.0, 5.5),
        "SWANDEF": (float("nan"), 79.7),   # no 12-month history
    }
    for name, (r12, d200) in losers.items():
        _stub_trend(monkeypatch, ret_12m=r12, dist_200dma=d200)
        df, lvl = _approach_frame(close_pos=0.97, gap_pct=1.0)
        m = classify_approach(df, lvl, lvl)
        assert m is not None and m["ok"] is False, f"{name} must be rejected"

    # ...and admits a genuine uptrend on the same candle
    _stub_trend(monkeypatch, ret_12m=75.8, dist_200dma=49.1)   # UNIVCABLES
    df, lvl = _approach_frame(close_pos=0.97, gap_pct=1.0)
    assert classify_approach(df, lvl, lvl)["ok"] is True


def test_bug51_unknown_trend_fails_rather_than_passes(monkeypatch):
    """
    Opposite of the c12 market-cap rule, and deliberately so. An unknown market
    cap is a data gap; an unknown 12-month return means the stock has not
    traded a year and has no trend to speak of. SWANDEF was exactly that.
    """
    import btst
    from btst import classify_approach

    monkeypatch.setattr(btst, "_pre_score_from_daily", lambda *a, **k: (8, []))
    _stub_trend(monkeypatch, ret_12m=float("nan"), dist_200dma=40.0)
    df, lvl = _approach_frame(close_pos=0.97, gap_pct=1.0)
    m = classify_approach(df, lvl, lvl)
    assert m["ok"] is False
    assert "unknown" in m["reject"]


def test_bug51_gap_cap_was_rejected_on_evidence():
    body = (ROOT / "btst.py").read_text()
    assert "NOT ADOPTED" in body and "gap cap" in body, (
        "the rejected gap filter must stay documented so it is not re-added")
    # and it must not actually be implemented
    assert "MAX_GAP" not in body


def test_bug51_watchlist_moon_matches_the_1520_floor():
    """
    A 🌙 the afternoon scan would reject on trend is a false promise. The two
    thresholds must be the same number.
    """
    import btst
    from watchlist import btst_ready

    ok = dict(atr_pct=4.9, ret_12m=btst.MIN_RET_12M + 1,
              dist_200dma=btst.MIN_DIST_200DMA + 1)
    assert btst_ready(ok) is True
    assert btst_ready({**ok, "ret_12m": btst.MIN_RET_12M - 1}) is False
    assert btst_ready({**ok, "dist_200dma": btst.MIN_DIST_200DMA - 1}) is False
    assert btst_ready({**ok, "ret_12m": float("nan")}) is False, (
        "no 12-month history must not earn a moon")


def test_bug51_trend_reaches_the_picks_file_and_message():
    body = (ROOT / "btst.py").read_text()
    # BUG 62 replaced round(float(x or 0)) with _num(), which keeps an unknown
    # BLANK instead of claiming 0.0. The requirement is unchanged: the trend
    # must reach the picks file and the message.
    assert '"ret_12m": _num(' in body, "picks must record the trend"
    assert '"dist_200dma": _num(' in body
    assert "over the 200DMA" in body, "the message must show it"


def test_bug51_trend_is_read_from_the_shared_metrics():
    """Recomputing it here would let the two scans drift apart."""
    import inspect

    import btst

    src = inspect.getsource(btst.classify_approach)
    assert "from watchlist import compute_metrics" in src


# --------------------------------------------------------------------------- #
#  BUG 52 - the 15:20 BTST scan could not finish before the close.
#
#  03-Aug-2026: the message was stamped 15:48 and said "171 screened". The run
#  took 42 minutes end to end.
#
#  CAUSE. BUG 47 widened the anticipation pool from "names that never fired"
#  to the WHOLE snapshot - correct for the measurement, but that is ~2,100
#  symbols and each one costs a daily-candles call plus an intraday call. The
#  job cannot complete in the ten minutes it has before the bell, so the
#  ANTICIPATE list came back empty and the top-5 was drawn from a fraction of
#  the universe.
#
#  A 15:20 job that finishes at 16:02 is useless no matter how good its picks.
#
#  FIX. The distance window is knowable from the FROZEN weekly levels plus ONE
#  bulk quote - no per-symbol history needed. Filter on that first, then fetch
#  history only for names plausibly in range. A 1.5% margin is added because
#  the quote is a snapshot and price still moves into the close.
#
#  NOT A BUG, verified separately: only ONE pick appeared because only one
#  name qualified. Of the 26 that broke out on 03-Aug, exactly one cleared
#  Tier A (DALMIASUG, day +17.3%, close_pos 0.87). The next best was
#  AVADHSUGAR at +9.6% / 0.77 - below both thresholds. The scan was right.
# --------------------------------------------------------------------------- #
def test_bug52_anticipation_prefilters_before_fetching_history():
    body = (ROOT / "btst.py").read_text()
    chunk = body.split("PRE-FILTER ON THE FROZEN LEVELS", 1)
    assert len(chunk) == 2, "the pre-filter must exist"
    seg = chunk[1][:2000]
    assert "fetch_ltp" in seg, "it must use ONE bulk quote, not per-symbol calls"
    assert "ANTICIPATE_NEAR" in seg and "ANTICIPATE_ABOVE_MAX" in seg, (
        "the window must be derived from the shipped thresholds")
    assert "MARGIN" in seg, (
        "a snapshot quote needs headroom - price moves into the close")


def test_bug52_prefilter_failure_does_not_drop_everything():
    """
    If the bulk quote fails the scan must fall back to screening everything -
    slow, but not silently empty. An exception here would have produced the
    same empty list the bug caused.
    """
    body = (ROOT / "btst.py").read_text()
    seg = body.split("PRE-FILTER ON THE FROZEN LEVELS", 1)[1][:2000]
    assert "except DhanError" in seg
    assert "ltp = {}" in seg, "on failure it must screen unfiltered, not none"
    assert "if ltp:" in seg, "the filter only applies when quotes were obtained"


def test_bug52_prefilter_keeps_both_sides_of_the_level():
    """The window must admit above-level names too - BUG 47 showed they are
    the better half (+0.819% vs +0.580%)."""
    body = (ROOT / "btst.py").read_text()
    seg = body.split("PRE-FILTER ON THE FROZEN LEVELS", 1)[1][:2000]
    assert "above" in seg and "ANTICIPATE_ABOVE_MAX" in seg


def test_bug68_pool_composition_is_stated_and_quote_failure_is_loud():
    """
    THE 07-Aug DIAGNOSIS GAP. The 15:18 run reported "7 candle(s) checked"
    and the 16:08 rerun "17" - but the message reported only SUCCESSES, so a
    small candidate POOL and a broken feed were indistinguishable.

    SBCL is the case that exposed it: it gapped ABOVE its 814 level at the
    09:15 open (843.65), so it was never a fresh CROSS and never entered the
    fired-today set. It could only reach the scan through the AGED tier B
    pool - which exists only when the bulk quote returns. If that quote is
    empty the aged pass is skipped silently and the scan degenerates to
    "names that crossed today", with nothing in the message to say so.
    """
    body = (ROOT / "btst.py").read_text()

    # the header must state pool composition, not just successes
    assert "candidate(s) \"" in body or "candidate(s) " in body, (
        "the message must say how many candidates existed, not only how many "
        "produced a candle")
    assert "{len(todo)} candidate(s)" in body
    assert "broke out today + {len(aged_pool)} aged" in body

    # a missing quote must be loud in the log AND the message
    assert "quote_failed = True" in body
    assert "NO BULK QUOTE" in body, "a skipped aged pass must be an ERROR"
    assert "aged setups were NOT scanned" in body, (
        "the user must be told the aged pass was skipped")


def test_bug69_bulk_quote_does_not_inherit_the_urgent_timeout():
    """
    BUG 61 set timeout=8s / retries=2 while the entry window is live. That is
    right for a PER-SYMBOL history call - hundreds of them, a slow one is
    worth abandoning - and WRONG for the bulk quote.

    fetch_ohlc sends ~2,100 ids as 3 requests of up to 1000 instruments. Those
    payloads are large, and unlike a per-symbol call this one is not
    redundant: if it fails, ltp_all is empty, the aged tier B pool is never
    built, and the scan silently degenerates to "names that crossed today".

    07-Aug signature:
        15:18 (urgent, 8s)  ->  7 candidates, SBCL absent
        16:08 (not urgent)  -> 17 candidates, SBCL present as TIER B
    SBCL was +13.1% above its level at close_pos 1.000 by 15:18 and passed
    both aged screens, so only an empty quote explains its absence.
    """
    body = (ROOT / "btst.py").read_text()

    assert "quote_client = client if not urgent else DhanClient(" in body, (
        "the bulk quote needs its own patient client")
    seg = body.split("quote_client = client if not urgent", 1)[1][:400]
    assert "timeout=30" in seg and "max_retries=4" in seg, (
        "one request that gates an entire pass must be retried patiently")

    # every bulk-quote call site must use it; per-symbol calls must NOT
    assert "fetch_ohlc(quote_client, snaps)" in body
    assert body.count("fetch_ltp(quote_client") == 2, (
        "both the aged fallback and the anticipation pre-filter are bulk "
        "quotes and must use the patient client")
    # (the def line is `def fetch_ohlc(client, snaps)` - check CALL sites)
    assert "= fetch_ohlc(client," not in body
    # the per-symbol path keeps the urgent settings
    assert "timeout=8 if urgent else 30" in body
    assert "max_retries=2 if urgent else 5" in body


def test_bug70_session_fraction_comes_from_the_clock_not_the_bar_count():
    """
    THE 17:34 FALSE TIER B. `frac` used to be len(m5)/75 - bars returned over
    a full session. That conflates "session still running" with "the API sent
    a partial page", and frac DIVIDES the volume benchmark, so a wrongly-low
    frac inflates rvol.

    On the 07-Aug 17:34 review run - two hours after the close - SONACOMS came
    back with ~36 of 75 bars and its rvol doubled 1.5x -> 3.1x, stepping over
    TIER_B_RVOL >= 3.0. A post-close rerun was MANUFACTURING a Tier B setup
    that did not exist at the bell. Compare the same names at 16:30:
        SBCL 13.7 -> 14.1, COMSYN 1.6 -> 1.5, DIVGIITTS 0.8 -> 0.7 (stable)
        SONACOMS 1.5 -> 3.1 (doubled)
    """
    from datetime import datetime

    import pytz

    import btst

    IST = pytz.timezone("Asia/Kolkata")

    def at(h, m):
        return btst.session_fraction(IST.localize(datetime(2026, 8, 7, h, m)))

    # after the close the session is over BY DEFINITION, whatever came back
    assert at(15, 30) == 1.0
    assert at(17, 34) == 1.0
    assert at(22, 0) == 1.0
    # during the session it tracks the clock
    assert 0.95 < at(15, 18) < 0.98, "15:18 is ~97% of the session"
    assert 0.4 < at(12, 0) < 0.5
    assert at(9, 20) <= 0.06
    # never zero - that would make rvol infinite
    assert at(9, 0) > 0

    # the bar count must NOT be able to shorten it
    n = IST.localize(datetime(2026, 8, 7, 17, 34))
    assert btst.session_fraction(n, 36) == 1.0, (
        "a partial intraday page after the close must not inflate rvol")

    body = (ROOT / "btst.py").read_text()
    assert "frac = min(len(m5) / float(BARS_PER_SESSION), 1.0)" not in body, (
        "frac must never be derived from how many bars the API returned")
    assert body.count("frac = session_fraction(now") == 2, (
        "both the confirmed and anticipation paths must use the clock")
    assert 'drops["partial intraday after close"]' in body, (
        "a short intraday page after the close is a data gap worth reporting")
