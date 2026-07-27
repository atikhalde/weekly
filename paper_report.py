"""
Post-market paper-trading report -> formatted Excel -> Telegram.

Runs the paper-trading simulation over a recent window and delivers a styled
.xlsx workbook to the chat after the close.

    python paper_report.py                      # this week, snapshot symbols
    python paper_report.py --weeks 4
    python paper_report.py RATNAVEER TMB --weeks 8
    python paper_report.py --no-send --out report.xlsx   # local only

Workbook layout
    Summary   headline P&L, win rate, profit factor, exit breakdown
    Trades    one row per trade: dates, times, entry, qty, SL, levels, P&L, R
    Open      positions still running, marked to the latest close

Rules are paper.py's, unchanged: entry on the live 5m signal, stop at the entry
candle low - 0.02%, exit on the first 5m close below the 9-EMA.
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime, time as dtime, timedelta
from pathlib import Path

import pandas as pd

from config import load_config
from dhan import IST, DhanClient, DhanError, last_n_years
from paper import EMA_LEN, SL_BUFFER, PaperTrade, fetch_5m, simulate
from strategy import build_snapshot, replay_week, week_start_of
from telegram import Telegram

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")
log = logging.getLogger("paper_report")

HEADERS = [
    ("symbol", "Symbol", 12), ("signal_date", "Date", 12),
    ("signal_time", "Signal Bar", 11), ("signal_close", "Signal Close", 12),
    ("entry_time", "Entry Bar", 11), ("entry", "Entry Fill", 11),
    ("slippage_pct", "Slip %", 9), ("qty", "Qty", 7),
    ("invested", "Invested", 12), ("stop", "Stop Loss", 11),
    ("level_26w", "26W Level", 11), ("level_52w", "52W Level", 11),
    ("exit_date", "Exit Date", 12), ("exit_time", "Exit Bar", 10),
    ("exit", "Exit Fill", 11), ("exit_reason", "Why", 8),
    ("exit_note", "Note", 18),
    ("bars_held", "Bars", 7), ("pnl", "P&L (Rs)", 12),
    ("pnl_pct", "P&L %", 9), ("r_multiple", "R", 8),
    ("mfe_pct", "MFE %", 9), ("mae_pct", "MAE %", 9),
    ("rsi", "RSI", 8), ("macd_hist", "MACD H", 10),
    ("trigger", "Trigger", 10), ("week", "Week", 12),
]


# --------------------------------------------------------------------------- #
#  Excel
# --------------------------------------------------------------------------- #
def build_workbook(trades: list[PaperTrade], capital: float, path: Path,
                   window: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    green = Font(color="107C10", bold=True)
    red = Font(color="C42B1C", bold=True)
    title_font = Font(bold=True, size=14)

    closed = [t for t in trades if t.exit_reason in ("SL", "EMA9")]
    openx = [t for t in trades if t.exit_reason == "OPEN"]
    wins = [t for t in trades if t.pnl > 0]
    losses = [t for t in trades if t.pnl <= 0]
    total = sum(t.pnl for t in trades)
    deployed = sum(t.invested for t in trades)
    gross_w = sum(t.pnl for t in wins)
    gross_l = abs(sum(t.pnl for t in losses))

    # ---------------------------------------------------------------- Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Paper Trading Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated {datetime.now(IST):%d-%b-%Y %H:%M} IST"
    ws["A3"] = f"Window: {window}"
    ws["A4"] = (f"Rules: entry on live 5m breakout signal | "
                f"stop = entry candle low -{SL_BUFFER*100:.2f}% | "
                f"exit = first 5m close below {EMA_LEN}-EMA")
    ws["A5"] = f"Capital deployed per stock: Rs {capital:,.0f}"
    for r in ("A2", "A3", "A4", "A5"):
        ws[r].font = Font(size=10, color="555555")

    rows = [
        ("", ""),
        ("Trades", len(trades)),
        ("Closed", len(closed)),
        ("Still open", len(openx)),
        ("", ""),
        ("Net P&L (Rs)", round(total, 2)),
        ("Return on deployed (%)", round(total / deployed * 100, 2) if deployed else 0),
        ("Total deployed (Rs)", round(deployed, 2)),
        ("", ""),
        ("Win rate (%)", round(len(wins) / len(trades) * 100, 1) if trades else 0),
        ("Wins", len(wins)),
        ("Losses", len(losses)),
        ("Avg win (Rs)", round(gross_w / len(wins), 2) if wins else 0),
        ("Avg loss (Rs)", round(-gross_l / len(losses), 2) if losses else 0),
        ("Profit factor", round(gross_w / gross_l, 2) if gross_l else "n/a"),
        ("Expectancy per trade (Rs)", round(total / len(trades), 2) if trades else 0),
        ("Avg R multiple", round(sum(t.r_multiple for t in trades) / len(trades), 2)
         if trades else 0),
        ("Avg bars held", round(sum(t.bars_held for t in trades) / len(trades), 1)
         if trades else 0),
    ]
    r = 7
    for label, val in rows:
        if label:
            ws.cell(r, 1, label).font = Font(bold=True, size=10)
            c = ws.cell(r, 2, val)
            if isinstance(val, (int, float)) and "P&L" in label or "Expectancy" in label:
                c.font = green if (isinstance(val, (int, float)) and val > 0) else (
                    red if isinstance(val, (int, float)) and val < 0 else Font())
                c.number_format = "#,##0.00"
        r += 1

    r += 1
    ws.cell(r, 1, "Exit breakdown").font = Font(bold=True, size=11)
    r += 1
    for h, w in (("Reason", 14), ("Trades", 10), ("P&L (Rs)", 14), ("Avg (Rs)", 14)):
        cell = ws.cell(r, ["Reason", "Trades", "P&L (Rs)", "Avg (Rs)"].index(h) + 1, h)
        cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
    by: dict[str, list[PaperTrade]] = {}
    for t in trades:
        by.setdefault(t.exit_reason, []).append(t)
    for reason, grp in sorted(by.items()):
        r += 1
        pnl = sum(x.pnl for x in grp)
        ws.cell(r, 1, reason).border = border
        ws.cell(r, 2, len(grp)).border = border
        c = ws.cell(r, 3, round(pnl, 2)); c.border = border; c.number_format = "#,##0.00"
        c.font = green if pnl > 0 else red
        c = ws.cell(r, 4, round(pnl / len(grp), 2)); c.border = border
        c.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    # ---------------------------------------------------------- Trades sheets
    def write_sheet(sheet, subset: list[PaperTrade]) -> None:
        for col, (_key, label, width) in enumerate(HEADERS, start=1):
            cell = sheet.cell(1, col, label)
            cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[get_column_letter(col)].width = width
        for i, t in enumerate(subset, start=2):
            for col, (key, _label, _w) in enumerate(HEADERS, start=1):
                val = getattr(t, key)
                if isinstance(val, float):
                    val = round(val, 2)
                cell = sheet.cell(i, col, val)
                cell.border = border
                if key in ("entry", "exit", "stop", "level_26w", "level_52w",
                           "invested", "pnl"):
                    cell.number_format = "#,##0.00"
                if key in ("pnl", "pnl_pct", "r_multiple"):
                    cell.font = green if (isinstance(val, (int, float)) and val > 0) else (
                        red if isinstance(val, (int, float)) and val < 0 else Font())
                if key == "exit_reason":
                    cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        if subset:
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(HEADERS))}{len(subset) + 1}")

    write_sheet(wb.create_sheet("Trades"), sorted(
        trades, key=lambda t: (t.entry_date, t.entry_time)))
    if openx:
        write_sheet(wb.create_sheet("Open"), openx)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def caption(trades: list[PaperTrade], capital: float, window: str) -> str:
    if not trades:
        return (f"📄 <b>Paper Trading Report</b>\n{window}\n\n"
                "No trades in this window.")
    total = sum(t.pnl for t in trades)
    deployed = sum(t.invested for t in trades)
    wins = [t for t in trades if t.pnl > 0]
    openx = [t for t in trades if t.exit_reason == "OPEN"]
    icon = "🟢" if total > 0 else ("🔴" if total < 0 else "⚪")
    lines = [
        "📄 <b>Paper Trading Report</b>",
        f"<i>{window}</i>",
        "",
        f"{icon} <b>Net P&L Rs {total:,.0f}</b> "
        f"({total/deployed*100 if deployed else 0:+.2f}% on deployed)",
        f"Trades <b>{len(trades)}</b> · win rate <b>{len(wins)/len(trades)*100:.0f}%</b>"
        f" · open <b>{len(openx)}</b>",
        f"Rs {capital:,.0f} per stock",
    ]
    best = max(trades, key=lambda t: t.pnl)
    worst = min(trades, key=lambda t: t.pnl)
    if best.pnl > 0:
        lines.append(f"Best  {best.symbol} {best.pnl:+,.0f} ({best.pnl_pct:+.2f}%)")
    if worst.pnl < 0:
        lines.append(f"Worst {worst.symbol} {worst.pnl:+,.0f} ({worst.pnl_pct:+.2f}%)")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("symbols", nargs="*")
    ap.add_argument("--weeks", type=int, default=1,
                    help="weeks of history to replay (default 1 = this week)")
    ap.add_argument("--capital", type=float, default=100000.0)
    ap.add_argument("--out", default=None, help="xlsx path")
    ap.add_argument("--no-send", action="store_true", help="build the file only")
    ap.add_argument("--fill", choices=["signal-close", "next-open"],
                    default="signal-close",
                    help="signal-close (default): fill at the close of the "
                         "signal candle - matches the indicator's BUY arrow. "
                         "next-open: fill at the next bar's open, which is what "
                         "a market order placed on the alert would realistically "
                         "get; use it to size expected slippage.")
    ap.add_argument("--config", default=None)
    args = ap.parse_args()

    cfg = load_config(args.config)
    interval = cfg.runtime.bar_interval_min
    tg = Telegram(cfg.secrets.telegram_bot_token, cfg.secrets.telegram_chat_id,
                  dry_run=cfg.runtime.dry_run)

    # Fail fast on a partial deploy. Paper trading needs the entry candle's low,
    # which strategy.Signal only carries in recent versions. Without this the
    # run would spend minutes fetching data and then die on AttributeError.
    from dataclasses import fields as _dc_fields

    from strategy import Signal as _Signal

    if "bar_low" not in {f.name for f in _dc_fields(_Signal)}:
        log.error("strategy.py is out of date: Signal has no `bar_low` field.")
        log.error("Paper trading needs the entry candle low to place the stop.")
        log.error("Upload the matching strategy.py and re-run.")
        return 2

    symbols = [s.upper() for s in args.symbols]
    if not symbols:
        p = cfg.paths["snapshot"]
        if not p.exists():
            log.error("%s not found - run build_snapshot.py first", p)
            return 1
        symbols = sorted(pd.read_csv(p, dtype=str)["symbol"].str.upper().unique())

    if not cfg.secrets.dhan_access_token:
        log.error("DHAN_ACCESS_TOKEN not set")
        return 2
    client = DhanClient(cfg.secrets.dhan_client_id, cfg.secrets.dhan_access_token,
                        data_rate=cfg.runtime.data_rate_per_sec,
                        quote_rate=cfg.runtime.quote_rate_per_sec)

    log.info("loading instruments ...")
    ins = DhanClient.fetch_instruments(cfg.universe.exchange_segments,
                                       cfg.universe.series,
                                       exclude_etf=cfg.universe.exclude_etf)
    sec_map = {i.symbol.upper(): (i.security_id, i.exchange_segment) for i in ins}

    today = datetime.now(IST).date()
    end_week = week_start_of(today)
    start_week = end_week - pd.Timedelta(weeks=max(args.weeks - 1, 0))
    from_date, to_date = last_n_years(cfg.runtime.history_years)
    window = f"{start_week.date()} to {today}"

    log.info("paper trading %s over %d symbol(s)", window, len(symbols))

    trades: list[PaperTrade] = []
    for n, sym in enumerate(symbols, 1):
        if sym not in sec_map:
            continue
        sid, seg = sec_map[sym]
        try:
            daily = client.daily_candles(sid, seg, from_date, to_date)
        except DhanError:
            continue
        if daily.empty:
            continue

        five = fetch_5m(client, sid, seg,
                        datetime.combine(start_week.date(), dtime(9, 0)).replace(tzinfo=IST),
                        datetime.now(IST), interval)
        if five.empty:
            continue
        five["_day"] = (pd.to_datetime(five["datetime"]).dt.tz_convert(IST)
                        .dt.tz_localize(None).dt.normalize())
        d = daily.copy()
        d["_day"] = pd.to_datetime(d["datetime"]).dt.tz_localize(None).dt.normalize()

        busy_until = None   # per-symbol: no overlapping positions
        wk = start_week
        while wk <= end_week:
            hist = d[d["_day"] < wk]
            snap = build_snapshot(sym, sid, seg, hist, cfg.strategy, wk) \
                if not hist.empty else None
            if snap is not None:
                wbars = five[(five["_day"] >= wk) & (five["_day"] < wk + pd.Timedelta(days=7))]
                if not wbars.empty:
                    for sig in replay_week(snap, cfg.strategy, wbars).signals:
                        # Skip a signal that lands while an earlier trade in
                        # this symbol is still open - real capital cannot hold
                        # two positions in one name.
                        if busy_until is not None and \
                                pd.Timestamp(sig.bar_time) <= busy_until:
                            continue
                        ts = pd.to_datetime(five["datetime"])
                        after = five[ts > pd.Timestamp(sig.bar_time)]
                        before = five[ts <= pd.Timestamp(sig.bar_time)]
                        tr = simulate(sig, after, args.capital, before, args.fill)
                        trades.append(tr)
                        if tr.exit_date:
                            busy_until = pd.Timestamp(
                                f"{tr.exit_date} {tr.exit_time}").tz_localize(IST)
            wk += pd.Timedelta(days=7)

        if n % 100 == 0:
            log.info("  %d/%d scanned, %d trades", n, len(symbols), len(trades))

    out = Path(args.out) if args.out else Path(
        f"paper_report_{today:%Y-%m-%d}.xlsx")
    build_workbook(trades, args.capital, out, window)
    log.info("wrote %s (%d trades, %.1f KB)", out, len(trades),
             out.stat().st_size / 1024)

    if args.no_send:
        return 0
    if tg.send_document(out, caption(trades, args.capital, window)):
        log.info("report sent to Telegram")
        return 0
    log.error("Telegram upload failed")
    return 1


if __name__ == "__main__":
    sys.exit(main())
